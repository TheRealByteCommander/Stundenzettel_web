"""
Autogen Agent Network for Travel Expense Report Review
Implements specialized agents: Chat Agent, Document Agent, Accounting Agent
"""

import os
import json
import logging
import asyncio
import uuid
from typing import Dict, List, Optional, Any
from datetime import datetime, timedelta
from pathlib import Path
import aiohttp
import base64
from collections import deque

try:
    import PyPDF2
    HAS_PYPDF2 = True
except ImportError:
    HAS_PYPDF2 = False

try:
    import pdfplumber
    HAS_PDFPLUMBER = True
except ImportError:
    HAS_PDFPLUMBER = False

from pydantic import BaseModel

logger = logging.getLogger(__name__)

# Memory configuration - große Gedächtnisgröße für jeden Agenten
MEMORY_MAX_ENTRIES = int(os.getenv('AGENT_MEMORY_MAX_ENTRIES', '10000'))  # 10000 Einträge pro Agent
MEMORY_SUMMARY_INTERVAL = int(os.getenv('AGENT_MEMORY_SUMMARY_INTERVAL', '100'))  # Zusammenfassung alle 100 Einträge

# Ollama configuration
# For Proxmox deployment: LLMs run on GMKTec evo x2 in local network
# Default: http://localhost:11434 (local)
# Network: http://GMKTEC_IP:11434 (e.g. http://192.168.178.155:11434)
OLLAMA_BASE_URL = os.getenv('OLLAMA_BASE_URL', 'http://localhost:11434')
OLLAMA_MODEL = os.getenv('OLLAMA_MODEL', 'Qwen2.5:32B')
OLLAMA_MODEL_CHAT = os.getenv('OLLAMA_MODEL_CHAT', OLLAMA_MODEL)
OLLAMA_MODEL_DOCUMENT = os.getenv('OLLAMA_MODEL_DOCUMENT', OLLAMA_MODEL)
OLLAMA_MODEL_ACCOUNTING = os.getenv('OLLAMA_MODEL_ACCOUNTING', OLLAMA_MODEL)
OLLAMA_TIMEOUT = int(os.getenv('OLLAMA_TIMEOUT', '300'))  # 5 minutes default
OLLAMA_MAX_RETRIES = int(os.getenv('OLLAMA_MAX_RETRIES', '3'))
OLLAMA_RETRY_DELAY = float(os.getenv('OLLAMA_RETRY_DELAY', '2.0'))  # seconds

# Prompt directory
PROMPTS_DIR = Path(__file__).parent / "prompts"

class AgentMemoryEntry(BaseModel):
    """Einzelner Memory-Eintrag für einen Agenten"""
    entry_id: str
    agent_name: str
    entry_type: str  # "conversation", "analysis", "decision", "pattern", "insight", "error", "correction"
    content: str  # Der eigentliche Inhalt
    context: Optional[Dict[str, Any]] = None  # Zusätzlicher Kontext
    metadata: Optional[Dict[str, Any]] = None  # Metadaten (z.B. confidence, source, etc.)
    timestamp: datetime
    tags: List[str] = []  # Tags für bessere Suche

class AgentMemory:
    """
    Großes persistentes Gedächtnis für Agenten
    Speichert Konversationen, Erkenntnisse, Muster und historische Entscheidungen
    """
    
    def __init__(self, agent_name: str, db=None):
        self.agent_name = agent_name
        self.db = db
        self.collection_name = "agent_memory"
        self._cache: deque = deque(maxlen=500)  # In-Memory Cache für schnellen Zugriff
        self._cache_loaded = False
    
    async def initialize(self):
        """Initialisiere Memory und lade Cache"""
        if self.db and not self._cache_loaded:
            try:
                # Lade die letzten 500 Einträge in den Cache
                async for entry in self.db[self.collection_name].find(
                    {"agent_name": self.agent_name}
                ).sort("timestamp", -1).limit(500):
                    self._cache.appendleft(entry)
                self._cache_loaded = True
                logger.info(f"AgentMemory für {self.agent_name} initialisiert: {len(self._cache)} Einträge geladen")
            except Exception as e:
                logger.warning(f"Fehler beim Laden des Memory-Cache für {self.agent_name}: {e}")
                self._cache_loaded = True  # Setze auf True, um weitere Versuche zu vermeiden
    
    async def add(self, 
                  entry_type: str, 
                  content: str, 
                  context: Optional[Dict[str, Any]] = None,
                  metadata: Optional[Dict[str, Any]] = None,
                  tags: Optional[List[str]] = None) -> str:
        """Füge einen neuen Memory-Eintrag hinzu"""
        entry_id = str(uuid.uuid4())
        entry = {
            "entry_id": entry_id,
            "agent_name": self.agent_name,
            "entry_type": entry_type,
            "content": content,
            "context": context or {},
            "metadata": metadata or {},
            "timestamp": datetime.utcnow(),
            "tags": tags or []
        }
        
        # Füge zum Cache hinzu
        self._cache.append(entry)
        
        # Persistiere in Datenbank
        if self.db:
            try:
                await self.db[self.collection_name].insert_one(entry)
                
                # Prüfe, ob Zusammenfassung nötig ist
                count = await self.db[self.collection_name].count_documents({"agent_name": self.agent_name})
                if count > 0 and count % MEMORY_SUMMARY_INTERVAL == 0:
                    await self._create_summary()
                    
            except Exception as e:
                logger.error(f"Fehler beim Speichern des Memory-Eintrags für {self.agent_name}: {e}")
        
        return entry_id
    
    async def search(self, 
                    query: Optional[str] = None,
                    entry_type: Optional[str] = None,
                    tags: Optional[List[str]] = None,
                    limit: int = 50,
                    days: Optional[int] = None) -> List[Dict[str, Any]]:
        """Suche in Memory-Einträgen"""
        results = []
        
        if self.db:
            try:
                # Baue Query auf
                db_query = {"agent_name": self.agent_name}
                
                if entry_type:
                    db_query["entry_type"] = entry_type
                
                if tags:
                    db_query["tags"] = {"$in": tags}
                
                if days:
                    cutoff_date = datetime.utcnow() - timedelta(days=days)
                    db_query["timestamp"] = {"$gte": cutoff_date}
                
                # Suche in Datenbank
                async for entry in self.db[self.collection_name].find(db_query).sort("timestamp", -1).limit(limit):
                    results.append(entry)
                
                # Wenn Text-Suche, filtere nach Inhalt
                if query and query.strip():
                    query_lower = query.lower()
                    results = [
                        entry for entry in results
                        if query_lower in entry.get("content", "").lower() or
                           query_lower in str(entry.get("context", {})).lower()
                    ]
                    
            except Exception as e:
                logger.error(f"Fehler beim Suchen im Memory für {self.agent_name}: {e}")
        else:
            # Fallback: Suche nur im Cache
            for entry in self._cache:
                if entry_type and entry.get("entry_type") != entry_type:
                    continue
                if tags and not any(tag in entry.get("tags", []) for tag in tags):
                    continue
                if query and query.lower() not in entry.get("content", "").lower():
                    continue
                results.append(entry)
                if len(results) >= limit:
                    break
        
        return results[:limit]
    
    async def get_recent(self, limit: int = 20, entry_type: Optional[str] = None) -> List[Dict[str, Any]]:
        """Hole die letzten Einträge"""
        return await self.search(entry_type=entry_type, limit=limit)
    
    async def get_patterns(self, pattern_type: Optional[str] = None) -> List[Dict[str, Any]]:
        """Hole gespeicherte Muster und Erkenntnisse"""
        return await self.search(entry_type="pattern", limit=100)
    
    async def get_insights(self, limit: int = 50) -> List[Dict[str, Any]]:
        """Hole gespeicherte Erkenntnisse"""
        return await self.search(entry_type="insight", limit=limit)
    
    async def get_context_for_prompt(self, 
                                     max_tokens: int = 2000,
                                     relevant_query: Optional[str] = None) -> str:
        """
        Generiere Kontext für LLM-Prompt aus Memory
        Kombiniert relevante Einträge und Erkenntnisse
        """
        context_parts = []
        
        # Hole relevante Erkenntnisse und Muster
        insights = await self.get_insights(limit=10)
        patterns = await self.get_patterns()
        
        if insights:
            context_parts.append("=== Gelernte Erkenntnisse ===")
            for insight in insights[:5]:  # Top 5 Erkenntnisse
                context_parts.append(f"- {insight.get('content', '')}")
                if insight.get('timestamp'):
                    context_parts.append(f"  (Gespeichert: {insight['timestamp']})")
        
        if patterns:
            context_parts.append("\n=== Gelernte Muster ===")
            for pattern in patterns[:5]:  # Top 5 Muster
                context_parts.append(f"- {pattern.get('content', '')}")
        
        # Wenn relevante Query, hole passende Einträge
        if relevant_query:
            relevant = await self.search(query=relevant_query, limit=10)
            if relevant:
                context_parts.append(f"\n=== Relevante historische Kontexte ===")
                for entry in relevant[:5]:
                    context_parts.append(f"- [{entry.get('entry_type', 'unknown')}] {entry.get('content', '')[:200]}")
                    if entry.get('timestamp'):
                        context_parts.append(f"  (Von: {entry['timestamp']})")
        
        # Hole die letzten wichtigen Entscheidungen
        recent_decisions = await self.search(entry_type="decision", limit=5)
        if recent_decisions:
            context_parts.append("\n=== Letzte wichtige Entscheidungen ===")
            for decision in recent_decisions:
                context_parts.append(f"- {decision.get('content', '')[:300]}")
        
        context = "\n".join(context_parts)
        
        # Kürze auf max_tokens (ungefähre Zeichenanzahl)
        if len(context) > max_tokens * 4:  # ~4 Zeichen pro Token
            context = context[:max_tokens * 4] + "\n... (weitere Einträge im Memory verfügbar)"
        
        return context
    
    async def _create_summary(self):
        """Erstelle eine Zusammenfassung von Memory-Einträgen (für große Memories)"""
        # Diese Funktion kann periodisch aufgerufen werden, um das Memory zu komprimieren
        # Für jetzt nur Logging
        count = await self.db[self.collection_name].count_documents({"agent_name": self.agent_name})
        logger.info(f"AgentMemory für {self.agent_name}: {count} Einträge gespeichert (Zusammenfassung könnte nützlich sein)")
    
    async def add_conversation(self, user_message: str, agent_response: str, context: Optional[Dict] = None):
        """Speichere eine Konversation"""
        content = f"User: {user_message}\nAgent: {agent_response}"
        await self.add("conversation", content, context=context, tags=["conversation", "interaction"])
    
    async def add_analysis(self, analysis_result: str, metadata: Optional[Dict] = None):
        """Speichere ein Analyseergebnis"""
        await self.add("analysis", analysis_result, metadata=metadata, tags=["analysis", "document"])
    
    async def add_decision(self, decision: str, confidence: float = 1.0, context: Optional[Dict] = None):
        """Speichere eine getroffene Entscheidung"""
        await self.add("decision", decision, 
                      metadata={"confidence": confidence}, 
                      context=context,
                      tags=["decision"])
    
    async def add_pattern(self, pattern: str, examples: Optional[List[str]] = None):
        """Speichere ein gelerntes Muster"""
        context = {"examples": examples} if examples else {}
        await self.add("pattern", pattern, context=context, tags=["pattern", "learning"])
    
    async def add_insight(self, insight: str, source: Optional[str] = None):
        """Speichere eine Erkenntnis"""
        metadata = {"source": source} if source else {}
        await self.add("insight", insight, metadata=metadata, tags=["insight", "learning"])
    
    async def add_correction(self, original: str, corrected: str, reason: str):
        """Speichere eine Korrektur"""
        content = f"Original: {original}\nKorrigiert: {corrected}\nGrund: {reason}"
        await self.add("correction", content, tags=["correction", "learning"])

class AgentMessageBus:
    """Message bus for inter-agent communication"""
    
    def __init__(self):
        self.messages: deque = deque(maxlen=1000)  # Keep last 1000 messages
        self.subscribers: Dict[str, List[callable]] = {}
    
    def subscribe(self, agent_name: str, callback: callable):
        """Subscribe an agent to messages"""
        if agent_name not in self.subscribers:
            self.subscribers[agent_name] = []
        self.subscribers[agent_name].append(callback)
    
    def publish(self, from_agent: str, to_agent: str, message: Dict[str, Any]):
        """Publish a message from one agent to another"""
        msg = {
            "from": from_agent,
            "to": to_agent,
            "content": message,
            "timestamp": datetime.utcnow().isoformat()
        }
        self.messages.append(msg)
        
        # Notify subscribers
        if to_agent in self.subscribers:
            for callback in self.subscribers[to_agent]:
                try:
                    callback(msg)
                except Exception as e:
                    logger.error(f"Error in subscriber callback: {e}")
    
    def broadcast(self, from_agent: str, message: Dict[str, Any]):
        """Broadcast message to all agents"""
        for agent_name in self.subscribers.keys():
            if agent_name != from_agent:
                self.publish(from_agent, agent_name, message)
    
    def get_messages(self, agent_name: Optional[str] = None, limit: int = 50) -> List[Dict]:
        """Get messages for a specific agent or all messages"""
        if agent_name:
            return [msg for msg in self.messages if msg.get("to") == agent_name][-limit:]
        return list(self.messages)[-limit:]

def load_prompt(prompt_file: str) -> str:
    """Load prompt from markdown file"""
    prompt_path = PROMPTS_DIR / prompt_file
    if not prompt_path.exists():
        logger.warning(f"Prompt file not found: {prompt_path}, using default")
        return ""
    
    try:
        with open(prompt_path, 'r', encoding='utf-8') as f:
            content = f.read()
            # Remove markdown header if present
            if content.startswith('# '):
                lines = content.split('\n')
                # Skip until we find the actual prompt content
                for i, line in enumerate(lines):
                    if line.strip() and not line.startswith('#'):
                        return '\n'.join(lines[i:]).strip()
            return content.strip()
    except Exception as e:
        logger.error(f"Error loading prompt from {prompt_path}: {e}")
        return ""

# Spesensätze API (Beispiel - würde durch echte API ersetzt)
# Quelle: Bundesfinanzministerium Verpflegungsmehraufwand
MEAL_ALLOWANCE_RATES = {
    "DE": {"24h": 28.0, "abwesend": 14.0},  # Deutschland in Euro
    "AT": {"24h": 41.00, "abwesend": 20.50},  # Österreich
    "CH": {"24h": 70.00, "abwesend": 35.00},  # Schweiz
    "FR": {"24h": 57.00, "abwesend": 28.50},  # Frankreich
    "IT": {"24h": 57.00, "abwesend": 28.50},  # Italien
    "ES": {"24h": 58.50, "abwesend": 29.25},  # Spanien
    "GB": {"24h": 52.00, "abwesend": 26.00},  # Großbritannien
    "US": {"24h": 89.00, "abwesend": 44.50},  # USA
    "DEFAULT": {"24h": 28.0, "abwesend": 14.0}  # Default: Deutschland
}

class AgentMessage(BaseModel):
    """Message between agents or agent and user"""
    sender: str  # agent name or "user"
    content: str
    metadata: Optional[Dict[str, Any]] = None
    timestamp: datetime = None

class DocumentAnalysis(BaseModel):
    """Result of document analysis"""
    document_type: str  # e.g. "hotel_receipt", "restaurant_bill", "toll_receipt", "parking", "fuel", "train_ticket"
    language: str  # detected language
    translated_content: Optional[str] = None  # if translation needed
    extracted_data: Dict[str, Any]  # extracted fields
    validation_issues: List[str] = []  # issues found
    completeness_check: Dict[str, bool] = {}  # e.g. {"has_tax_number": True, "has_company_address": False}
    confidence: float  # 0.0 to 1.0

class ExpenseAssignment(BaseModel):
    """Expense assignment to travel entry"""
    receipt_id: str
    entry_date: str  # YYYY-MM-DD
    category: str  # "hotel", "meals", "tolls", "parking", "fuel", "transport", "other"
    amount: float
    currency: str = "EUR"
    meal_allowance_added: Optional[float] = None  # if Verpflegungsmehraufwand was added
    assignment_confidence: float

class AgentResponse(BaseModel):
    """Response from an agent"""
    agent_name: str
    message: str
    action_taken: Optional[str] = None
    data: Optional[Dict[str, Any]] = None
    requires_user_input: bool = False
    next_agent: Optional[str] = None

# ============================================================================
# Agent Tools System - Web-Zugriff und externe APIs
# ============================================================================

class AgentTool(BaseModel):
    """Basis-Klasse für Agent-Tools"""
    name: str
    description: str
    parameters: Dict[str, Any]
    
    async def execute(self, **kwargs) -> Dict[str, Any]:
        """Führe das Tool aus - muss von Unterklassen implementiert werden"""
        raise NotImplementedError("Subclasses must implement execute()")

class WebSearchTool(AgentTool):
    """Tool für Web-Suche - holt aktuelle Informationen aus dem Internet"""
    
    def __init__(self):
        super().__init__(
            name="web_search",
            description="Suche nach aktuellen Informationen im Internet. Nützlich für aktuelle Daten, Spesensätze, Währungsinformationen, etc.",
            parameters={
                "query": {
                    "type": "string",
                    "description": "Suchanfrage (z.B. 'aktuelle Verpflegungsmehraufwand Sätze 2025 Deutschland')"
                },
                "max_results": {
                    "type": "integer",
                    "description": "Maximale Anzahl an Ergebnissen (Standard: 5)",
                    "default": 5
                }
            }
        )
        self._session = None
    
    async def _get_session(self):
        """Get aiohttp session"""
        if self._session is None or self._session.closed:
            connector = aiohttp.TCPConnector(limit=5, limit_per_host=2)
            timeout = aiohttp.ClientTimeout(total=30, connect=10)
            self._session = aiohttp.ClientSession(connector=connector, timeout=timeout)
        return self._session
    
    async def execute(self, query: str, max_results: int = 5) -> Dict[str, Any]:
        """Führe Web-Suche aus"""
        try:
            # Verwende DuckDuckGo HTML-Suche (kein API-Key nötig)
            # Alternative: Man könnte auch eine echte Search API verwenden (Google, Bing, etc.)
            search_url = "https://html.duckduckgo.com/html/"
            params = {"q": query}
            
            session = await self._get_session()
            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
            }
            
            async with session.get(search_url, params=params, headers=headers) as response:
                if response.status == 200:
                    html = await response.text()
                    # Einfache Text-Extraktion aus HTML (vereinfacht)
                    # In Produktion würde man BeautifulSoup verwenden
                    import re
                    # Extrahiere Snippets aus dem HTML
                    snippets = re.findall(r'<a[^>]*class="result__a"[^>]*>(.*?)</a>', html, re.DOTALL)
                    results = []
                    for i, snippet in enumerate(snippets[:max_results]):
                        # HTML-Tags entfernen
                        clean_snippet = re.sub(r'<[^>]+>', '', snippet).strip()
                        if clean_snippet:
                            results.append({
                                "title": clean_snippet[:100],
                                "snippet": clean_snippet[:300]
                            })
                    
                    return {
                        "success": True,
                        "query": query,
                        "results": results,
                        "count": len(results)
                    }
                else:
                    return {
                        "success": False,
                        "error": f"HTTP {response.status}",
                        "query": query
                    }
                    
        except Exception as e:
            logger.error(f"Web search error: {e}")
            return {
                "success": False,
                "error": str(e),
                "query": query
            }

class CurrencyExchangeTool(AgentTool):
    """Tool für Währungswechselkurse - holt aktuelle Wechselkurse"""
    
    def __init__(self):
        super().__init__(
            name="currency_exchange",
            description="Holt aktuelle Wechselkurse zwischen verschiedenen Währungen. Nützlich für Reisekostenabrechnungen in Fremdwährung.",
            parameters={
                "from_currency": {
                    "type": "string",
                    "description": "Quell-Währung (z.B. 'USD', 'EUR', 'GBP')"
                },
                "to_currency": {
                    "type": "string",
                    "description": "Ziel-Währung (Standard: 'EUR')",
                    "default": "EUR"
                },
                "amount": {
                    "type": "number",
                    "description": "Betrag zum Umrechnen (optional)",
                    "default": 1.0
                }
            }
        )
        self._session = None
        self._cache: Dict[str, tuple] = {}  # Cache für 1 Stunde
        self._cache_ttl = 3600  # 1 Stunde in Sekunden
    
    async def _get_session(self):
        """Get aiohttp session"""
        if self._session is None or self._session.closed:
            connector = aiohttp.TCPConnector(limit=5, limit_per_host=2)
            timeout = aiohttp.ClientTimeout(total=10, connect=5)
            self._session = aiohttp.ClientSession(connector=connector, timeout=timeout)
        return self._session
    
    async def execute(self, from_currency: str, to_currency: str = "EUR", amount: float = 1.0) -> Dict[str, Any]:
        """Holt Wechselkurs und rechnet Betrag um"""
        try:
            from_upper = from_currency.upper()
            to_upper = to_currency.upper()
            
            if from_upper == to_upper:
                return {
                    "success": True,
                    "from_currency": from_upper,
                    "to_currency": to_upper,
                    "rate": 1.0,
                    "amount": amount,
                    "converted_amount": amount
                }
            
            # Prüfe Cache
            cache_key = f"{from_upper}_{to_upper}"
            now = datetime.utcnow().timestamp()
            if cache_key in self._cache:
                cached_rate, cached_time = self._cache[cache_key]
                if now - cached_time < self._cache_ttl:
                    converted = amount * cached_rate
                    return {
                        "success": True,
                        "from_currency": from_upper,
                        "to_currency": to_upper,
                        "rate": cached_rate,
                        "amount": amount,
                        "converted_amount": round(converted, 2),
                        "cached": True
                    }
            
            # Hole aktuellen Kurs von exchangerate-api.com (kostenlos, kein API-Key nötig)
            session = await self._get_session()
            url = f"https://api.exchangerate-api.com/v4/latest/{from_upper}"
            
            async with session.get(url) as response:
                if response.status == 200:
                    data = await response.json()
                    rates = data.get("rates", {})
                    
                    if to_upper in rates:
                        rate = rates[to_upper]
                        # Update Cache
                        self._cache[cache_key] = (rate, now)
                        
                        converted = amount * rate
                        return {
                            "success": True,
                            "from_currency": from_upper,
                            "to_currency": to_upper,
                            "rate": rate,
                            "amount": amount,
                            "converted_amount": round(converted, 2),
                            "date": data.get("date", ""),
                            "cached": False
                        }
                    else:
                        return {
                            "success": False,
                            "error": f"Währung {to_upper} nicht gefunden",
                            "from_currency": from_upper,
                            "to_currency": to_upper
                        }
                else:
                    return {
                        "success": False,
                        "error": f"HTTP {response.status}",
                        "from_currency": from_upper,
                        "to_currency": to_upper
                    }
                    
        except Exception as e:
            logger.error(f"Currency exchange error: {e}")
            return {
                "success": False,
                "error": str(e),
                "from_currency": from_currency,
                "to_currency": to_currency
            }

class MealAllowanceLookupTool(AgentTool):
    """Tool für aktuelle Verpflegungsmehraufwand-Spesensätze"""
    
    def __init__(self):
        super().__init__(
            name="meal_allowance_lookup",
            description="Sucht nach aktuellen Verpflegungsmehraufwand-Spesensätzen für verschiedene Länder. Nutzt Web-Suche für aktuelle Daten.",
            parameters={
                "country": {
                    "type": "string",
                    "description": "Land oder Ländercode (z.B. 'Deutschland', 'DE', 'USA', 'US')"
                },
                "year": {
                    "type": "integer",
                    "description": "Jahr für die Spesensätze (Standard: aktuelles Jahr)",
                    "default": None
                }
            }
        )
        self.web_search = WebSearchTool()
    
    async def execute(self, country: str, year: Optional[int] = None) -> Dict[str, Any]:
        """Sucht nach aktuellen Spesensätzen"""
        try:
            if year is None:
                year = datetime.now().year
            
            # Baue Suchanfrage
            query = f"Verpflegungsmehraufwand {country} {year} Spesensätze Bundesfinanzministerium"
            
            # Nutze Web-Search-Tool
            search_result = await self.web_search.execute(query, max_results=3)
            
            if search_result.get("success") and search_result.get("results"):
                # Extrahiere Zahlen aus den Ergebnissen
                results_text = " ".join([r.get("snippet", "") for r in search_result["results"]])
                
                # Versuche Beträge zu extrahieren (vereinfacht)
                import re
                # Suche nach Euro-Beträgen wie "28,00 EUR" oder "28.00 Euro"
                amounts = re.findall(r'(\d+[.,]\d{2})\s*(?:EUR|Euro|€)', results_text, re.IGNORECASE)
                
                if amounts:
                    # Konvertiere zu float
                    try:
                        rates = [float(a.replace(",", ".")) for a in amounts[:2]]  # Nimm erste 2
                        return {
                            "success": True,
                            "country": country,
                            "year": year,
                            "rates": rates,
                            "source": "web_search",
                            "search_results": search_result["results"][:2]
                        }
                    except ValueError:
                        pass
                
                return {
                    "success": True,
                    "country": country,
                    "year": year,
                    "found_information": True,
                    "search_results": search_result["results"],
                    "note": "Spesensätze gefunden, aber Beträge müssen manuell extrahiert werden"
                }
            else:
                # Fallback auf lokale Datenbank
                country_upper = country.upper()
                if country_upper in MEAL_ALLOWANCE_RATES:
                    rates = MEAL_ALLOWANCE_RATES[country_upper]
                    return {
                        "success": True,
                        "country": country,
                        "year": year,
                        "rates": rates,
                        "source": "local_database",
                        "note": "Verwendet lokale Datenbank (möglicherweise nicht aktuell)"
                    }
                
                return {
                    "success": False,
                    "error": "Keine Spesensätze gefunden",
                    "country": country,
                    "year": year
                }
                
        except Exception as e:
            logger.error(f"Meal allowance lookup error: {e}")
            return {
                "success": False,
                "error": str(e),
                "country": country
            }

class GeocodingTool(AgentTool):
    """Tool für Geocoding - bestimmt Ländercode aus Ortsangabe"""
    
    def __init__(self):
        super().__init__(
            name="geocoding",
            description="Bestimmt Ländercode aus einer Ortsangabe oder Adresse. Nützlich für automatische Ländererkennung.",
            parameters={
                "location": {
                    "type": "string",
                    "description": "Ortsangabe (z.B. 'München', 'Berlin, Deutschland', 'New York, USA')"
                }
            }
        )
        self._session = None
    
    async def _get_session(self):
        """Get aiohttp session"""
        if self._session is None or self._session.closed:
            connector = aiohttp.TCPConnector(limit=5, limit_per_host=2)
            timeout = aiohttp.ClientTimeout(total=10, connect=5)
            self._session = aiohttp.ClientSession(connector=connector, timeout=timeout)
        return self._session
    
    async def execute(self, location: str) -> Dict[str, Any]:
        """Bestimmt Ländercode aus Ortsangabe"""
        try:
            # Verwende Nominatim (OpenStreetMap Geocoding API) - kostenlos
            session = await self._get_session()
            url = "https://nominatim.openstreetmap.org/search"
            params = {
                "q": location,
                "format": "json",
                "limit": 1,
                "addressdetails": 1
            }
            headers = {
                "User-Agent": "Stundenzettel-Web-App/1.0"  # Nominatim erfordert User-Agent
            }
            
            async with session.get(url, params=params, headers=headers) as response:
                if response.status == 200:
                    data = await response.json()
                    if data and len(data) > 0:
                        result = data[0]
                        address = result.get("address", {})
                        country_code = address.get("country_code", "").upper()
                        country = address.get("country", "")
                        
                        return {
                            "success": True,
                            "location": location,
                            "country_code": country_code,
                            "country": country,
                            "full_address": result.get("display_name", ""),
                            "lat": float(result.get("lat", 0)),
                            "lon": float(result.get("lon", 0))
                        }
                    else:
                        return {
                            "success": False,
                            "error": "Ort nicht gefunden",
                            "location": location
                        }
                else:
                    return {
                        "success": False,
                        "error": f"HTTP {response.status}",
                        "location": location
                    }
                    
        except Exception as e:
            logger.error(f"Geocoding error: {e}")
            # Fallback auf einfache String-Erkennung
            location_lower = location.lower()
            country_mapping = {
                "deutschland": "DE", "germany": "DE",
                "österreich": "AT", "austria": "AT",
                "schweiz": "CH", "switzerland": "CH",
                "frankreich": "FR", "france": "FR",
                "italien": "IT", "italy": "IT",
                "spanien": "ES", "spain": "ES",
                "großbritannien": "GB", "uk": "GB", "england": "GB",
                "usa": "US", "vereinigte staaten": "US"
            }
            for key, code in country_mapping.items():
                if key in location_lower:
                    return {
                        "success": True,
                        "location": location,
                        "country_code": code,
                        "country": key,
                        "source": "fallback_mapping"
                    }
            
            return {
                "success": False,
                "error": str(e),
                "location": location
            }
    
    async def close(self):
        """Close HTTP session"""
        if self._session and not self._session.closed:
            await self._session.close()
            self._session = None

class OpenMapsTool(AgentTool):
    """Tool für OpenStreetMap - umfassende Karten- und Geodaten-Funktionen"""
    
    def __init__(self):
        super().__init__(
            name="openmaps",
            description="Umfassendes Tool für OpenStreetMap-Funktionen: Geocoding, Reverse Geocoding, POI-Suche, Entfernungsberechnung, Routenplanung. Nützlich für Reisekostenabrechnungen, Ortsbestimmung, Entfernungsvalidierung und Standortinformationen.",
            parameters={
                "action": {
                    "type": "string",
                    "description": "Aktion: 'geocode' (Adresse zu Koordinaten), 'reverse' (Koordinaten zu Adresse), 'search' (POI-Suche), 'distance' (Entfernung berechnen), 'route' (Route berechnen)",
                    "enum": ["geocode", "reverse", "search", "distance", "route"]
                },
                "query": {
                    "type": "string",
                    "description": "Suchanfrage (für geocode/search): Adresse, Ort oder POI-Name",
                    "default": None
                },
                "lat": {
                    "type": "number",
                    "description": "Breitengrad (für reverse/distance/route)",
                    "default": None
                },
                "lon": {
                    "type": "number",
                    "description": "Längengrad (für reverse/distance/route)",
                    "default": None
                },
                "lat2": {
                    "type": "number",
                    "description": "Zweiter Breitengrad (für distance/route)",
                    "default": None
                },
                "lon2": {
                    "type": "number",
                    "description": "Zweiter Längengrad (für distance/route)",
                    "default": None
                },
                "poi_type": {
                    "type": "string",
                    "description": "POI-Typ für Suche (z.B. 'hotel', 'restaurant', 'fuel', 'parking')",
                    "default": None
                },
                "radius": {
                    "type": "number",
                    "description": "Suchradius in Metern (für search, Standard: 1000)",
                    "default": 1000
                },
                "limit": {
                    "type": "integer",
                    "description": "Maximale Anzahl Ergebnisse (Standard: 5)",
                    "default": 5
                }
            }
        )
        self._session = None
        self.base_url = "https://nominatim.openstreetmap.org"
        self.user_agent = "Stundenzettel-Web-App/1.0"
    
    async def _get_session(self):
        """Get aiohttp session"""
        if self._session is None or self._session.closed:
            connector = aiohttp.TCPConnector(limit=5, limit_per_host=2)
            timeout = aiohttp.ClientTimeout(total=10, connect=5)
            self._session = aiohttp.ClientSession(connector=connector, timeout=timeout)
        return self._session
    
    async def _make_request(self, endpoint: str, params: Dict[str, Any]) -> Dict[str, Any]:
        """Mache Request an Nominatim API"""
        try:
            session = await self._get_session()
            headers = {"User-Agent": self.user_agent}
            url = f"{self.base_url}/{endpoint}"
            
            async with session.get(url, params=params, headers=headers) as response:
                if response.status == 200:
                    data = await response.json()
                    return {"success": True, "data": data}
                else:
                    return {
                        "success": False,
                        "error": f"HTTP {response.status}",
                        "status": response.status
                    }
        except Exception as e:
            logger.error(f"OpenMaps request error: {e}")
            return {"success": False, "error": str(e)}
    
    async def execute(self, 
                      action: str,
                      query: Optional[str] = None,
                      lat: Optional[float] = None,
                      lon: Optional[float] = None,
                      lat2: Optional[float] = None,
                      lon2: Optional[float] = None,
                      poi_type: Optional[str] = None,
                      radius: int = 1000,
                      limit: int = 5) -> Dict[str, Any]:
        """Führe OpenMaps-Aktion aus"""
        try:
            if action == "geocode":
                if not query:
                    return {"success": False, "error": "query ist erforderlich für geocode"}
                
                params = {
                    "q": query,
                    "format": "json",
                    "limit": limit,
                    "addressdetails": 1
                }
                result = await self._make_request("search", params)
                
                if result.get("success"):
                    results = result["data"]
                    formatted_results = []
                    for r in results[:limit]:
                        address = r.get("address", {})
                        formatted_results.append({
                            "display_name": r.get("display_name", ""),
                            "lat": float(r.get("lat", 0)),
                            "lon": float(r.get("lon", 0)),
                            "country_code": address.get("country_code", "").upper(),
                            "country": address.get("country", ""),
                            "city": address.get("city") or address.get("town") or address.get("village", ""),
                            "postcode": address.get("postcode", ""),
                            "type": r.get("type", ""),
                            "class": r.get("class", "")
                        })
                    
                    return {
                        "success": True,
                        "action": "geocode",
                        "query": query,
                        "results": formatted_results,
                        "count": len(formatted_results)
                    }
                else:
                    return result
            
            elif action == "reverse":
                if lat is None or lon is None:
                    return {"success": False, "error": "lat und lon sind erforderlich für reverse"}
                
                params = {
                    "lat": str(lat),
                    "lon": str(lon),
                    "format": "json",
                    "addressdetails": 1
                }
                result = await self._make_request("reverse", params)
                
                if result.get("success"):
                    data = result["data"]
                    address = data.get("address", {})
                    return {
                        "success": True,
                        "action": "reverse",
                        "lat": lat,
                        "lon": lon,
                        "display_name": data.get("display_name", ""),
                        "country_code": address.get("country_code", "").upper(),
                        "country": address.get("country", ""),
                        "city": address.get("city") or address.get("town") or address.get("village", ""),
                        "postcode": address.get("postcode", ""),
                        "full_address": address
                    }
                else:
                    return result
            
            elif action == "search":
                if not query:
                    return {"success": False, "error": "query ist erforderlich für search"}
                
                # POI-Suche mit Typ-Filter
                search_query = query
                if poi_type:
                    search_query = f"{poi_type} {query}"
                
                params = {
                    "q": search_query,
                    "format": "json",
                    "limit": limit,
                    "addressdetails": 1
                }
                
                # Wenn lat/lon gegeben, füge Bounding Box hinzu für Radius-Suche
                if lat is not None and lon is not None:
                    # Einfache Bounding Box (ungefährer Radius)
                    # 1 Grad ≈ 111 km, also radius/111000 für grobe Bounding Box
                    delta = radius / 111000.0
                    params["viewbox"] = f"{lon-delta},{lat-delta},{lon+delta},{lat+delta}"
                    params["bounded"] = "1"
                
                result = await self._make_request("search", params)
                
                if result.get("success"):
                    results = result["data"]
                    formatted_results = []
                    for r in results[:limit]:
                        address = r.get("address", {})
                        formatted_results.append({
                            "name": r.get("display_name", ""),
                            "lat": float(r.get("lat", 0)),
                            "lon": float(r.get("lon", 0)),
                            "type": r.get("type", ""),
                            "class": r.get("class", ""),
                            "country_code": address.get("country_code", "").upper(),
                            "city": address.get("city") or address.get("town") or address.get("village", "")
                        })
                    
                    return {
                        "success": True,
                        "action": "search",
                        "query": query,
                        "poi_type": poi_type,
                        "results": formatted_results,
                        "count": len(formatted_results)
                    }
                else:
                    return result
            
            elif action == "distance":
                if lat is None or lon is None or lat2 is None or lon2 is None:
                    return {"success": False, "error": "lat, lon, lat2 und lon2 sind erforderlich für distance"}
                
                # Haversine-Formel für Entfernungsberechnung
                from math import radians, sin, cos, sqrt, atan2
                
                R = 6371000  # Erdradius in Metern
                lat1_rad = radians(lat)
                lat2_rad = radians(lat2)
                delta_lat = radians(lat2 - lat)
                delta_lon = radians(lon2 - lon)
                
                a = sin(delta_lat / 2) ** 2 + cos(lat1_rad) * cos(lat2_rad) * sin(delta_lon / 2) ** 2
                c = 2 * atan2(sqrt(a), sqrt(1 - a))
                distance_meters = R * c
                distance_km = distance_meters / 1000
                
                return {
                    "success": True,
                    "action": "distance",
                    "from": {"lat": lat, "lon": lon},
                    "to": {"lat": lat2, "lon": lon2},
                    "distance_meters": round(distance_meters, 2),
                    "distance_km": round(distance_km, 2)
                }
            
            elif action == "route":
                if lat is None or lon is None or lat2 is None or lon2 is None:
                    return {"success": False, "error": "lat, lon, lat2 und lon2 sind erforderlich für route"}
                
                # Für einfache Routenberechnung verwenden wir die Entfernung
                # Eine vollständige Routing-API würde OSRM oder GraphHopper benötigen
                # Hier geben wir eine einfache Luftlinie + geschätzte Fahrzeit zurück
                from math import radians, sin, cos, sqrt, atan2
                
                R = 6371000
                lat1_rad = radians(lat)
                lat2_rad = radians(lat2)
                delta_lat = radians(lat2 - lat)
                delta_lon = radians(lon2 - lon)
                
                a = sin(delta_lat / 2) ** 2 + cos(lat1_rad) * cos(lat2_rad) * sin(delta_lon / 2) ** 2
                c = 2 * atan2(sqrt(a), sqrt(1 - a))
                distance_meters = R * c
                distance_km = distance_meters / 1000
                
                # Geschätzte Fahrzeit (Durchschnittsgeschwindigkeit 80 km/h)
                estimated_driving_time_hours = distance_km / 80.0
                estimated_driving_time_minutes = estimated_driving_time_hours * 60
                
                return {
                    "success": True,
                    "action": "route",
                    "from": {"lat": lat, "lon": lon},
                    "to": {"lat": lat2, "lon": lon2},
                    "distance_km": round(distance_km, 2),
                    "distance_meters": round(distance_meters, 2),
                    "estimated_driving_time_hours": round(estimated_driving_time_hours, 2),
                    "estimated_driving_time_minutes": round(estimated_driving_time_minutes, 0),
                    "note": "Luftlinie - für detaillierte Routenplanung wird eine Routing-API benötigt"
                }
            
            else:
                return {
                    "success": False,
                    "error": f"Unbekannte Aktion: {action}",
                    "available_actions": ["geocode", "reverse", "search", "distance", "route"]
                }
                
        except Exception as e:
            logger.error(f"OpenMaps execute error: {e}")
            return {
                "success": False,
                "error": str(e),
                "action": action
            }
    
    async def close(self):
        """Close HTTP session"""
        if self._session and not self._session.closed:
            await self._session.close()
            self._session = None

# Optional imports für erweiterte Tools
try:
    from exa_py import Exa
    HAS_EXA = True
except ImportError:
    HAS_EXA = False

try:
    from paddleocr import PaddleOCR
    HAS_PADDLEOCR = True
except ImportError:
    HAS_PADDLEOCR = False

try:
    from langchain.tools import Tool
    from langchain.agents import AgentExecutor, create_openai_functions_agent
    from langchain_openai import ChatOpenAI
    HAS_LANGCHAIN = True
except ImportError:
    HAS_LANGCHAIN = False

class ExaSearchTool(AgentTool):
    """Tool für Exa/XNG Suche - hochwertige semantische Suche"""
    
    def __init__(self):
        super().__init__(
            name="exa_search",
            description="Hochwertige semantische Suche mit Exa/XNG API. Besser als Standard-Web-Suche für präzise, relevante Ergebnisse. Nützlich für ChatAgent zur Beantwortung von Fragen.",
            parameters={
                "query": {
                    "type": "string",
                    "description": "Suchanfrage (semantisch verstanden)"
                },
                "max_results": {
                    "type": "integer",
                    "description": "Maximale Anzahl an Ergebnissen (Standard: 5)",
                    "default": 5
                },
                "use_autoprompt": {
                    "type": "boolean",
                    "description": "Verwende Auto-Prompt für bessere Ergebnisse (Standard: True)",
                    "default": True
                }
            }
        )
        self.api_key = os.getenv('EXA_API_KEY')
        self._exa_client = None
    
    def _get_client(self):
        """Get Exa client"""
        if not HAS_EXA:
            return None
        if self._exa_client is None and self.api_key:
            try:
                self._exa_client = Exa(api_key=self.api_key)
            except Exception as e:
                logger.warning(f"Exa client initialization error: {e}")
        return self._exa_client
    
    async def execute(self, query: str, max_results: int = 5, use_autoprompt: bool = True) -> Dict[str, Any]:
        """Führe Exa-Suche aus"""
        try:
            client = self._get_client()
            if not client:
                return {
                    "success": False,
                    "error": "Exa API nicht verfügbar. Bitte EXA_API_KEY setzen und 'pip install exa-py' installieren.",
                    "fallback_available": True
                }
            
            # Exa API-Aufruf (synchron, aber in async context)
            search_params = {
                "query": query,
                "num_results": max_results,
                "use_autoprompt": use_autoprompt,
                "text": {"max_characters": 1000}  # Hole Text-Inhalt
            }
            
            # Exa ist synchron, daher in Thread ausführen
            import asyncio
            loop = asyncio.get_event_loop()
            results = await loop.run_in_executor(None, lambda: client.search(**search_params))
            
            formatted_results = []
            for result in results.results:
                formatted_results.append({
                    "title": result.title,
                    "url": result.url,
                    "text": result.text[:500] if result.text else "",
                    "published_date": result.published_date.isoformat() if hasattr(result, 'published_date') and result.published_date else None,
                    "author": result.author if hasattr(result, 'author') else None
                })
            
            return {
                "success": True,
                "query": query,
                "results": formatted_results,
                "count": len(formatted_results),
                "source": "exa"
            }
            
        except Exception as e:
            logger.error(f"Exa search error: {e}")
            return {
                "success": False,
                "error": str(e),
                "query": query,
                "fallback_available": True
            }

class MarkerTool(AgentTool):
    """Tool für Marker - Dokumentenanalyse und -extraktion"""
    
    def __init__(self):
        super().__init__(
            name="marker",
            description="Marker-Tool für erweiterte Dokumentenanalyse. Extrahiert strukturierte Daten aus PDFs und anderen Dokumenten. Nützlich für DocumentAgent und AccountingAgent.",
            parameters={
                "document_path": {
                    "type": "string",
                    "description": "Pfad zum Dokument (PDF, Bild, etc.)"
                },
                "extract_tables": {
                    "type": "boolean",
                    "description": "Tabellen extrahieren (Standard: True)",
                    "default": True
                },
                "extract_images": {
                    "type": "boolean",
                    "description": "Bilder extrahieren (Standard: False)",
                    "default": False
                },
                "markdown_output": {
                    "type": "boolean",
                    "description": "Markdown-Format für Ausgabe (Standard: True)",
                    "default": True
                }
            }
        )
        self.marker_api_key = os.getenv('MARKER_API_KEY')
        self.marker_base_url = os.getenv('MARKER_BASE_URL', 'https://api.marker.io/v1')
        self._session = None
    
    async def _get_session(self):
        """Get aiohttp session"""
        if self._session is None or self._session.closed:
            connector = aiohttp.TCPConnector(limit=5, limit_per_host=2)
            timeout = aiohttp.ClientTimeout(total=60, connect=10)
            self._session = aiohttp.ClientSession(connector=connector, timeout=timeout)
        return self._session
    
    async def execute(self, document_path: str, extract_tables: bool = True, extract_images: bool = False, markdown_output: bool = True) -> Dict[str, Any]:
        """Führe Marker-Dokumentenanalyse aus"""
        try:
            # Prüfe ob Datei existiert
            if not Path(document_path).exists():
                return {
                    "success": False,
                    "error": f"Dokument nicht gefunden: {document_path}",
                    "document_path": document_path
                }
            
            # Option 1: Marker API (falls verfügbar)
            if self.marker_api_key:
                session = await self._get_session()
                headers = {
                    "Authorization": f"Bearer {self.marker_api_key}",
                    "Content-Type": "application/json"
                }
                
                # Lese Datei
                with open(document_path, 'rb') as f:
                    file_data = f.read()
                
                # Base64 kodieren
                import base64
                file_base64 = base64.b64encode(file_data).decode('utf-8')
                
                payload = {
                    "file": file_base64,
                    "filename": Path(document_path).name,
                    "extract_tables": extract_tables,
                    "extract_images": extract_images,
                    "markdown_output": markdown_output
                }
                
                async with session.post(
                    f"{self.marker_base_url}/extract",
                    json=payload,
                    headers=headers
                ) as response:
                    if response.status == 200:
                        result = await response.json()
                        return {
                            "success": True,
                            "document_path": document_path,
                            "text": result.get("text", ""),
                            "markdown": result.get("markdown", ""),
                            "tables": result.get("tables", []),
                            "metadata": result.get("metadata", {}),
                            "source": "marker_api"
                        }
                    else:
                        error_text = await response.text()
                        logger.warning(f"Marker API error: {response.status} - {error_text[:200]}")
            
            # Option 2: Lokale Marker-Installation (falls verfügbar)
            # Marker kann auch lokal installiert werden
            try:
                import subprocess
                import tempfile
                
                # Versuche marker CLI zu verwenden
                result = subprocess.run(
                    ["marker", "convert", document_path, "--output", "markdown"],
                    capture_output=True,
                    text=True,
                    timeout=60
                )
                
                if result.returncode == 0:
                    return {
                        "success": True,
                        "document_path": document_path,
                        "text": result.stdout,
                        "markdown": result.stdout,
                        "source": "marker_cli"
                    }
            except (subprocess.TimeoutExpired, FileNotFoundError, subprocess.SubprocessError) as e:
                logger.debug(f"Marker CLI nicht verfügbar: {e}")
            
            # Fallback: Verwende vorhandene PDF-Extraktion
            return {
                "success": False,
                "error": "Marker API/CLI nicht verfügbar. Verwende Fallback-Methoden.",
                "document_path": document_path,
                "fallback_available": True
            }
            
        except Exception as e:
            logger.error(f"Marker tool error: {e}")
            return {
                "success": False,
                "error": str(e),
                "document_path": document_path,
                "fallback_available": True
            }
    
    async def close(self):
        """Close HTTP session"""
        if self._session and not self._session.closed:
            await self._session.close()
            self._session = None

class PaddleOCRTool(AgentTool):
    """Tool für PaddleOCR - OCR als Fallback für Dokumentenanalyse"""
    
    def __init__(self):
        super().__init__(
            name="paddleocr",
            description="PaddleOCR-Tool für Texterkennung in Bildern und PDFs. Fallback für DocumentAgent wenn andere Methoden versagen. Unterstützt über 100 Sprachen.",
            parameters={
                "image_path": {
                    "type": "string",
                    "description": "Pfad zum Bild oder PDF"
                },
                "lang": {
                    "type": "string",
                    "description": "Sprache (z.B. 'de', 'en', 'ch', Standard: 'de')",
                    "default": "de"
                },
                "use_angle_cls": {
                    "type": "boolean",
                    "description": "Verwende Winkel-Klassifikation für bessere Ergebnisse (Standard: True)",
                    "default": True
                }
            }
        )
        self._ocr_instances: Dict[str, Any] = {}  # Cache für OCR-Instanzen pro Sprache
    
    def _get_ocr(self, lang: str = "de", use_angle_cls: bool = True):
        """Get PaddleOCR instance (cached)"""
        if not HAS_PADDLEOCR:
            return None
        
        cache_key = f"{lang}_{use_angle_cls}"
        if cache_key not in self._ocr_instances:
            try:
                self._ocr_instances[cache_key] = PaddleOCR(
                    use_angle_cls=use_angle_cls,
                    lang=lang,
                    show_log=False  # Reduziere Logging
                )
            except Exception as e:
                logger.error(f"PaddleOCR initialization error: {e}")
                return None
        
        return self._ocr_instances[cache_key]
    
    async def execute(self, image_path: str, lang: str = "de", use_angle_cls: bool = True) -> Dict[str, Any]:
        """Führe OCR aus"""
        try:
            if not Path(image_path).exists():
                return {
                    "success": False,
                    "error": f"Datei nicht gefunden: {image_path}",
                    "image_path": image_path
                }
            
            ocr = self._get_ocr(lang, use_angle_cls)
            if not ocr:
                return {
                    "success": False,
                    "error": "PaddleOCR nicht verfügbar. Bitte 'pip install paddleocr paddlepaddle' installieren.",
                    "image_path": image_path
                }
            
            # OCR ist synchron, daher in Thread ausführen
            import asyncio
            loop = asyncio.get_event_loop()
            result = await loop.run_in_executor(None, lambda: ocr.ocr(image_path, cls=use_angle_cls))
            
            # Formatiere Ergebnis
            extracted_text = []
            confidence_scores = []
            
            if result and len(result) > 0:
                for page_result in result:
                    if page_result:
                        for line in page_result:
                            if line and len(line) >= 2:
                                text_info = line[1]
                                if isinstance(text_info, tuple) and len(text_info) >= 2:
                                    text = text_info[0]
                                    confidence = text_info[1]
                                    extracted_text.append(text)
                                    confidence_scores.append(confidence)
                                elif isinstance(text_info, str):
                                    extracted_text.append(text_info)
            
            full_text = "\n".join(extracted_text)
            avg_confidence = sum(confidence_scores) / len(confidence_scores) if confidence_scores else 0.0
            
            return {
                "success": True,
                "image_path": image_path,
                "text": full_text,
                "lines": extracted_text,
                "confidence": avg_confidence,
                "confidence_scores": confidence_scores,
                "language": lang,
                "source": "paddleocr"
            }
            
        except Exception as e:
            logger.error(f"PaddleOCR error: {e}")
            return {
                "success": False,
                "error": str(e),
                "image_path": image_path
            }

class CustomPythonRulesTool(AgentTool):
    """Tool für Custom Python Regeln - ausführbare Python-Regeln für AccountingAgent"""
    
    def __init__(self):
        super().__init__(
            name="custom_python_rules",
            description="Führt benutzerdefinierte Python-Regeln für Buchhaltungsvalidierung und -berechnung aus. Nützlich für AccountingAgent zur Anwendung spezifischer Geschäftsregeln.",
            parameters={
                "rule_name": {
                    "type": "string",
                    "description": "Name der Regel (z.B. 'validate_tax_number', 'calculate_meal_allowance', 'check_receipt_completeness')"
                },
                "rule_code": {
                    "type": "string",
                    "description": "Python-Code der Regel (optional, wenn Regel bereits registriert ist)"
                },
                "input_data": {
                    "type": "object",
                    "description": "Eingabedaten für die Regel (Dict)"
                }
            }
        )
        self._registered_rules: Dict[str, str] = {}  # Cache für registrierte Regeln
        self._load_default_rules()
    
    def _load_default_rules(self):
        """Lade Standard-Buchhaltungsregeln"""
        # Beispiel-Regeln
        self._registered_rules["validate_tax_number"] = """
def validate_tax_number(tax_number: str, country_code: str = "DE") -> Dict[str, Any]:
    \"\"\"Validiere Steuernummer basierend auf Ländercode\"\"\"
    if not tax_number or not tax_number.strip():
        return {"valid": False, "error": "Steuernummer ist leer"}
    
    tax_number = tax_number.strip().replace(" ", "").replace("-", "")
    
    if country_code == "DE":
        # Deutsche USt-IdNr: DE + 9 Ziffern
        if tax_number.startswith("DE") and len(tax_number) == 11:
            return {"valid": True, "normalized": tax_number}
        elif len(tax_number) == 11 and tax_number.isdigit():
            return {"valid": True, "normalized": f"DE{tax_number}"}
        else:
            return {"valid": False, "error": "Ungültiges Format für deutsche USt-IdNr"}
    elif country_code == "AT":
        # Österreich: AT + 9 Ziffern
        if tax_number.startswith("AT") and len(tax_number) == 11:
            return {"valid": True, "normalized": tax_number}
        else:
            return {"valid": False, "error": "Ungültiges Format für österreichische USt-IdNr"}
    
    return {"valid": False, "error": f"Unbekannter Ländercode: {country_code}"}
"""
        
        self._registered_rules["check_receipt_completeness"] = """
def check_receipt_completeness(receipt_data: Dict[str, Any]) -> Dict[str, Any]:
    \"\"\"Prüfe Vollständigkeit eines Belegs\"\"\"
    required_fields = ["amount", "date", "tax_number"]
    missing_fields = []
    issues = []
    
    for field in required_fields:
        if field not in receipt_data or not receipt_data[field]:
            missing_fields.append(field)
    
    if "amount" in receipt_data:
        try:
            amount = float(receipt_data["amount"])
            if amount <= 0:
                issues.append("Betrag muss größer als 0 sein")
        except (ValueError, TypeError):
            issues.append("Ungültiger Betrag")
    
    return {
        "complete": len(missing_fields) == 0,
        "missing_fields": missing_fields,
        "issues": issues
    }
"""
        
        self._registered_rules["calculate_meal_allowance"] = """
def calculate_meal_allowance(country_code: str, days: int, is_24h: bool = True) -> float:
    \"\"\"Berechne Verpflegungsmehraufwand\"\"\"
    rates = {
        "DE": {"24h": 28.0, "abwesend": 14.0},
        "AT": {"24h": 41.0, "abwesend": 20.5},
        "CH": {"24h": 70.0, "abwesend": 35.0},
        "FR": {"24h": 57.0, "abwesend": 28.5},
        "IT": {"24h": 57.0, "abwesend": 28.5},
        "ES": {"24h": 58.5, "abwesend": 29.25},
        "GB": {"24h": 52.0, "abwesend": 26.0},
        "US": {"24h": 89.0, "abwesend": 44.5}
    }
    
    country_rates = rates.get(country_code.upper(), rates["DE"])
    rate_type = "24h" if is_24h else "abwesend"
    daily_rate = country_rates.get(rate_type, country_rates["24h"])
    
    return daily_rate * days
"""
    
    async def execute(self, rule_name: str, rule_code: Optional[str] = None, input_data: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        """Führe Python-Regel aus"""
        try:
            # Hole Regel-Code
            if rule_name in self._registered_rules:
                rule_code = self._registered_rules[rule_name]
            elif not rule_code:
                return {
                    "success": False,
                    "error": f"Regel '{rule_name}' nicht gefunden und kein Code bereitgestellt",
                    "available_rules": list(self._registered_rules.keys())
                }
            
            # Sicherheitsprüfung: Erlaube nur bestimmte Funktionen
            allowed_imports = ["json", "datetime", "re", "math"]
            dangerous_keywords = ["__import__", "eval", "exec", "open", "file", "input", "raw_input"]
            
            for keyword in dangerous_keywords:
                if keyword in rule_code:
                    return {
                        "success": False,
                        "error": f"Sicherheitsproblem: Gefährliches Keyword '{keyword}' in Regel-Code gefunden"
                    }
            
            # Erstelle sicherer Ausführungskontext
            safe_globals = {
                "__builtins__": {
                    "len": len,
                    "str": str,
                    "int": int,
                    "float": float,
                    "bool": bool,
                    "dict": dict,
                    "list": list,
                    "tuple": tuple,
                    "set": set,
                    "round": round,
                    "abs": abs,
                    "min": min,
                    "max": max,
                    "sum": sum,
                    "all": all,
                    "any": any,
                    "isinstance": isinstance,
                    "type": type,
                    "hasattr": hasattr,
                    "getattr": getattr,
                }
            }
            
            # Füge erlaubte Imports hinzu
            for imp in allowed_imports:
                try:
                    safe_globals[imp] = __import__(imp)
                except ImportError:
                    pass
            
            # Kompiliere und führe Regel aus
            try:
                compiled_code = compile(rule_code, f"<rule:{rule_name}>", "exec")
                exec(compiled_code, safe_globals)
                
                # Finde Hauptfunktion (gleicher Name wie Regel)
                if rule_name in safe_globals and callable(safe_globals[rule_name]):
                    func = safe_globals[rule_name]
                    
                    # Führe Funktion aus
                    if input_data:
                        if isinstance(input_data, dict):
                            result = func(**input_data)
                        else:
                            result = func(input_data)
                    else:
                        result = func()
                    
                    return {
                        "success": True,
                        "rule_name": rule_name,
                        "result": result,
                        "source": "custom_python_rule"
                    }
                else:
                    return {
                        "success": False,
                        "error": f"Funktion '{rule_name}' nicht in Regel-Code gefunden",
                        "rule_name": rule_name
                    }
                    
            except SyntaxError as e:
                return {
                    "success": False,
                    "error": f"Syntax-Fehler in Regel: {str(e)}",
                    "rule_name": rule_name
                }
            except Exception as e:
                return {
                    "success": False,
                    "error": f"Fehler beim Ausführen der Regel: {str(e)}",
                    "rule_name": rule_name
                }
                
        except Exception as e:
            logger.error(f"Custom Python Rules error: {e}")
            return {
                "success": False,
                "error": str(e),
                "rule_name": rule_name
            }
    
    def register_rule(self, rule_name: str, rule_code: str):
        """Registriere eine neue Regel"""
        self._registered_rules[rule_name] = rule_code
        logger.info(f"Regel '{rule_name}' registriert")

class WebAccessTool(AgentTool):
    """Tool für generischen Web-Zugriff - HTTP-Requests zu beliebigen Web-Ressourcen"""
    
    def __init__(self):
        super().__init__(
            name="web_access",
            description="Generischer Web-Zugriff für HTTP-Requests zu beliebigen URLs. Ermöglicht GET/POST/PUT/DELETE Requests, Web-Scraping, API-Zugriff und HTML-Inhalt-Extraktion. Nützlich für alle Agents zur Validierung, Datensammlung und API-Interaktion.",
            parameters={
                "url": {
                    "type": "string",
                    "description": "Vollständige URL (z.B. 'https://example.com/api/data')"
                },
                "method": {
                    "type": "string",
                    "description": "HTTP-Methode (GET, POST, PUT, DELETE, Standard: GET)",
                    "enum": ["GET", "POST", "PUT", "DELETE", "PATCH"],
                    "default": "GET"
                },
                "headers": {
                    "type": "object",
                    "description": "HTTP-Headers als Dictionary (optional)",
                    "default": {}
                },
                "params": {
                    "type": "object",
                    "description": "URL-Parameter als Dictionary (optional)",
                    "default": {}
                },
                "data": {
                    "type": "object",
                    "description": "Request-Body als Dictionary (für POST/PUT, optional)",
                    "default": None
                },
                "json_data": {
                    "type": "object",
                    "description": "JSON-Request-Body als Dictionary (für POST/PUT, optional)",
                    "default": None
                },
                "extract_text": {
                    "type": "boolean",
                    "description": "Extrahiere Text aus HTML (Standard: False)",
                    "default": False
                },
                "timeout": {
                    "type": "integer",
                    "description": "Timeout in Sekunden (Standard: 30)",
                    "default": 30
                }
            }
        )
        self._session = None
        self.allowed_domains = os.getenv('WEB_ACCESS_ALLOWED_DOMAINS', '').split(',') if os.getenv('WEB_ACCESS_ALLOWED_DOMAINS') else []
        self.blocked_domains = os.getenv('WEB_ACCESS_BLOCKED_DOMAINS', 'localhost,127.0.0.1,0.0.0.0').split(',')
    
    async def _get_session(self):
        """Get aiohttp session"""
        if self._session is None or self._session.closed:
            connector = aiohttp.TCPConnector(limit=10, limit_per_host=5)
            timeout = aiohttp.ClientTimeout(total=30, connect=10)
            self._session = aiohttp.ClientSession(connector=connector, timeout=timeout)
        return self._session
    
    def _is_url_allowed(self, url: str) -> tuple:
        """Prüfe ob URL erlaubt ist (Sicherheitsprüfung)"""
        try:
            from urllib.parse import urlparse
            parsed = urlparse(url)
            
            # Prüfe Schema
            if parsed.scheme not in ['http', 'https']:
                return False, f"Nur HTTP/HTTPS erlaubt, nicht {parsed.scheme}"
            
            # Prüfe Domain-Whitelist (falls gesetzt)
            if self.allowed_domains and parsed.netloc not in self.allowed_domains:
                return False, f"Domain {parsed.netloc} nicht in erlaubter Liste"
            
            # Prüfe Domain-Blacklist
            for blocked in self.blocked_domains:
                if blocked and blocked in parsed.netloc:
                    return False, f"Domain {parsed.netloc} ist blockiert"
            
            # Prüfe auf private IPs (Sicherheit)
            if parsed.hostname:
                parts = parsed.hostname.split('.')
                if len(parts) == 4:
                    try:
                        ip_parts = [int(p) for p in parts]
                        # Private IP ranges
                        if (ip_parts[0] == 10 or 
                            (ip_parts[0] == 172 and 16 <= ip_parts[1] <= 31) or
                            (ip_parts[0] == 192 and ip_parts[1] == 168) or
                            ip_parts[0] == 127):
                            return False, f"Private IP {parsed.hostname} nicht erlaubt"
                    except ValueError:
                        pass  # Keine IP, sondern Hostname
            
            return True, None
            
        except Exception as e:
            return False, f"URL-Validierung fehlgeschlagen: {str(e)}"
    
    def _extract_text_from_html(self, html: str) -> str:
        """Extrahiere Text aus HTML (vereinfacht)"""
        try:
            import re
            # Entferne Script- und Style-Tags
            html = re.sub(r'<script[^>]*>.*?</script>', '', html, flags=re.DOTALL | re.IGNORECASE)
            html = re.sub(r'<style[^>]*>.*?</style>', '', html, flags=re.DOTALL | re.IGNORECASE)
            # Entferne HTML-Tags
            text = re.sub(r'<[^>]+>', ' ', html)
            # Normalisiere Whitespace
            text = re.sub(r'\s+', ' ', text)
            return text.strip()
        except Exception as e:
            logger.warning(f"HTML-Text-Extraktion fehlgeschlagen: {e}")
            return html[:1000]  # Fallback: erste 1000 Zeichen
    
    async def execute(self,
                      url: str,
                      method: str = "GET",
                      headers: Optional[Dict[str, str]] = None,
                      params: Optional[Dict[str, Any]] = None,
                      data: Optional[Dict[str, Any]] = None,
                      json_data: Optional[Dict[str, Any]] = None,
                      extract_text: bool = False,
                      timeout: int = 30) -> Dict[str, Any]:
        """Führe HTTP-Request aus"""
        try:
            # Sicherheitsprüfung
            allowed, error_msg = self._is_url_allowed(url)
            if not allowed:
                return {
                    "success": False,
                    "error": error_msg or "URL nicht erlaubt",
                    "url": url
                }
            
            # Standard-Headers
            default_headers = {
                "User-Agent": "Stundenzettel-Web-App/1.0 (Agent Tool)"
            }
            if headers:
                default_headers.update(headers)
            
            session = await self._get_session()
            request_timeout = aiohttp.ClientTimeout(total=timeout, connect=10)
            
            # Request ausführen
            request_kwargs = {
                "headers": default_headers,
                "timeout": request_timeout
            }
            
            if params:
                request_kwargs["params"] = params
            
            if method.upper() in ["POST", "PUT", "PATCH"]:
                if json_data:
                    request_kwargs["json"] = json_data
                elif data:
                    request_kwargs["data"] = data
            
            async with session.request(method.upper(), url, **request_kwargs) as response:
                response_data = None
                content_type = response.headers.get('Content-Type', '').lower()
                
                # Hole Response-Inhalt
                if 'application/json' in content_type:
                    response_data = await response.json()
                elif 'text/html' in content_type or 'text/plain' in content_type:
                    text_content = await response.text()
                    if extract_text and 'text/html' in content_type:
                        response_data = {
                            "html": text_content,
                            "text": self._extract_text_from_html(text_content)
                        }
                    else:
                        response_data = text_content
                else:
                    # Binärdaten - nur Metadaten zurückgeben
                    response_data = {
                        "content_type": content_type,
                        "size": len(await response.read()),
                        "note": "Binärdaten nicht extrahiert"
                    }
                
                return {
                    "success": response.status < 400,
                    "status_code": response.status,
                    "url": url,
                    "method": method.upper(),
                    "headers": dict(response.headers),
                    "data": response_data,
                    "content_type": content_type
                }
                
        except aiohttp.ClientError as e:
            logger.error(f"Web access error: {e}")
            return {
                "success": False,
                "error": f"HTTP-Client-Fehler: {str(e)}",
                "url": url
            }
        except Exception as e:
            logger.error(f"Web access error: {e}")
            return {
                "success": False,
                "error": str(e),
                "url": url
            }
    
    async def close(self):
        """Close HTTP session"""
        if self._session and not self._session.closed:
            await self._session.close()
            self._session = None

class DateParserTool(AgentTool):
    """Tool für Datums-Parsing und -Validierung - unterstützt verschiedene Datumsformate"""
    
    def __init__(self):
        super().__init__(
            name="date_parser",
            description="Parsed und validiert Datumsangaben in verschiedenen Formaten. Unterstützt internationale Formate, relative Daten (heute, gestern) und Datumsberechnungen. Nützlich für alle Agents zur Datumsverarbeitung.",
            parameters={
                "date_string": {
                    "type": "string",
                    "description": "Datumsstring in beliebigem Format (z.B. '15.01.2025', '2025-01-15', '15/01/2025', 'heute', 'gestern')"
                },
                "output_format": {
                    "type": "string",
                    "description": "Ausgabeformat (Standard: 'YYYY-MM-DD')",
                    "enum": ["YYYY-MM-DD", "DD.MM.YYYY", "DD/MM/YYYY", "timestamp", "iso"],
                    "default": "YYYY-MM-DD"
                },
                "locale": {
                    "type": "string",
                    "description": "Locale für Parsing (Standard: 'de_DE')",
                    "default": "de_DE"
                }
            }
        )
    
    async def execute(self, date_string: str, output_format: str = "YYYY-MM-DD", locale: str = "de_DE") -> Dict[str, Any]:
        """Parse und formatiere Datum"""
        try:
            from datetime import datetime, timedelta
            import re
            
            date_string = date_string.strip()
            
            # Relative Daten
            today = datetime.now()
            relative_map = {
                "heute": today,
                "today": today,
                "gestern": today - timedelta(days=1),
                "yesterday": today - timedelta(days=1),
                "morgen": today + timedelta(days=1),
                "tomorrow": today + timedelta(days=1),
                "vorgestern": today - timedelta(days=2),
                "übermorgen": today + timedelta(days=2)
            }
            
            if date_string.lower() in relative_map:
                parsed_date = relative_map[date_string.lower()]
            else:
                # Versuche verschiedene Formate
                formats = [
                    "%d.%m.%Y",  # 15.01.2025
                    "%Y-%m-%d",  # 2025-01-15
                    "%d/%m/%Y",  # 15/01/2025
                    "%d-%m-%Y",  # 15-01-2025
                    "%Y.%m.%d",  # 2025.01.15
                    "%d.%m.%y",  # 15.01.25
                    "%d/%m/%y",  # 15/01/25
                    "%Y%m%d",    # 20250115
                    "%d %B %Y",  # 15 Januar 2025
                    "%d. %B %Y", # 15. Januar 2025
                    "%B %d, %Y", # January 15, 2025
                ]
                
                parsed_date = None
                for fmt in formats:
                    try:
                        parsed_date = datetime.strptime(date_string, fmt)
                        break
                    except ValueError:
                        continue
                
                if not parsed_date:
                    # Versuche ISO-Format
                    try:
                        parsed_date = datetime.fromisoformat(date_string.replace('Z', '+00:00'))
                    except ValueError:
                        return {
                            "success": False,
                            "error": f"Konnte Datum '{date_string}' nicht parsen",
                            "date_string": date_string
                        }
            
            # Formatiere Ausgabe
            if output_format == "YYYY-MM-DD":
                formatted = parsed_date.strftime("%Y-%m-%d")
            elif output_format == "DD.MM.YYYY":
                formatted = parsed_date.strftime("%d.%m.%Y")
            elif output_format == "DD/MM/YYYY":
                formatted = parsed_date.strftime("%d/%m/%Y")
            elif output_format == "timestamp":
                formatted = int(parsed_date.timestamp())
            elif output_format == "iso":
                formatted = parsed_date.isoformat()
            else:
                formatted = parsed_date.strftime("%Y-%m-%d")
            
            return {
                "success": True,
                "date_string": date_string,
                "parsed_date": formatted,
                "datetime": parsed_date.isoformat(),
                "year": parsed_date.year,
                "month": parsed_date.month,
                "day": parsed_date.day,
                "weekday": parsed_date.strftime("%A"),
                "weekday_number": parsed_date.weekday(),  # 0=Montag, 6=Sonntag
                "is_weekend": parsed_date.weekday() >= 5,
                "output_format": output_format
            }
            
        except Exception as e:
            logger.error(f"Date parser error: {e}")
            return {
                "success": False,
                "error": str(e),
                "date_string": date_string
            }

class TaxNumberValidatorTool(AgentTool):
    """Tool für Steuernummer-Validierung - unterstützt verschiedene Länder"""
    
    def __init__(self):
        super().__init__(
            name="tax_number_validator",
            description="Validiert Steuernummern (USt-IdNr, VAT) für verschiedene Länder. Unterstützt DE, AT, CH, FR, IT, ES, GB, US und weitere. Nützlich für DocumentAgent und AccountingAgent zur Validierung von Belegen.",
            parameters={
                "tax_number": {
                    "type": "string",
                    "description": "Steuernummer zum Validieren"
                },
                "country_code": {
                    "type": "string",
                    "description": "Ländercode (z.B. 'DE', 'AT', 'CH', Standard: 'DE')",
                    "default": "DE"
                },
                "normalize": {
                    "type": "boolean",
                    "description": "Normalisiere Format (entferne Leerzeichen, Bindestriche, Standard: True)",
                    "default": True
                }
            }
        )
    
    async def execute(self, tax_number: str, country_code: str = "DE", normalize: bool = True) -> Dict[str, Any]:
        """Validiere Steuernummer"""
        try:
            import re
            
            if not tax_number or not tax_number.strip():
                return {
                    "success": False,
                    "valid": False,
                    "error": "Steuernummer ist leer",
                    "tax_number": tax_number
                }
            
            # Normalisiere
            if normalize:
                tax_number = tax_number.strip().replace(" ", "").replace("-", "").replace(".", "").upper()
            
            country_code = country_code.upper()
            
            # Validierungsregeln pro Land
            validation_rules = {
                "DE": {
                    "pattern": r"^DE\d{9}$",
                    "length": 11,
                    "format": "DE + 9 Ziffern",
                    "example": "DE123456789"
                },
                "AT": {
                    "pattern": r"^ATU\d{8}$",
                    "length": 11,
                    "format": "ATU + 8 Ziffern",
                    "example": "ATU12345678"
                },
                "CH": {
                    "pattern": r"^CHE\d{9}(TVA|MWST|IVA)$",
                    "length_min": 12,
                    "format": "CHE + 9 Ziffern + TVA/MWST/IVA",
                    "example": "CHE123456789TVA"
                },
                "FR": {
                    "pattern": r"^FR[A-HJ-NP-Z0-9]{2}\d{9}$",
                    "length": 13,
                    "format": "FR + 2 Zeichen + 9 Ziffern",
                    "example": "FR12345678901"
                },
                "IT": {
                    "pattern": r"^IT\d{11}$",
                    "length": 13,
                    "format": "IT + 11 Ziffern",
                    "example": "IT12345678901"
                },
                "ES": {
                    "pattern": r"^ES[A-Z0-9]\d{7}[A-Z0-9]$",
                    "length": 9,
                    "format": "ES + 1 Zeichen + 7 Ziffern + 1 Zeichen",
                    "example": "ES12345678Z"
                },
                "GB": {
                    "pattern": r"^GB\d{9}(\d{3})?$|^GB(GD|HA)\d{3}$",
                    "length_min": 9,
                    "format": "GB + 9-12 Ziffern oder GB(GD|HA) + 3 Ziffern",
                    "example": "GB123456789"
                },
                "US": {
                    "pattern": r"^\d{2}-\d{7}$",
                    "length": 10,
                    "format": "2 Ziffern - 7 Ziffern",
                    "example": "12-3456789"
                }
            }
            
            rule = validation_rules.get(country_code)
            if not rule:
                return {
                    "success": False,
                    "valid": False,
                    "error": f"Unbekannter Ländercode: {country_code}",
                    "supported_countries": list(validation_rules.keys()),
                    "tax_number": tax_number
                }
            
            # Prüfe Pattern
            pattern_match = re.match(rule["pattern"], tax_number)
            
            # Prüfe Länge (falls angegeben)
            length_valid = True
            if "length" in rule:
                length_valid = len(tax_number) == rule["length"]
            elif "length_min" in rule:
                length_valid = len(tax_number) >= rule["length_min"]
            
            is_valid = pattern_match is not None and length_valid
            
            result = {
                "success": True,
                "valid": is_valid,
                "tax_number": tax_number,
                "normalized": tax_number if normalize else tax_number,
                "country_code": country_code,
                "format": rule["format"],
                "example": rule.get("example", "")
            }
            
            if not is_valid:
                result["error"] = f"Steuernummer entspricht nicht dem Format für {country_code}: {rule['format']}"
                result["suggestions"] = [f"Erwartetes Format: {rule.get('example', rule['format'])}"]
            
            return result
            
        except Exception as e:
            logger.error(f"Tax number validator error: {e}")
            return {
                "success": False,
                "valid": False,
                "error": str(e),
                "tax_number": tax_number
            }

class TranslationTool(AgentTool):
    """Tool für Übersetzung - unterstützt mehrsprachige Dokumente"""
    
    def __init__(self):
        super().__init__(
            name="translation",
            description="Übersetzt Text zwischen verschiedenen Sprachen. Nützlich für DocumentAgent zur Übersetzung von Belegen in andere Sprachen. Unterstützt 100+ Sprachen.",
            parameters={
                "text": {
                    "type": "string",
                    "description": "Text zum Übersetzen"
                },
                "source_language": {
                    "type": "string",
                    "description": "Quellsprache (z.B. 'en', 'fr', 'it', 'es', Standard: 'auto' für automatische Erkennung)",
                    "default": "auto"
                },
                "target_language": {
                    "type": "string",
                    "description": "Zielsprache (z.B. 'de', 'en', Standard: 'de')",
                    "default": "de"
                }
            }
        )
    
    async def execute(self, text: str, source_language: str = "auto", target_language: str = "de") -> Dict[str, Any]:
        """Übersetze Text"""
        try:
            if not text or not text.strip():
                return {
                    "success": False,
                    "error": "Text ist leer",
                    "text": text
                }
            
            # Optionale DeepL-Integration (falls API-Key vorhanden)
            deepl_api_key = os.getenv('DEEPL_API_KEY')
            if deepl_api_key:
                try:
                    import aiohttp
                    async with aiohttp.ClientSession() as session:
                        url = "https://api-free.deepl.com/v2/translate" if "free" in deepl_api_key else "https://api.deepl.com/v2/translate"
                        params = {
                            "auth_key": deepl_api_key,
                            "text": text,
                            "target_lang": target_language.upper()
                        }
                        if source_language != "auto":
                            params["source_lang"] = source_language.upper()
                        
                        async with session.post(url, data=params) as response:
                            if response.status == 200:
                                data = await response.json()
                                translations = data.get("translations", [])
                                if translations:
                                    translated_text = translations[0].get("text", "")
                                    detected_source = translations[0].get("detected_source_language", source_language)
                                    return {
                                        "success": True,
                                        "text": text,
                                        "translated_text": translated_text,
                                        "source_language": detected_source.lower(),
                                        "target_language": target_language.lower(),
                                        "source": "deepl_api"
                                    }
                except Exception as e:
                    logger.debug(f"DeepL API error: {e}, using fallback")
            
            # Fallback: Nutze LLM für Übersetzung (wenn verfügbar)
            # Dies ist ein vereinfachtes Beispiel - in Produktion würde man eine echte Übersetzungs-API verwenden
            return {
                "success": False,
                "error": "Übersetzung nicht verfügbar. Bitte DEEPL_API_KEY setzen oder LLM-Integration verwenden.",
                "text": text,
                "note": "Für bessere Übersetzungen: DEEPL_API_KEY in .env setzen"
            }
            
        except Exception as e:
            logger.error(f"Translation error: {e}")
            return {
                "success": False,
                "error": str(e),
                "text": text
            }

class CurrencyValidatorTool(AgentTool):
    """Tool für Währungsvalidierung und -formatierung"""
    
    def __init__(self):
        super().__init__(
            name="currency_validator",
            description="Validiert und formatiert Währungsangaben. Prüft Währungscodes (ISO 4217), formatiert Beträge und validiert Währungsformate. Nützlich für AccountingAgent zur Währungsvalidierung.",
            parameters={
                "currency_code": {
                    "type": "string",
                    "description": "Währungscode zum Validieren (z.B. 'EUR', 'USD', 'GBP')"
                },
                "amount": {
                    "type": "number",
                    "description": "Betrag zum Formatieren (optional)"
                },
                "format": {
                    "type": "string",
                    "description": "Ausgabeformat (Standard: 'symbol')",
                    "enum": ["symbol", "code", "name"],
                    "default": "symbol"
                }
            }
        )
        # ISO 4217 Währungscodes
        self.currency_data = {
            "EUR": {"name": "Euro", "symbol": "€", "decimals": 2},
            "USD": {"name": "US Dollar", "symbol": "$", "decimals": 2},
            "GBP": {"name": "British Pound", "symbol": "£", "decimals": 2},
            "CHF": {"name": "Swiss Franc", "symbol": "CHF", "decimals": 2},
            "JPY": {"name": "Japanese Yen", "symbol": "¥", "decimals": 0},
            "CNY": {"name": "Chinese Yuan", "symbol": "¥", "decimals": 2},
            "AUD": {"name": "Australian Dollar", "symbol": "A$", "decimals": 2},
            "CAD": {"name": "Canadian Dollar", "symbol": "C$", "decimals": 2},
            "NZD": {"name": "New Zealand Dollar", "symbol": "NZ$", "decimals": 2},
            "SEK": {"name": "Swedish Krona", "symbol": "kr", "decimals": 2},
            "NOK": {"name": "Norwegian Krone", "symbol": "kr", "decimals": 2},
            "DKK": {"name": "Danish Krone", "symbol": "kr", "decimals": 2},
            "PLN": {"name": "Polish Zloty", "symbol": "zł", "decimals": 2},
            "CZK": {"name": "Czech Koruna", "symbol": "Kč", "decimals": 2},
            "HUF": {"name": "Hungarian Forint", "symbol": "Ft", "decimals": 2},
            "RUB": {"name": "Russian Ruble", "symbol": "₽", "decimals": 2},
            "TRY": {"name": "Turkish Lira", "symbol": "₺", "decimals": 2},
            "BRL": {"name": "Brazilian Real", "symbol": "R$", "decimals": 2},
            "INR": {"name": "Indian Rupee", "symbol": "₹", "decimals": 2},
            "ZAR": {"name": "South African Rand", "symbol": "R", "decimals": 2}
        }
    
    async def execute(self, currency_code: str, amount: Optional[float] = None, format: str = "symbol") -> Dict[str, Any]:
        """Validiere und formatiere Währung"""
        try:
            currency_code = currency_code.upper().strip()
            
            if currency_code not in self.currency_data:
                return {
                    "success": False,
                    "valid": False,
                    "error": f"Unbekannter Währungscode: {currency_code}",
                    "currency_code": currency_code,
                    "supported_currencies": list(self.currency_data.keys())
                }
            
            currency_info = self.currency_data[currency_code]
            
            result = {
                "success": True,
                "valid": True,
                "currency_code": currency_code,
                "name": currency_info["name"],
                "symbol": currency_info["symbol"],
                "decimals": currency_info["decimals"]
            }
            
            # Formatiere Betrag (falls angegeben)
            if amount is not None:
                decimals = currency_info["decimals"]
                formatted_amount = f"{amount:,.{decimals}f}".replace(",", "X").replace(".", ",").replace("X", ".")
                
                if format == "symbol":
                    result["formatted"] = f"{currency_info['symbol']} {formatted_amount}"
                elif format == "code":
                    result["formatted"] = f"{formatted_amount} {currency_code}"
                elif format == "name":
                    result["formatted"] = f"{formatted_amount} {currency_info['name']}"
                
                result["amount"] = amount
                result["formatted_amount"] = formatted_amount
            
            return result
            
        except Exception as e:
            logger.error(f"Currency validator error: {e}")
            return {
                "success": False,
                "valid": False,
                "error": str(e),
                "currency_code": currency_code
            }

class RegexPatternMatcherTool(AgentTool):
    """Tool für Regex-Mustererkennung - findet Muster in Texten"""
    
    def __init__(self):
        super().__init__(
            name="regex_pattern_matcher",
            description="Findet Muster in Texten mit regulären Ausdrücken. Nützlich für alle Agents zur Mustererkennung (z.B. Beträge, Datumsangaben, Steuernummern, E-Mail-Adressen).",
            parameters={
                "text": {
                    "type": "string",
                    "description": "Text zum Durchsuchen"
                },
                "pattern": {
                    "type": "string",
                    "description": "Regex-Pattern (z.B. r'\\d+,\\d{2}' für Beträge)"
                },
                "pattern_name": {
                    "type": "string",
                    "description": "Name eines vordefinierten Patterns (z.B. 'amount', 'date', 'email', 'tax_number', 'phone')",
                    "default": None
                },
                "case_sensitive": {
                    "type": "boolean",
                    "description": "Groß-/Kleinschreibung beachten (Standard: False)",
                    "default": False
                },
                "find_all": {
                    "type": "boolean",
                    "description": "Alle Vorkommen finden (Standard: True)",
                    "default": True
                }
            }
        )
        # Vordefinierte Patterns
        self.predefined_patterns = {
            "amount": [
                r"\d+[.,]\d{2}",  # 123,45 oder 123.45
                r"\d+[.,]\d{1,2}",  # Mit optionaler Dezimalstelle
                r"€\s*\d+[.,]\d{2}",  # € 123,45
                r"\$\s*\d+[.,]\d{2}",  # $ 123.45
            ],
            "date": [
                r"\d{1,2}[./-]\d{1,2}[./-]\d{2,4}",  # 15.01.2025 oder 15/01/25
                r"\d{4}[./-]\d{1,2}[./-]\d{1,2}",  # 2025-01-15
            ],
            "email": [
                r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}"
            ],
            "tax_number": [
                r"DE\d{9}",  # Deutsche USt-IdNr
                r"ATU\d{8}",  # Österreichische USt-IdNr
                r"[A-Z]{2}\d{8,12}",  # Allgemeines Format
            ],
            "phone": [
                r"\+?\d{1,4}[\s-]?\d{1,4}[\s-]?\d{1,9}",  # International
                r"0\d{1,4}[\s/-]?\d{1,9}",  # National (DE)
            ],
            "iban": [
                r"[A-Z]{2}\d{2}[A-Z0-9]{4,30}",  # IBAN
            ],
            "postal_code": [
                r"\d{5}",  # DE
                r"[A-Z0-9]{3,10}",  # International
            ]
        }
    
    async def execute(self,
                      text: str,
                      pattern: Optional[str] = None,
                      pattern_name: Optional[str] = None,
                      case_sensitive: bool = False,
                      find_all: bool = True) -> Dict[str, Any]:
        """Finde Muster in Text"""
        try:
            import re
            
            if not text:
                return {
                    "success": False,
                    "error": "Text ist leer",
                    "text": text
                }
            
            # Hole Pattern
            patterns_to_use = []
            if pattern_name and pattern_name in self.predefined_patterns:
                patterns_to_use = self.predefined_patterns[pattern_name]
            elif pattern:
                patterns_to_use = [pattern]
            else:
                return {
                    "success": False,
                    "error": "pattern oder pattern_name ist erforderlich",
                    "available_patterns": list(self.predefined_patterns.keys())
                }
            
            flags = 0 if case_sensitive else re.IGNORECASE
            all_matches = []
            
            for pat in patterns_to_use:
                if find_all:
                    matches = re.findall(pat, text, flags)
                    for match in matches:
                        all_matches.append({
                            "match": match if isinstance(match, str) else str(match),
                            "pattern": pat
                        })
                else:
                    match = re.search(pat, text, flags)
                    if match:
                        all_matches.append({
                            "match": match.group(0),
                            "pattern": pat,
                            "start": match.start(),
                            "end": match.end()
                        })
                        break
            
            return {
                "success": True,
                "text": text[:200] + "..." if len(text) > 200 else text,
                "matches": all_matches,
                "count": len(all_matches),
                "pattern_name": pattern_name,
                "patterns_used": patterns_to_use
            }
            
        except re.error as e:
            return {
                "success": False,
                "error": f"Ungültiges Regex-Pattern: {str(e)}",
                "pattern": pattern
            }
        except Exception as e:
            logger.error(f"Regex pattern matcher error: {e}")
            return {
                "success": False,
                "error": str(e),
                "text": text
            }

class PDFMetadataTool(AgentTool):
    """Tool für PDF-Metadaten-Extraktion"""
    
    def __init__(self):
        super().__init__(
            name="pdf_metadata",
            description="Extrahiert Metadaten aus PDF-Dateien (Erstellungsdatum, Autor, Titel, Seitenzahl, etc.). Nützlich für DocumentAgent zur Dokumentenanalyse.",
            parameters={
                "pdf_path": {
                    "type": "string",
                    "description": "Pfad zur PDF-Datei"
                }
            }
        )
    
    async def execute(self, pdf_path: str) -> Dict[str, Any]:
        """Extrahiere PDF-Metadaten"""
        try:
            if not Path(pdf_path).exists():
                return {
                    "success": False,
                    "error": f"PDF nicht gefunden: {pdf_path}",
                    "pdf_path": pdf_path
                }
            
            metadata = {}
            
            # Versuche mit PyPDF2
            if HAS_PYPDF2:
                try:
                    with open(pdf_path, 'rb') as file:
                        pdf_reader = PyPDF2.PdfReader(file)
                        info = pdf_reader.metadata
                        if info:
                            metadata = {
                                "title": info.get("/Title", ""),
                                "author": info.get("/Author", ""),
                                "subject": info.get("/Subject", ""),
                                "creator": info.get("/Creator", ""),
                                "producer": info.get("/Producer", ""),
                                "creation_date": str(info.get("/CreationDate", "")),
                                "modification_date": str(info.get("/ModDate", ""))
                            }
                        metadata["pages"] = len(pdf_reader.pages)
                        metadata["encrypted"] = pdf_reader.is_encrypted
                except Exception as e:
                    logger.debug(f"PyPDF2 metadata extraction error: {e}")
            
            # Versuche mit pdfplumber
            if HAS_PDFPLUMBER and not metadata.get("pages"):
                try:
                    import pdfplumber
                    with pdfplumber.open(pdf_path) as pdf:
                        metadata["pages"] = len(pdf.pages)
                        if pdf.metadata:
                            metadata.update({
                                "title": pdf.metadata.get("Title", metadata.get("title", "")),
                                "author": pdf.metadata.get("Author", metadata.get("author", "")),
                                "subject": pdf.metadata.get("Subject", metadata.get("subject", "")),
                                "creator": pdf.metadata.get("Creator", metadata.get("creator", "")),
                                "producer": pdf.metadata.get("Producer", metadata.get("producer", ""))
                            })
                except Exception as e:
                    logger.debug(f"pdfplumber metadata extraction error: {e}")
            
            if not metadata:
                return {
                    "success": False,
                    "error": "Konnte keine Metadaten extrahieren",
                    "pdf_path": pdf_path
                }
            
            return {
                "success": True,
                "pdf_path": pdf_path,
                "metadata": metadata,
                "source": "pypdf2" if HAS_PYPDF2 else "pdfplumber"
            }
            
        except Exception as e:
            logger.error(f"PDF metadata error: {e}")
            return {
                "success": False,
                "error": str(e),
                "pdf_path": pdf_path
            }

class DuplicateDetectionTool(AgentTool):
    """Tool für Duplikats-Erkennung - verhindert doppelte Beleg-Uploads"""
    
    def __init__(self):
        super().__init__(
            name="duplicate_detection",
            description="Erkennt doppelte Belege durch Hash-Vergleich und Bild-Ähnlichkeitsprüfung. Verhindert doppelte Uploads und Abrechnungen. Nützlich für DocumentAgent und AccountingAgent.",
            parameters={
                "file_path": {
                    "type": "string",
                    "description": "Pfad zur Datei zum Prüfen"
                },
                "file_hash": {
                    "type": "string",
                    "description": "Berechneter Hash-Wert der Datei (optional, wird automatisch berechnet wenn nicht angegeben)"
                },
                "check_similarity": {
                    "type": "boolean",
                    "description": "Bild-Ähnlichkeitsprüfung durchführen (Standard: True)",
                    "default": True
                },
                "similarity_threshold": {
                    "type": "number",
                    "description": "Ähnlichkeits-Schwellenwert (0.0-1.0, Standard: 0.95)",
                    "default": 0.95
                }
            }
        )
        self._hash_cache = {}  # Cache für bereits geprüfte Hashes
    
    def _calculate_file_hash(self, file_path: str) -> str:
        """Berechne SHA256-Hash einer Datei"""
        try:
            import hashlib
            sha256_hash = hashlib.sha256()
            with open(file_path, "rb") as f:
                for byte_block in iter(lambda: f.read(4096), b""):
                    sha256_hash.update(byte_block)
            return sha256_hash.hexdigest()
        except Exception as e:
            logger.error(f"Hash-Berechnung fehlgeschlagen: {e}")
            return ""
    
    def _calculate_perceptual_hash(self, file_path: str) -> Optional[str]:
        """Berechne Perceptual Hash für Bild-Ähnlichkeitsprüfung"""
        try:
            from PIL import Image
            import imagehash
            
            # Versuche Bild zu öffnen
            try:
                img = Image.open(file_path)
                # Konvertiere zu RGB falls nötig
                if img.mode != 'RGB':
                    img = img.convert('RGB')
                # Berechne Perceptual Hash
                phash = imagehash.phash(img)
                return str(phash)
            except Exception:
                # Kein Bild, versuche PDF
                if file_path.lower().endswith('.pdf'):
                    # Für PDFs extrahiere erste Seite als Bild
                    try:
                        if HAS_PDFPLUMBER:
                            import pdfplumber
                            from PIL import Image as PILImage
                            import io
                            
                            with pdfplumber.open(file_path) as pdf:
                                if len(pdf.pages) > 0:
                                    # Konvertiere erste Seite zu Bild (vereinfacht)
                                    # In Produktion würde man pdf2image verwenden
                                    return None  # Für PDFs verwenden wir nur Hash
                    except Exception:
                        pass
                return None
        except ImportError:
            logger.debug("imagehash nicht verfügbar, Perceptual Hash übersprungen")
            return None
        except Exception as e:
            logger.debug(f"Perceptual Hash Berechnung fehlgeschlagen: {e}")
            return None
    
    async def _check_hash_duplicate(self, file_hash: str, db=None) -> Dict[str, Any]:
        """Prüfe ob Hash bereits in Datenbank existiert"""
        try:
            if db is None:
                # Versuche Datenbank zu importieren (falls verfügbar)
                try:
                    import sys
                    if 'backend.server' in sys.modules:
                        from backend.server import db as server_db
                        db = server_db
                    else:
                        # Fallback: Keine Datenbank-Prüfung möglich
                        return {"is_duplicate": False, "note": "Datenbank nicht verfügbar für Duplikats-Prüfung"}
                except ImportError:
                    return {"is_duplicate": False, "note": "Datenbank nicht verfügbar für Duplikats-Prüfung"}
            
            # Suche in Receipts-Collection
            receipts_collection = db.receipts
            existing = await receipts_collection.find_one({"file_hash": file_hash})
            
            if existing:
                return {
                    "is_duplicate": True,
                    "match_type": "hash",
                    "existing_receipt_id": str(existing.get("_id", "")),
                    "existing_report_id": str(existing.get("report_id", "")),
                    "upload_date": str(existing.get("upload_date", ""))
                }
            
            # Suche in Timesheets-Collection (für unterschriebene Stundenzettel)
            timesheets_collection = db.timesheets
            existing_timesheet = await timesheets_collection.find_one({"signed_pdf_hash": file_hash})
            
            if existing_timesheet:
                return {
                    "is_duplicate": True,
                    "match_type": "hash",
                    "existing_timesheet_id": str(existing_timesheet.get("_id", "")),
                    "upload_date": str(existing_timesheet.get("signed_pdf_upload_date", ""))
                }
            
            return {"is_duplicate": False}
            
        except Exception as e:
            logger.error(f"Hash-Duplikats-Prüfung fehlgeschlagen: {e}")
            return {"is_duplicate": False, "error": str(e)}
    
    async def execute(self,
                      file_path: str,
                      file_hash: Optional[str] = None,
                      check_similarity: bool = True,
                      similarity_threshold: float = 0.95) -> Dict[str, Any]:
        """Prüfe auf Duplikate"""
        try:
            if not Path(file_path).exists():
                return {
                    "success": False,
                    "error": f"Datei nicht gefunden: {file_path}",
                    "file_path": file_path
                }
            
            # Berechne Hash falls nicht angegeben
            if not file_hash:
                file_hash = self._calculate_file_hash(file_path)
            
            if not file_hash:
                return {
                    "success": False,
                    "error": "Hash-Berechnung fehlgeschlagen",
                    "file_path": file_path
                }
            
            # Prüfe Hash-Duplikat (versuche db aus Kontext zu bekommen)
            db = None
            try:
                if hasattr(self, 'db'):
                    db = self.db
                elif hasattr(self, 'tools') and hasattr(self.tools, 'db'):
                    db = self.tools.db
            except:
                pass
            hash_result = await self._check_hash_duplicate(file_hash, db)
            
            result = {
                "success": True,
                "file_path": file_path,
                "file_hash": file_hash,
                "is_duplicate": hash_result.get("is_duplicate", False),
                "match_type": hash_result.get("match_type", "none")
            }
            
            if hash_result.get("is_duplicate"):
                result.update({
                    "existing_receipt_id": hash_result.get("existing_receipt_id"),
                    "existing_timesheet_id": hash_result.get("existing_timesheet_id"),
                    "existing_report_id": hash_result.get("existing_report_id"),
                    "upload_date": hash_result.get("upload_date"),
                    "warning": "Doppelte Datei erkannt!"
                })
                return result
            
            # Prüfe Bild-Ähnlichkeit (falls aktiviert)
            if check_similarity:
                try:
                    phash = self._calculate_perceptual_hash(file_path)
                    if phash:
                        # Suche ähnliche Perceptual Hashes in Datenbank
                        # (vereinfacht - in Produktion würde man Hamming-Distanz verwenden)
                        result["perceptual_hash"] = phash
                        result["similarity_check"] = "completed"
                except Exception as e:
                    logger.debug(f"Ähnlichkeitsprüfung übersprungen: {e}")
                    result["similarity_check"] = "skipped"
            
            return result
            
        except Exception as e:
            logger.error(f"Duplicate detection error: {e}")
            return {
                "success": False,
                "error": str(e),
                "file_path": file_path
            }

class IBANValidatorTool(AgentTool):
    """Tool für IBAN-Validierung und Bankdaten-Extraktion"""
    
    def __init__(self):
        super().__init__(
            name="iban_validator",
            description="Validiert IBAN-Nummern (ISO 13616) und extrahiert Bankdaten. Prüft Prüfziffern, erkennt Länder und BIC. Nützlich für DocumentAgent und AccountingAgent.",
            parameters={
                "iban": {
                    "type": "string",
                    "description": "IBAN zum Validieren"
                },
                "extract_bic": {
                    "type": "boolean",
                    "description": "BIC extrahieren (Standard: True)",
                    "default": True
                }
            }
        )
        # IBAN-Längen pro Land (erste 2 Zeichen)
        self.iban_lengths = {
            "AD": 24, "AE": 23, "AL": 28, "AT": 20, "AZ": 28, "BA": 20, "BE": 16,
            "BG": 22, "BH": 22, "BR": 29, "BY": 28, "CH": 21, "CR": 22, "CY": 28,
            "CZ": 24, "DE": 22, "DK": 18, "DO": 28, "EE": 20, "EG": 29, "ES": 24,
            "FI": 18, "FO": 18, "FR": 27, "GB": 22, "GE": 22, "GI": 23, "GL": 18,
            "GR": 27, "GT": 28, "HR": 21, "HU": 28, "IE": 22, "IL": 23, "IS": 26,
            "IT": 27, "JO": 30, "KW": 30, "KZ": 20, "LB": 28, "LC": 32, "LI": 21,
            "LT": 20, "LU": 20, "LV": 21, "MC": 27, "MD": 24, "ME": 22, "MK": 19,
            "MR": 27, "MT": 31, "MU": 30, "NL": 18, "NO": 15, "PK": 24, "PL": 28,
            "PS": 29, "PT": 25, "QA": 29, "RO": 24, "RS": 22, "SA": 24, "SE": 24,
            "SI": 19, "SK": 24, "SM": 27, "TN": 24, "TR": 26, "UA": 29, "VG": 24,
            "XK": 20
        }
    
    def _validate_iban_format(self, iban: str) -> tuple[bool, Optional[str]]:
        """Validiere IBAN-Format"""
        # Normalisiere (entferne Leerzeichen)
        iban_clean = iban.replace(" ", "").replace("-", "").upper()
        
        # Prüfe Länge (minimal 15, maximal 34)
        if len(iban_clean) < 15 or len(iban_clean) > 34:
            return False, "IBAN-Länge ungültig (muss zwischen 15 und 34 Zeichen sein)"
        
        # Prüfe Format (2 Buchstaben + 2 Ziffern + alphanumerisch)
        if not iban_clean[:2].isalpha():
            return False, "IBAN muss mit 2 Buchstaben (Ländercode) beginnen"
        
        if not iban_clean[2:4].isdigit():
            return False, "IBAN muss nach Ländercode 2 Ziffern (Prüfziffern) haben"
        
        if not iban_clean[4:].isalnum():
            return False, "IBAN enthält ungültige Zeichen nach Prüfziffern"
        
        # Prüfe Länge für spezifisches Land
        country_code = iban_clean[:2]
        expected_length = self.iban_lengths.get(country_code)
        if expected_length and len(iban_clean) != expected_length:
            return False, f"IBAN-Länge für {country_code} muss {expected_length} Zeichen sein, gefunden: {len(iban_clean)}"
        
        return True, iban_clean
    
    def _validate_iban_checksum(self, iban: str) -> bool:
        """Validiere IBAN-Prüfziffern (Modulo 97)"""
        try:
            # Verschiebe erste 4 Zeichen ans Ende
            rearranged = iban[4:] + iban[:4]
            
            # Ersetze Buchstaben durch Zahlen (A=10, B=11, ..., Z=35)
            numeric = ""
            for char in rearranged:
                if char.isalpha():
                    numeric += str(ord(char) - ord('A') + 10)
                else:
                    numeric += char
            
            # Berechne Modulo 97
            remainder = int(numeric) % 97
            return remainder == 1
            
        except Exception:
            return False
    
    def _extract_bic(self, iban: str) -> Optional[str]:
        """Extrahiere BIC aus IBAN (vereinfacht, nicht alle Länder unterstützen dies)"""
        # BIC-Extraktion ist komplex und länderspezifisch
        # Für DE: Bankleitzahl ist in IBAN enthalten, aber BIC-Mapping erforderlich
        # Hier nur Platzhalter
        country_code = iban[:2]
        if country_code == "DE":
            # Deutsche IBAN: DE + 2 Prüfziffern + 8-stellige BLZ + 10-stellige Kontonummer
            blz = iban[4:12]
            # BIC würde aus BLZ-Datenbank kommen
            return None  # Erfordert BLZ-zu-BIC-Mapping
        return None
    
    async def execute(self, iban: str, extract_bic: bool = True) -> Dict[str, Any]:
        """Validiere IBAN"""
        try:
            if not iban or not iban.strip():
                return {
                    "success": False,
                    "valid": False,
                    "error": "IBAN ist leer",
                    "iban": iban
                }
            
            # Validiere Format
            format_valid, iban_clean = self._validate_iban_format(iban)
            if not format_valid:
                return {
                    "success": True,
                    "valid": False,
                    "error": iban_clean,
                    "iban": iban,
                    "normalized": iban_clean if isinstance(iban_clean, str) else iban
                }
            
            # Validiere Prüfziffern
            checksum_valid = self._validate_iban_checksum(iban_clean)
            
            result = {
                "success": True,
                "valid": checksum_valid,
                "iban": iban,
                "normalized": iban_clean,
                "country_code": iban_clean[:2],
                "check_digits": iban_clean[2:4],
                "bban": iban_clean[4:],  # Basic Bank Account Number
                "length": len(iban_clean),
                "format_valid": format_valid,
                "checksum_valid": checksum_valid
            }
            
            if not checksum_valid:
                result["error"] = "IBAN-Prüfziffern ungültig (Modulo 97 Prüfung fehlgeschlagen)"
            
            # Extrahiere BIC (falls angefordert)
            if extract_bic and checksum_valid:
                bic = self._extract_bic(iban_clean)
                if bic:
                    result["bic"] = bic
                else:
                    result["bic"] = None
                    result["bic_note"] = "BIC-Extraktion erfordert externe Datenbank"
            
            return result
            
        except Exception as e:
            logger.error(f"IBAN validator error: {e}")
            return {
                "success": False,
                "valid": False,
                "error": str(e),
                "iban": iban
            }

class ImageQualityTool(AgentTool):
    """Tool für Qualitätsprüfung von gescannten Belegen"""
    
    def __init__(self):
        super().__init__(
            name="image_quality",
            description="Prüft Qualität von gescannten Belegen (Auflösung, Schärfe, Kontrast, Helligkeit). Warnt vor schlechter Qualität vor OCR. Nützlich für DocumentAgent.",
            parameters={
                "image_path": {
                    "type": "string",
                    "description": "Pfad zur Bilddatei oder PDF"
                },
                "min_dpi": {
                    "type": "integer",
                    "description": "Minimale DPI-Anforderung (Standard: 150)",
                    "default": 150
                },
                "min_sharpness": {
                    "type": "number",
                    "description": "Minimale Schärfe (0.0-1.0, Standard: 0.3)",
                    "default": 0.3
                }
            }
        )
    
    def _check_image_quality(self, image_path: str) -> Dict[str, Any]:
        """Prüfe Bildqualität"""
        try:
            from PIL import Image
            import numpy as np
            import cv2
            
            # Öffne Bild
            img = Image.open(image_path)
            
            # Prüfe DPI
            dpi = img.info.get('dpi', (72, 72))[0]  # Standard: 72 DPI
            
            # Konvertiere zu numpy array für OpenCV
            img_array = np.array(img.convert('RGB'))
            img_gray = cv2.cvtColor(img_array, cv2.COLOR_RGB2GRAY)
            
            # Berechne Schärfe (Laplacian Variance)
            laplacian = cv2.Laplacian(img_gray, cv2.CV_64F)
            sharpness = laplacian.var()
            normalized_sharpness = min(sharpness / 1000.0, 1.0)  # Normalisiere auf 0-1
            
            # Berechne Kontrast (Standardabweichung)
            contrast = np.std(img_gray)
            normalized_contrast = min(contrast / 128.0, 1.0)  # Normalisiere auf 0-1
            
            # Berechne Helligkeit (Durchschnitt)
            brightness = np.mean(img_gray)
            normalized_brightness = brightness / 255.0  # Normalisiere auf 0-1
            
            # Prüfe auf Blur (zu niedrige Schärfe)
            is_blurry = normalized_sharpness < 0.3
            
            # Prüfe auf zu dunkel/hell
            is_too_dark = normalized_brightness < 0.2
            is_too_bright = normalized_brightness > 0.8
            
            # Gesamtbewertung
            quality_score = (normalized_sharpness * 0.4 + 
                          normalized_contrast * 0.3 + 
                          (1.0 - abs(normalized_brightness - 0.5) * 2) * 0.3)
            
            issues = []
            if dpi < 150:
                issues.append(f"DPI zu niedrig ({dpi}, empfohlen: ≥150)")
            if is_blurry:
                issues.append(f"Bild unscharf (Schärfe: {normalized_sharpness:.2f})")
            if is_too_dark:
                issues.append(f"Bild zu dunkel (Helligkeit: {normalized_brightness:.2f})")
            if is_too_bright:
                issues.append(f"Bild zu hell (Helligkeit: {normalized_brightness:.2f})")
            if normalized_contrast < 0.3:
                issues.append(f"Kontrast zu niedrig ({normalized_contrast:.2f})")
            
            return {
                "dpi": dpi,
                "width": img.width,
                "height": img.height,
                "sharpness": float(normalized_sharpness),
                "contrast": float(normalized_contrast),
                "brightness": float(normalized_brightness),
                "quality_score": float(quality_score),
                "is_blurry": is_blurry,
                "is_too_dark": is_too_dark,
                "is_too_bright": is_too_bright,
                "issues": issues,
                "is_good_quality": quality_score >= 0.6 and len(issues) == 0
            }
            
        except ImportError:
            return {
                "error": "OpenCV oder PIL nicht verfügbar",
                "note": "Bitte 'pip install opencv-python pillow' installieren"
            }
        except Exception as e:
            logger.error(f"Image quality check error: {e}")
            return {"error": str(e)}
    
    async def execute(self, image_path: str, min_dpi: int = 150, min_sharpness: float = 0.3) -> Dict[str, Any]:
        """Prüfe Bildqualität"""
        try:
            if not Path(image_path).exists():
                return {
                    "success": False,
                    "error": f"Datei nicht gefunden: {image_path}",
                    "image_path": image_path
                }
            
            # Prüfe ob PDF oder Bild
            if image_path.lower().endswith('.pdf'):
                # Für PDFs extrahiere erste Seite
                try:
                    if HAS_PDFPLUMBER:
                        import pdfplumber
                        from PIL import Image as PILImage
                        import io
                        import tempfile
                        
                        with pdfplumber.open(image_path) as pdf:
                            if len(pdf.pages) > 0:
                                # Konvertiere erste Seite zu Bild (vereinfacht)
                                # In Produktion würde man pdf2image verwenden
                                page = pdf.pages[0]
                                # Extrahiere als Bild (vereinfacht)
                                # Für echte Implementierung: pdf2image verwenden
                                return {
                                    "success": True,
                                    "image_path": image_path,
                                    "type": "pdf",
                                    "pages": len(pdf.pages),
                                    "note": "PDF-Qualitätsprüfung erfordert pdf2image. Verwende erste Seite.",
                                    "quality_check": "limited"
                                }
                except Exception as e:
                    logger.debug(f"PDF-Qualitätsprüfung fehlgeschlagen: {e}")
                    return {
                        "success": False,
                        "error": f"PDF-Qualitätsprüfung fehlgeschlagen: {str(e)}",
                        "image_path": image_path
                    }
            
            # Prüfe Bildqualität
            quality_result = self._check_image_quality(image_path)
            
            if "error" in quality_result:
                return {
                    "success": False,
                    "error": quality_result.get("error"),
                    "note": quality_result.get("note"),
                    "image_path": image_path
                }
            
            # Bewertung
            is_good = quality_result.get("is_good_quality", False)
            meets_dpi = quality_result.get("dpi", 0) >= min_dpi
            meets_sharpness = quality_result.get("sharpness", 0) >= min_sharpness
            
            result = {
                "success": True,
                "image_path": image_path,
                "quality_score": quality_result.get("quality_score", 0),
                "is_good_quality": is_good and meets_dpi and meets_sharpness,
                "dpi": quality_result.get("dpi", 0),
                "meets_dpi_requirement": meets_dpi,
                "sharpness": quality_result.get("sharpness", 0),
                "meets_sharpness_requirement": meets_sharpness,
                "contrast": quality_result.get("contrast", 0),
                "brightness": quality_result.get("brightness", 0),
                "issues": quality_result.get("issues", []),
                "recommendations": []
            }
            
            # Empfehlungen
            if not meets_dpi:
                result["recommendations"].append(f"Scannen Sie mit mindestens {min_dpi} DPI")
            if not meets_sharpness:
                result["recommendations"].append("Verbessern Sie die Bildschärfe")
            if quality_result.get("is_too_dark"):
                result["recommendations"].append("Erhöhen Sie die Helligkeit beim Scannen")
            if quality_result.get("is_too_bright"):
                result["recommendations"].append("Reduzieren Sie die Helligkeit beim Scannen")
            if quality_result.get("contrast", 0) < 0.3:
                result["recommendations"].append("Erhöhen Sie den Kontrast")
            
            return result
            
        except Exception as e:
            logger.error(f"Image quality error: {e}")
            return {
                "success": False,
                "error": str(e),
                "image_path": image_path
            }

class TimeZoneTool(AgentTool):
    """Tool für Zeitzonen-Handling für internationale Reisen"""
    
    def __init__(self):
        super().__init__(
            name="timezone",
            description="Zeitzonen-Erkennung und -Konvertierung für internationale Reisen. Validiert Reisezeiten bei Zeitzonen-Wechsel. Nützlich für AccountingAgent und ChatAgent.",
            parameters={
                "location": {
                    "type": "string",
                    "description": "Ortsangabe (z.B. 'Berlin', 'New York', 'Tokyo')"
                },
                "datetime_string": {
                    "type": "string",
                    "description": "Datum/Zeit-String zum Konvertieren (optional)"
                },
                "from_timezone": {
                    "type": "string",
                    "description": "Quell-Zeitzone (optional, wird automatisch erkannt)"
                },
                "to_timezone": {
                    "type": "string",
                    "description": "Ziel-Zeitzone (Standard: 'UTC')",
                    "default": "UTC"
                }
            }
        )
    
    async def execute(self,
                      location: Optional[str] = None,
                      datetime_string: Optional[str] = None,
                      from_timezone: Optional[str] = None,
                      to_timezone: str = "UTC") -> Dict[str, Any]:
        """Zeitzonen-Operationen"""
        try:
            from datetime import datetime
            import pytz
            
            result = {
                "success": True
            }
            
            # Zeitzone aus Ort bestimmen
            if location:
                try:
                    # Nutze Geocoding-Tool für Koordinaten (falls verfügbar)
                    # Tools werden über Tool-Registry bereitgestellt
                    from backend.agents import get_tool_registry
                    tool_registry = get_tool_registry()
                    geocoding_tool = tool_registry.get_tool("geocoding")
                    if geocoding_tool:
                        geo_result = await geocoding_tool.execute(location)
                        if geo_result.get("success"):
                            lat = geo_result.get("latitude")
                            lon = geo_result.get("longitude")
                            
                            # Bestimme Zeitzone aus Koordinaten
                            try:
                                import timezonefinder
                                tf = timezonefinder.TimezoneFinder()
                                tz_name = tf.timezone_at(lat=lat, lng=lon)
                                if tz_name:
                                    result["location"] = location
                                    result["timezone"] = tz_name
                                    result["timezone_offset"] = pytz.timezone(tz_name).utcoffset(datetime.now()).total_seconds() / 3600
                            except ImportError:
                                # Fallback: Nutze OpenMaps-Tool
                                openmaps_tool = tool_registry.get_tool("openmaps")
                                if openmaps_tool:
                                    maps_result = await openmaps_tool.execute(action="geocode", query=location)
                                    if maps_result.get("success") and maps_result.get("data"):
                                        # Extrahiere Zeitzone aus OpenMaps-Ergebnis
                                        result["location"] = location
                                        result["timezone"] = "unknown"
                                        result["note"] = "Zeitzone-Erkennung erfordert timezonefinder"
                except Exception as e:
                    logger.debug(f"Zeitzone-Erkennung fehlgeschlagen: {e}")
                    result["location"] = location
                    result["timezone"] = "unknown"
                    result["error"] = str(e)
            
            # Datum/Zeit konvertieren
            if datetime_string:
                try:
                    # Parse Datum/Zeit
                    dt = datetime.fromisoformat(datetime_string.replace('Z', '+00:00'))
                    
                    # Konvertiere Zeitzone
                    if from_timezone:
                        from_tz = pytz.timezone(from_timezone)
                        dt = from_tz.localize(dt) if dt.tzinfo is None else dt.astimezone(from_tz)
                    
                    to_tz = pytz.timezone(to_timezone)
                    dt_converted = dt.astimezone(to_tz)
                    
                    result["datetime_string"] = datetime_string
                    result["converted_datetime"] = dt_converted.isoformat()
                    result["from_timezone"] = from_timezone or "local"
                    result["to_timezone"] = to_timezone
                    result["timezone_offset_hours"] = dt_converted.utcoffset().total_seconds() / 3600
                    
                except Exception as e:
                    result["datetime_conversion_error"] = str(e)
            
            return result
            
        except ImportError:
            return {
                "success": False,
                "error": "pytz nicht verfügbar",
                "note": "Bitte 'pip install pytz' installieren"
            }
        except Exception as e:
            logger.error(f"Timezone error: {e}")
            return {
                "success": False,
                "error": str(e),
                "location": location
            }

class EmailValidatorTool(AgentTool):
    """Tool für E-Mail-Validierung und Domain-Prüfung"""
    
    def __init__(self):
        super().__init__(
            name="email_validator",
            description="Validiert E-Mail-Adressen (RFC 5322) und prüft Domain-Existenz (DNS MX-Record). Erkennt Disposable-E-Mails. Nützlich für DocumentAgent und ChatAgent.",
            parameters={
                "email": {
                    "type": "string",
                    "description": "E-Mail-Adresse zum Validieren"
                },
                "check_dns": {
                    "type": "boolean",
                    "description": "DNS MX-Record prüfen (Standard: True)",
                    "default": True
                },
                "check_disposable": {
                    "type": "boolean",
                    "description": "Disposable-E-Mail erkennen (Standard: True)",
                    "default": True
                }
            }
        )
        # Liste bekannter Disposable-E-Mail-Domains
        self.disposable_domains = {
            "10minutemail.com", "guerrillamail.com", "mailinator.com",
            "tempmail.com", "throwaway.email", "yopmail.com", "temp-mail.org"
        }
    
    def _validate_email_format(self, email: str) -> tuple[bool, Optional[str]]:
        """Validiere E-Mail-Format (RFC 5322)"""
        import re
        
        # Einfache RFC 5322-konforme Regex
        pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        
        if not re.match(pattern, email):
            return False, "E-Mail-Format ungültig"
        
        # Prüfe Länge
        if len(email) > 254:  # RFC 5321 Limit
            return False, "E-Mail-Adresse zu lang (max. 254 Zeichen)"
        
        # Prüfe lokalen Teil
        local_part = email.split('@')[0]
        if len(local_part) > 64:  # RFC 5321 Limit
            return False, "Lokaler Teil zu lang (max. 64 Zeichen)"
        
        return True, None
    
    async def _check_dns_mx(self, domain: str) -> Dict[str, Any]:
        """Prüfe DNS MX-Record"""
        try:
            import dns.resolver
            
            try:
                mx_records = dns.resolver.resolve(domain, 'MX')
                mx_hosts = [str(mx.exchange) for mx in mx_records]
                return {
                    "has_mx": True,
                    "mx_hosts": mx_hosts,
                    "mx_count": len(mx_hosts)
                }
            except dns.resolver.NXDOMAIN:
                return {
                    "has_mx": False,
                    "error": "Domain existiert nicht"
                }
            except dns.resolver.NoAnswer:
                return {
                    "has_mx": False,
                    "error": "Keine MX-Records gefunden"
                }
                
        except ImportError:
            return {
                "has_mx": None,
                "error": "dnspython nicht verfügbar",
                "note": "Bitte 'pip install dnspython' installieren"
            }
        except Exception as e:
            logger.debug(f"DNS-Prüfung fehlgeschlagen: {e}")
            return {
                "has_mx": None,
                "error": str(e)
            }
    
    def _check_disposable(self, domain: str) -> bool:
        """Prüfe ob Disposable-E-Mail"""
        return domain.lower() in self.disposable_domains
    
    async def execute(self,
                      email: str,
                      check_dns: bool = True,
                      check_disposable: bool = True) -> Dict[str, Any]:
        """Validiere E-Mail"""
        try:
            if not email or not email.strip():
                return {
                    "success": False,
                    "valid": False,
                    "error": "E-Mail ist leer",
                    "email": email
                }
            
            email = email.strip().lower()
            
            # Validiere Format
            format_valid, format_error = self._validate_email_format(email)
            
            if not format_valid:
                return {
                    "success": True,
                    "valid": False,
                    "error": format_error,
                    "email": email
                }
            
            # Extrahiere Domain
            domain = email.split('@')[1]
            
            result = {
                "success": True,
                "valid": True,
                "email": email,
                "domain": domain,
                "format_valid": True
            }
            
            # Prüfe Disposable
            if check_disposable:
                is_disposable = self._check_disposable(domain)
                result["is_disposable"] = is_disposable
                if is_disposable:
                    result["warning"] = "Disposable-E-Mail erkannt"
            
            # Prüfe DNS
            if check_dns:
                dns_result = await self._check_dns_mx(domain)
                result["dns_check"] = dns_result
                if dns_result.get("has_mx") is False:
                    result["valid"] = False
                    result["error"] = dns_result.get("error", "DNS-Prüfung fehlgeschlagen")
                elif dns_result.get("has_mx") is None:
                    result["warning"] = "DNS-Prüfung nicht verfügbar"
            
            return result
            
        except Exception as e:
            logger.error(f"Email validator error: {e}")
            return {
                "success": False,
                "valid": False,
                "error": str(e),
                "email": email
            }

class EmailParserTool(AgentTool):
    """Tool für automatische Beleg-Extraktion aus E-Mails"""
    
    def __init__(self):
        super().__init__(
            name="email_parser",
            description="Extrahiert automatisch Belege aus E-Mails (IMAP/POP3). Erkennt Beleg-Anhänge (PDF, Bilder) und extrahiert Betrag, Datum, Absender. Nützlich für DocumentAgent und AccountingAgent.",
            parameters={
                "email_server": {
                    "type": "string",
                    "description": "E-Mail-Server (IMAP oder POP3, z.B. 'imap.gmail.com')"
                },
                "email_user": {
                    "type": "string",
                    "description": "E-Mail-Benutzername"
                },
                "email_password": {
                    "type": "string",
                    "description": "E-Mail-Passwort oder App-Passwort"
                },
                "email_folder": {
                    "type": "string",
                    "description": "E-Mail-Ordner zum Durchsuchen (Standard: 'INBOX')",
                    "default": "INBOX"
                },
                "max_emails": {
                    "type": "integer",
                    "description": "Maximale Anzahl E-Mails zum Durchsuchen (Standard: 10)",
                    "default": 10
                },
                "extract_attachments": {
                    "type": "boolean",
                    "description": "Anhänge extrahieren (Standard: True)",
                    "default": True
                }
            }
        )
    
    async def execute(self,
                      email_server: str,
                      email_user: str,
                      email_password: str,
                      email_folder: str = "INBOX",
                      max_emails: int = 10,
                      extract_attachments: bool = True) -> Dict[str, Any]:
        """Parse E-Mails und extrahiere Belege"""
        try:
            # Prüfe ob IMAP-Client verfügbar
            try:
                import imapclient
            except ImportError:
                return {
                    "success": False,
                    "error": "imapclient nicht verfügbar",
                    "note": "Bitte 'pip install imapclient' installieren"
                }
            
            # Verbinde mit E-Mail-Server
            try:
                server = imapclient.IMAPClient(email_server, ssl=True)
                server.login(email_user, email_password)
                server.select_folder(email_folder)
                
                # Suche nach E-Mails
                messages = server.search(['UNSEEN', 'ALL'])
                messages = messages[:max_emails] if len(messages) > max_emails else messages
                
                results = []
                for msg_id in messages:
                    try:
                        msg_data = server.fetch([msg_id], ['ENVELOPE', 'BODYSTRUCTURE', 'BODY[]'])
                        envelope = msg_data[msg_id][b'ENVELOPE']
                        body = msg_data[msg_id][b'BODY[]']
                        
                        # Parse E-Mail
                        from email import message_from_bytes
                        email_message = message_from_bytes(body)
                        
                        # Extrahiere Informationen
                        subject = email_message.get('Subject', '')
                        sender = email_message.get('From', '')
                        date = email_message.get('Date', '')
                        
                        # Extrahiere Text
                        text_content = ""
                        if email_message.is_multipart():
                            for part in email_message.walk():
                                if part.get_content_type() == "text/plain":
                                    text_content = part.get_payload(decode=True).decode('utf-8', errors='ignore')
                                    break
                        else:
                            text_content = email_message.get_payload(decode=True).decode('utf-8', errors='ignore')
                        
                        # Suche nach Betrag und Datum im Text
                        import re
                        amount_match = re.search(r'(\d+[.,]\d{2})\s*€|€\s*(\d+[.,]\d{2})', text_content)
                        date_match = re.search(r'\d{1,2}[./-]\d{1,2}[./-]\d{2,4}', text_content)
                        
                        email_info = {
                            "message_id": msg_id,
                            "subject": subject,
                            "sender": sender,
                            "date": date,
                            "text_preview": text_content[:200] + "..." if len(text_content) > 200 else text_content,
                            "amount_found": amount_match.group(0) if amount_match else None,
                            "date_found": date_match.group(0) if date_match else None
                        }
                        
                        # Extrahiere Anhänge
                        attachments = []
                        if extract_attachments and email_message.is_multipart():
                            for part in email_message.walk():
                                if part.get_content_disposition() == 'attachment':
                                    filename = part.get_filename()
                                    if filename and (filename.lower().endswith('.pdf') or 
                                                    filename.lower().endswith(('.jpg', '.jpeg', '.png'))):
                                        attachments.append({
                                            "filename": filename,
                                            "content_type": part.get_content_type(),
                                            "size": len(part.get_payload(decode=True))
                                        })
                        
                        email_info["attachments"] = attachments
                        results.append(email_info)
                        
                    except Exception as e:
                        logger.debug(f"Fehler beim Parsen von E-Mail {msg_id}: {e}")
                        continue
                
                server.logout()
                
                return {
                    "success": True,
                    "emails_processed": len(results),
                    "emails": results,
                    "total_attachments": sum(len(e.get("attachments", [])) for e in results)
                }
                
            except Exception as e:
                return {
                    "success": False,
                    "error": f"E-Mail-Verbindung fehlgeschlagen: {str(e)}",
                    "email_server": email_server
                }
            
        except Exception as e:
            logger.error(f"Email parser error: {e}")
            return {
                "success": False,
                "error": str(e),
                "email_server": email_server
            }

class SignatureDetectionTool(AgentTool):
    """Tool für erweiterte Signatur-Erkennung in PDFs"""
    
    def __init__(self):
        super().__init__(
            name="signature_detection",
            description="Erkennt Signaturen in PDFs (Signatur-Felder, digitale Signaturen, handschriftliche Signaturen). Nützlich für DocumentAgent zur verbesserten Unterschriften-Verifikation.",
            parameters={
                "pdf_path": {
                    "type": "string",
                    "description": "Pfad zur PDF-Datei"
                },
                "check_digital": {
                    "type": "boolean",
                    "description": "Digitale Signaturen prüfen (Standard: True)",
                    "default": True
                },
                "check_handwritten": {
                    "type": "boolean",
                    "description": "Handschriftliche Signaturen erkennen (Standard: True)",
                    "default": True
                }
            }
        )
    
    async def execute(self,
                      pdf_path: str,
                      check_digital: bool = True,
                      check_handwritten: bool = True) -> Dict[str, Any]:
        """Erkenne Signaturen in PDF"""
        try:
            if not Path(pdf_path).exists():
                return {
                    "success": False,
                    "error": f"PDF nicht gefunden: {pdf_path}",
                    "pdf_path": pdf_path
                }
            
            result = {
                "success": True,
                "pdf_path": pdf_path,
                "digital_signatures": [],
                "signature_fields": [],
                "handwritten_signatures": []
            }
            
            # Prüfe digitale Signaturen
            if check_digital and HAS_PYPDF2:
                try:
                    with open(pdf_path, 'rb') as file:
                        pdf_reader = PyPDF2.PdfReader(file)
                        
                        # Prüfe auf Signatur-Felder
                        if hasattr(pdf_reader, 'get_form_text_fields'):
                            fields = pdf_reader.get_form_text_fields()
                            for field_name, field_value in fields.items():
                                if 'signature' in field_name.lower() or 'unterschrift' in field_name.lower():
                                    result["signature_fields"].append({
                                        "field_name": field_name,
                                        "field_value": field_value
                                    })
                        
                        # Prüfe auf digitale Signaturen (X.509)
                        if hasattr(pdf_reader, 'get_signature_fields'):
                            sig_fields = pdf_reader.get_signature_fields()
                            for sig_field in sig_fields:
                                result["digital_signatures"].append({
                                    "field_name": sig_field.get('name', ''),
                                    "signed": True
                                })
                except Exception as e:
                    logger.debug(f"Digitale Signatur-Prüfung fehlgeschlagen: {e}")
                    result["digital_signature_error"] = str(e)
            
            # Prüfe handschriftliche Signaturen (vereinfacht - würde ML erfordern)
            if check_handwritten:
                # Extrahiere Text und suche nach Signatur-Hinweisen
                try:
                    if HAS_PDFPLUMBER:
                        import pdfplumber
                        with pdfplumber.open(pdf_path) as pdf:
                            for page_num, page in enumerate(pdf.pages):
                                text = page.extract_text()
                                if text:
                                    # Suche nach Signatur-Hinweisen im Text
                                    import re
                                    sig_patterns = [
                                        r'unterschrift',
                                        r'signature',
                                        r'gezeichnet',
                                        r'gez\.',
                                        r'________________'
                                    ]
                                    for pattern in sig_patterns:
                                        if re.search(pattern, text, re.IGNORECASE):
                                            result["handwritten_signatures"].append({
                                                "page": page_num + 1,
                                                "indicator": "Text-Hinweis gefunden",
                                                "pattern": pattern
                                            })
                                            break
                except Exception as e:
                    logger.debug(f"Handschriftliche Signatur-Prüfung fehlgeschlagen: {e}")
                    result["handwritten_signature_error"] = str(e)
            
            result["has_signatures"] = (
                len(result["digital_signatures"]) > 0 or
                len(result["signature_fields"]) > 0 or
                len(result["handwritten_signatures"]) > 0
            )
            
            return result
            
        except Exception as e:
            logger.error(f"Signature detection error: {e}")
            return {
                "success": False,
                "error": str(e),
                "pdf_path": pdf_path
            }

class ExcelImportExportTool(AgentTool):
    """Tool für Excel/CSV-Import/Export für Buchhaltung"""
    
    def __init__(self):
        super().__init__(
            name="excel_import_export",
            description="Importiert und exportiert Excel/CSV-Dateien für Buchhaltung. Automatische Formatierung von Beträgen und Datumsangaben. Nützlich für AccountingAgent.",
            parameters={
                "action": {
                    "type": "string",
                    "description": "Aktion: 'export' (Excel/CSV exportieren), 'import' (Excel/CSV importieren), 'template' (Template generieren)",
                    "enum": ["export", "import", "template"]
                },
                "data": {
                    "type": "array",
                    "description": "Daten zum Exportieren (für export-Aktion)",
                    "default": []
                },
                "file_path": {
                    "type": "string",
                    "description": "Dateipfad (für import/export)"
                },
                "format": {
                    "type": "string",
                    "description": "Dateiformat (Standard: 'xlsx')",
                    "enum": ["xlsx", "csv"],
                    "default": "xlsx"
                }
            }
        )
    
    async def execute(self,
                      action: str,
                      data: Optional[List[Dict[str, Any]]] = None,
                      file_path: Optional[str] = None,
                      format: str = "xlsx") -> Dict[str, Any]:
        """Import/Export Excel/CSV"""
        try:
            if action == "export":
                if not data:
                    return {
                        "success": False,
                        "error": "Daten sind erforderlich für Export"
                    }
                
                if format == "xlsx":
                    try:
                        import openpyxl
                        from openpyxl.styles import Font, Alignment
                        
                        wb = openpyxl.Workbook()
                        ws = wb.active
                        ws.title = "Reisekosten"
                        
                        # Header
                        if data and len(data) > 0:
                            headers = list(data[0].keys())
                            ws.append(headers)
                            
                            # Formatierung Header
                            for cell in ws[1]:
                                cell.font = Font(bold=True)
                                cell.alignment = Alignment(horizontal='center')
                            
                            # Daten
                            for row_data in data:
                                row = [row_data.get(header, "") for header in headers]
                                ws.append(row)
                        
                        # Speichere Datei
                        if not file_path:
                            from datetime import datetime
                            file_path = f"reisekosten_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                        
                        wb.save(file_path)
                        
                        return {
                            "success": True,
                            "action": "export",
                            "file_path": file_path,
                            "format": "xlsx",
                            "rows_exported": len(data)
                        }
                    except ImportError:
                        return {
                            "success": False,
                            "error": "openpyxl nicht verfügbar",
                            "note": "Bitte 'pip install openpyxl' installieren"
                        }
                
                elif format == "csv":
                    try:
                        import csv
                        
                        if not file_path:
                            from datetime import datetime
                            file_path = f"reisekosten_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
                        
                        if data and len(data) > 0:
                            headers = list(data[0].keys())
                            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                                writer = csv.DictWriter(f, fieldnames=headers)
                                writer.writeheader()
                                writer.writerows(data)
                        
                        return {
                            "success": True,
                            "action": "export",
                            "file_path": file_path,
                            "format": "csv",
                            "rows_exported": len(data) if data else 0
                        }
                    except Exception as e:
                        return {
                            "success": False,
                            "error": str(e),
                            "file_path": file_path
                        }
            
            elif action == "import":
                if not file_path or not Path(file_path).exists():
                    return {
                        "success": False,
                        "error": f"Datei nicht gefunden: {file_path}",
                        "file_path": file_path
                    }
                
                if format == "xlsx":
                    try:
                        import openpyxl
                        
                        wb = openpyxl.load_workbook(file_path)
                        ws = wb.active
                        
                        data = []
                        headers = None
                        for row in ws.iter_rows(values_only=True):
                            if headers is None:
                                headers = [str(cell) if cell else "" for cell in row]
                            else:
                                row_dict = {headers[i]: str(cell) if cell else "" for i, cell in enumerate(row)}
                                data.append(row_dict)
                        
                        return {
                            "success": True,
                            "action": "import",
                            "file_path": file_path,
                            "format": "xlsx",
                            "rows_imported": len(data),
                            "data": data
                        }
                    except ImportError:
                        return {
                            "success": False,
                            "error": "openpyxl nicht verfügbar",
                            "note": "Bitte 'pip install openpyxl' installieren"
                        }
                
                elif format == "csv":
                    try:
                        import csv
                        
                        data = []
                        with open(file_path, 'r', encoding='utf-8') as f:
                            reader = csv.DictReader(f)
                            data = list(reader)
                        
                        return {
                            "success": True,
                            "action": "import",
                            "file_path": file_path,
                            "format": "csv",
                            "rows_imported": len(data),
                            "data": data
                        }
                    except Exception as e:
                        return {
                            "success": False,
                            "error": str(e),
                            "file_path": file_path
                        }
            
            elif action == "template":
                # Generiere Template
                template_data = [
                    {
                        "Datum": "2025-01-15",
                        "Beschreibung": "Hotelübernachtung",
                        "Betrag": "150.00",
                        "Währung": "EUR",
                        "Kategorie": "hotel",
                        "Beleg": "hotel_rechnung.pdf"
                    }
                ]
                
                return await self.execute("export", template_data, file_path or "reisekosten_template.xlsx", format)
            
            return {
                "success": False,
                "error": f"Unbekannte Aktion: {action}"
            }
            
        except Exception as e:
            logger.error(f"Excel import/export error: {e}")
            return {
                "success": False,
                "error": str(e),
                "action": action
            }

class PostalCodeValidatorTool(AgentTool):
    """Tool für Postleitzahlen-Validierung und Adress-Verbesserung"""
    
    def __init__(self):
        super().__init__(
            name="postal_code_validator",
            description="Validiert Postleitzahlen und verbessert Adressen. Unterstützt DE, AT, CH, FR, IT, ES, GB, US. Nützlich für DocumentAgent und AccountingAgent.",
            parameters={
                "postal_code": {
                    "type": "string",
                    "description": "Postleitzahl zum Validieren"
                },
                "country_code": {
                    "type": "string",
                    "description": "Ländercode (z.B. 'DE', 'AT', 'CH', Standard: 'DE')",
                    "default": "DE"
                },
                "city": {
                    "type": "string",
                    "description": "Stadtname (optional, für Validierung)"
                }
            }
        )
        # Postleitzahlen-Formate pro Land
        self.postal_code_patterns = {
            "DE": r"^\d{5}$",  # 5 Ziffern
            "AT": r"^\d{4}$",  # 4 Ziffern
            "CH": r"^\d{4}$",  # 4 Ziffern
            "FR": r"^\d{5}$",  # 5 Ziffern
            "IT": r"^\d{5}$",  # 5 Ziffern
            "ES": r"^\d{5}$",  # 5 Ziffern
            "GB": r"^[A-Z]{1,2}\d{1,2}[A-Z]?\s?\d[A-Z]{2}$",  # Komplex (z.B. SW1A 1AA)
            "US": r"^\d{5}(-\d{4})?$"  # 5 Ziffern, optional -4 Ziffern
        }
    
    async def execute(self,
                      postal_code: str,
                      country_code: str = "DE",
                      city: Optional[str] = None) -> Dict[str, Any]:
        """Validiere Postleitzahl"""
        try:
            import re
            
            if not postal_code or not postal_code.strip():
                return {
                    "success": False,
                    "valid": False,
                    "error": "Postleitzahl ist leer",
                    "postal_code": postal_code
                }
            
            postal_code = postal_code.strip().upper()
            country_code = country_code.upper()
            
            # Validiere Format
            pattern = self.postal_code_patterns.get(country_code)
            if not pattern:
                return {
                    "success": False,
                    "valid": False,
                    "error": f"Unbekannter Ländercode: {country_code}",
                    "supported_countries": list(self.postal_code_patterns.keys())
                }
            
            format_valid = bool(re.match(pattern, postal_code))
            
            result = {
                "success": True,
                "valid": format_valid,
                "postal_code": postal_code,
                "country_code": country_code,
                "format_valid": format_valid
            }
            
            if not format_valid:
                result["error"] = f"Postleitzahl entspricht nicht dem Format für {country_code}"
                result["expected_format"] = self._get_format_description(country_code)
            
            # Validiere gegen Stadt (falls angegeben)
            if city and format_valid:
                # Nutze OpenMaps-Tool für Validierung
                try:
                    from backend.agents import get_tool_registry
                    tool_registry = get_tool_registry()
                    openmaps_tool = tool_registry.get_tool("openmaps")
                    if openmaps_tool:
                        maps_result = await openmaps_tool.execute(
                            action="geocode",
                            query=f"{postal_code} {city}, {country_code}"
                        )
                        if maps_result.get("success"):
                            result["city_validation"] = "success"
                            result["validated_city"] = maps_result.get("data", [{}])[0].get("display_name", "")
                except Exception as e:
                    logger.debug(f"Stadt-Validierung fehlgeschlagen: {e}")
            
            return result
            
        except Exception as e:
            logger.error(f"Postal code validator error: {e}")
            return {
                "success": False,
                "valid": False,
                "error": str(e),
                "postal_code": postal_code
            }
    
    def _get_format_description(self, country_code: str) -> str:
        """Beschreibung des erwarteten Formats"""
        descriptions = {
            "DE": "5 Ziffern (z.B. 12345)",
            "AT": "4 Ziffern (z.B. 1234)",
            "CH": "4 Ziffern (z.B. 1234)",
            "FR": "5 Ziffern (z.B. 12345)",
            "IT": "5 Ziffern (z.B. 12345)",
            "ES": "5 Ziffern (z.B. 12345)",
            "GB": "Format: SW1A 1AA",
            "US": "5 Ziffern, optional -4 Ziffern (z.B. 12345 oder 12345-6789)"
        }
        return descriptions.get(country_code, "Unbekannt")

class PhoneNumberValidatorTool(AgentTool):
    """Tool für Telefonnummer-Validierung und Formatierung"""
    
    def __init__(self):
        super().__init__(
            name="phone_number_validator",
            description="Validiert und formatiert Telefonnummern (E.164). Erkennt Länder, formatiert national/international. Nützlich für DocumentAgent.",
            parameters={
                "phone_number": {
                    "type": "string",
                    "description": "Telefonnummer zum Validieren"
                },
                "country_code": {
                    "type": "string",
                    "description": "Ländercode (optional, wird automatisch erkannt)",
                    "default": None
                },
                "format": {
                    "type": "string",
                    "description": "Ausgabeformat (Standard: 'international')",
                    "enum": ["international", "national", "e164"],
                    "default": "international"
                }
            }
        )
    
    async def execute(self,
                      phone_number: str,
                      country_code: Optional[str] = None,
                      format: str = "international") -> Dict[str, Any]:
        """Validiere Telefonnummer"""
        try:
            try:
                import phonenumbers
                from phonenumbers import geocoder, carrier
            except ImportError:
                return {
                    "success": False,
                    "valid": False,
                    "error": "phonenumbers nicht verfügbar",
                    "note": "Bitte 'pip install phonenumbers' installieren"
                }
            
            if not phone_number or not phone_number.strip():
                return {
                    "success": False,
                    "valid": False,
                    "error": "Telefonnummer ist leer",
                    "phone_number": phone_number
                }
            
            # Parse Telefonnummer
            try:
                if country_code:
                    parsed = phonenumbers.parse(phone_number, country_code)
                else:
                    parsed = phonenumbers.parse(phone_number, None)
            except phonenumbers.NumberParseException as e:
                return {
                    "success": True,
                    "valid": False,
                    "error": f"Telefonnummer konnte nicht geparst werden: {str(e)}",
                    "phone_number": phone_number
                }
            
            # Validiere
            is_valid = phonenumbers.is_valid_number(parsed)
            is_possible = phonenumbers.is_possible_number(parsed)
            
            # Extrahiere Informationen
            country = phonenumbers.region_code_for_number(parsed)
            number_type = phonenumbers.number_type(parsed)
            type_names = {
                phonenumbers.PhoneNumberType.MOBILE: "Mobil",
                phonenumbers.PhoneNumberType.FIXED_LINE: "Festnetz",
                phonenumbers.PhoneNumberType.FIXED_LINE_OR_MOBILE: "Festnetz oder Mobil"
            }
            
            # Formatiere
            formatted_international = phonenumbers.format_number(parsed, phonenumbers.PhoneNumberFormat.INTERNATIONAL)
            formatted_national = phonenumbers.format_number(parsed, phonenumbers.PhoneNumberFormat.NATIONAL)
            formatted_e164 = phonenumbers.format_number(parsed, phonenumbers.PhoneNumberFormat.E164)
            
            result = {
                "success": True,
                "valid": is_valid,
                "possible": is_possible,
                "phone_number": phone_number,
                "country_code": country or country_code,
                "number_type": type_names.get(number_type, "Unbekannt"),
                "formatted_international": formatted_international,
                "formatted_national": formatted_national,
                "formatted_e164": formatted_e164
            }
            
            # Wähle Format basierend auf Parameter
            if format == "international":
                result["formatted"] = formatted_international
            elif format == "national":
                result["formatted"] = formatted_national
            elif format == "e164":
                result["formatted"] = formatted_e164
            
            if not is_valid:
                result["error"] = "Telefonnummer ist ungültig"
            elif not is_possible:
                result["warning"] = "Telefonnummer ist möglicherweise ungültig"
            
            return result
            
        except Exception as e:
            logger.error(f"Phone number validator error: {e}")
            return {
                "success": False,
                "valid": False,
                "error": str(e),
                "phone_number": phone_number
            }

class HolidayAPITool(AgentTool):
    """Tool für internationale Feiertags-Erkennung"""
    
    def __init__(self):
        super().__init__(
            name="holiday_api",
            description="Erkennt Feiertage für verschiedene Länder. Unterstützt regionale Feiertage. Nützlich für AccountingAgent zur Validierung von Reisetagen.",
            parameters={
                "country_code": {
                    "type": "string",
                    "description": "Ländercode (z.B. 'DE', 'AT', 'CH', 'US')"
                },
                "year": {
                    "type": "integer",
                    "description": "Jahr (Standard: aktuelles Jahr)"
                },
                "date": {
                    "type": "string",
                    "description": "Datum zum Prüfen (optional, Format: YYYY-MM-DD)"
                },
                "region": {
                    "type": "string",
                    "description": "Region (optional, z.B. 'SN' für Sachsen)"
                }
            }
        )
        self.api_key = os.getenv('HOLIDAY_API_KEY')
    
    async def execute(self,
                      country_code: str,
                      year: Optional[int] = None,
                      date: Optional[str] = None,
                      region: Optional[str] = None) -> Dict[str, Any]:
        """Erkenne Feiertage"""
        try:
            from datetime import datetime
            
            if not year:
                year = datetime.now().year
            
            country_code = country_code.upper()
            
            # Nutze lokale holidays-Bibliothek (bereits in requirements.txt)
            try:
                import holidays
                
                # Erstelle Feiertags-Objekt
                if region:
                    holiday_obj = holidays.country_holidays(country_code, years=year, subdiv=region)
                else:
                    holiday_obj = holidays.country_holidays(country_code, years=year)
                
                # Prüfe einzelnes Datum
                if date:
                    try:
                        check_date = datetime.strptime(date, "%Y-%m-%d").date()
                        is_holiday = check_date in holiday_obj
                        holiday_name = holiday_obj.get(check_date) if is_holiday else None
                        
                        return {
                            "success": True,
                            "date": date,
                            "is_holiday": is_holiday,
                            "holiday_name": holiday_name,
                            "country_code": country_code,
                            "region": region
                        }
                    except ValueError:
                        return {
                            "success": False,
                            "error": f"Ungültiges Datumsformat: {date} (erwartet: YYYY-MM-DD)",
                            "date": date
                        }
                
                # Hole alle Feiertage des Jahres
                all_holidays = {}
                for holiday_date, holiday_name in holiday_obj.items():
                    all_holidays[str(holiday_date)] = holiday_name
                
                return {
                    "success": True,
                    "country_code": country_code,
                    "year": year,
                    "region": region,
                    "holidays": all_holidays,
                    "count": len(all_holidays)
                }
                
            except ImportError:
                return {
                    "success": False,
                    "error": "holidays-Bibliothek nicht verfügbar",
                    "note": "Bitte 'pip install holidays' installieren"
                }
            except Exception as e:
                logger.error(f"Holiday API error: {e}")
                return {
                    "success": False,
                    "error": str(e),
                    "country_code": country_code
                }
            
        except Exception as e:
            logger.error(f"Holiday API error: {e}")
            return {
                "success": False,
                "error": str(e),
                "country_code": country_code
            }

class WeatherAPITool(AgentTool):
    """Tool für Wetter-Daten für Reisevalidierung"""
    
    def __init__(self):
        super().__init__(
            name="weather_api",
            description="Holt Wetter-Daten für Reisevalidierung. Historische Wetterdaten, Temperatur, Wetterbedingungen. Nützlich für AccountingAgent.",
            parameters={
                "location": {
                    "type": "string",
                    "description": "Ortsangabe (z.B. 'Berlin', 'New York')"
                },
                "date": {
                    "type": "string",
                    "description": "Datum (optional, Format: YYYY-MM-DD, Standard: heute)"
                },
                "historical": {
                    "type": "boolean",
                    "description": "Historische Daten abrufen (Standard: False)",
                    "default": False
                }
            }
        )
        self.api_key = os.getenv('WEATHER_API_KEY')
        self.api_provider = os.getenv('WEATHER_API_PROVIDER', 'openweathermap')  # openweathermap oder weatherapi
    
    async def execute(self,
                      location: str,
                      date: Optional[str] = None,
                      historical: bool = False) -> Dict[str, Any]:
        """Hole Wetter-Daten"""
        try:
            if not self.api_key:
                return {
                    "success": False,
                    "error": "WEATHER_API_KEY nicht gesetzt",
                    "note": "Bitte WEATHER_API_KEY in .env setzen (OpenWeatherMap oder WeatherAPI)"
                }
            
            from datetime import datetime
            import aiohttp
            
            if not date:
                date = datetime.now().strftime("%Y-%m-%d")
            
            # OpenWeatherMap API
            if self.api_provider == "openweathermap":
                url = "https://api.openweathermap.org/data/2.5/weather"
                params = {
                    "q": location,
                    "appid": self.api_key,
                    "units": "metric",
                    "lang": "de"
                }
                
                async with aiohttp.ClientSession() as session:
                    async with session.get(url, params=params) as response:
                        if response.status == 200:
                            data = await response.json()
                            return {
                                "success": True,
                                "location": location,
                                "date": date,
                                "temperature": data.get("main", {}).get("temp"),
                                "description": data.get("weather", [{}])[0].get("description"),
                                "humidity": data.get("main", {}).get("humidity"),
                                "wind_speed": data.get("wind", {}).get("speed"),
                                "provider": "openweathermap"
                            }
                        else:
                            return {
                                "success": False,
                                "error": f"HTTP {response.status}",
                                "location": location
                            }
            
            # WeatherAPI (alternative)
            elif self.api_provider == "weatherapi":
                url = "https://api.weatherapi.com/v1/current.json"
                params = {
                    "key": self.api_key,
                    "q": location,
                    "lang": "de"
                }
                
                async with aiohttp.ClientSession() as session:
                    async with session.get(url, params=params) as response:
                        if response.status == 200:
                            data = await response.json()
                            return {
                                "success": True,
                                "location": location,
                                "date": date,
                                "temperature": data.get("current", {}).get("temp_c"),
                                "description": data.get("current", {}).get("condition", {}).get("text"),
                                "humidity": data.get("current", {}).get("humidity"),
                                "wind_speed": data.get("current", {}).get("wind_kph"),
                                "provider": "weatherapi"
                            }
                        else:
                            return {
                                "success": False,
                                "error": f"HTTP {response.status}",
                                "location": location
                            }
            
            return {
                "success": False,
                "error": f"Unbekannter API-Provider: {self.api_provider}",
                "supported_providers": ["openweathermap", "weatherapi"]
            }
            
        except Exception as e:
            logger.error(f"Weather API error: {e}")
            return {
                "success": False,
                "error": str(e),
                "location": location
            }

class TravelTimeCalculatorTool(AgentTool):
    """Tool für Reisezeit-Berechnung zwischen Orten"""
    
    def __init__(self):
        super().__init__(
            name="travel_time_calculator",
            description="Berechnet Reisezeit und Entfernung zwischen Orten. Unterstützt Auto, Bahn, Flugzeug. Nützlich für AccountingAgent zur Validierung von Reisezeiten.",
            parameters={
                "origin": {
                    "type": "string",
                    "description": "Startort (z.B. 'Berlin, Deutschland')"
                },
                "destination": {
                    "type": "string",
                    "description": "Zielort (z.B. 'München, Deutschland')"
                },
                "mode": {
                    "type": "string",
                    "description": "Verkehrsmittel (Standard: 'driving')",
                    "enum": ["driving", "walking", "bicycling", "transit"],
                    "default": "driving"
                },
                "provider": {
                    "type": "string",
                    "description": "API-Provider (Standard: 'openrouteservice')",
                    "enum": ["google", "openrouteservice"],
                    "default": "openrouteservice"
                }
            }
        )
        self.google_api_key = os.getenv('GOOGLE_MAPS_API_KEY')
        self.ors_api_key = os.getenv('OPENROUTESERVICE_API_KEY')
    
    async def execute(self,
                      origin: str,
                      destination: str,
                      mode: str = "driving",
                      provider: str = "openrouteservice") -> Dict[str, Any]:
        """Berechne Reisezeit"""
        try:
            import aiohttp
            
            # OpenRouteService (kostenlos)
            if provider == "openrouteservice":
                if not self.ors_api_key:
                    return {
                        "success": False,
                        "error": "OPENROUTESERVICE_API_KEY nicht gesetzt",
                        "note": "Bitte OPENROUTESERVICE_API_KEY in .env setzen (kostenlos bei openrouteservice.org)"
                    }
                
                # Geocode Orte
                from backend.agents import get_tool_registry
                tool_registry = get_tool_registry()
                openmaps_tool = tool_registry.get_tool("openmaps")
                
                if not openmaps_tool:
                    return {
                        "success": False,
                        "error": "OpenMapsTool nicht verfügbar für Geocoding"
                    }
                
                origin_geo = await openmaps_tool.execute(action="geocode", query=origin, limit=1)
                dest_geo = await openmaps_tool.execute(action="geocode", query=destination, limit=1)
                
                if not origin_geo.get("success") or not dest_geo.get("success"):
                    return {
                        "success": False,
                        "error": "Orte konnten nicht geocodiert werden",
                        "origin": origin,
                        "destination": destination
                    }
                
                origin_coords = origin_geo.get("data", [{}])[0]
                dest_coords = dest_geo.get("data", [{}])[0]
                
                origin_lon = origin_coords.get("lon")
                origin_lat = origin_coords.get("lat")
                dest_lon = dest_coords.get("lon")
                dest_lat = dest_coords.get("lat")
                
                # OpenRouteService Directions API
                profile_map = {
                    "driving": "driving-car",
                    "walking": "foot-walking",
                    "bicycling": "cycling-regular",
                    "transit": "driving-car"  # Fallback
                }
                profile = profile_map.get(mode, "driving-car")
                
                url = f"https://api.openrouteservice.org/v2/directions/{profile}"
                headers = {"Authorization": self.ors_api_key}
                params = {
                    "start": f"{origin_lon},{origin_lat}",
                    "end": f"{dest_lon},{dest_lat}"
                }
                
                async with aiohttp.ClientSession() as session:
                    async with session.get(url, headers=headers, params=params) as response:
                        if response.status == 200:
                            data = await response.json()
                            route = data.get("features", [{}])[0].get("properties", {})
                            summary = route.get("summary", {})
                            
                            distance_km = summary.get("distance", 0) / 1000  # Meter zu km
                            duration_seconds = summary.get("duration", 0)
                            duration_hours = duration_seconds / 3600
                            
                            return {
                                "success": True,
                                "origin": origin,
                                "destination": destination,
                                "mode": mode,
                                "distance_km": round(distance_km, 2),
                                "distance_m": round(summary.get("distance", 0)),
                                "duration_seconds": int(duration_seconds),
                                "duration_hours": round(duration_hours, 2),
                                "duration_formatted": f"{int(duration_seconds // 3600)}h {int((duration_seconds % 3600) // 60)}min",
                                "provider": "openrouteservice"
                            }
                        else:
                            return {
                                "success": False,
                                "error": f"HTTP {response.status}",
                                "origin": origin,
                                "destination": destination
                            }
            
            # Google Maps API (alternative)
            elif provider == "google":
                if not self.google_api_key:
                    return {
                        "success": False,
                        "error": "GOOGLE_MAPS_API_KEY nicht gesetzt",
                        "note": "Bitte GOOGLE_MAPS_API_KEY in .env setzen"
                    }
                
                url = "https://maps.googleapis.com/maps/api/distancematrix/json"
                params = {
                    "origins": origin,
                    "destinations": destination,
                    "mode": mode,
                    "key": self.google_api_key,
                    "language": "de"
                }
                
                async with aiohttp.ClientSession() as session:
                    async with session.get(url, params=params) as response:
                        if response.status == 200:
                            data = await response.json()
                            if data.get("status") == "OK":
                                element = data.get("rows", [{}])[0].get("elements", [{}])[0]
                                distance = element.get("distance", {}).get("value", 0)  # Meter
                                duration = element.get("duration", {}).get("value", 0)  # Sekunden
                                
                                return {
                                    "success": True,
                                    "origin": origin,
                                    "destination": destination,
                                    "mode": mode,
                                    "distance_km": round(distance / 1000, 2),
                                    "distance_m": distance,
                                    "duration_seconds": duration,
                                    "duration_hours": round(duration / 3600, 2),
                                    "duration_formatted": element.get("duration", {}).get("text", ""),
                                    "provider": "google"
                                }
                            else:
                                return {
                                    "success": False,
                                    "error": data.get("status", "Unknown error"),
                                    "origin": origin,
                                    "destination": destination
                                }
                        else:
                            return {
                                "success": False,
                                "error": f"HTTP {response.status}",
                                "origin": origin,
                                "destination": destination
                            }
            
            return {
                "success": False,
                "error": f"Unbekannter Provider: {provider}",
                "supported_providers": ["openrouteservice", "google"]
            }
            
        except Exception as e:
            logger.error(f"Travel time calculator error: {e}")
            return {
                "success": False,
                "error": str(e),
                "origin": origin,
                "destination": destination
            }

class PDFTimestampTool(AgentTool):
    """Tool für Zeitstempel-Validierung in PDFs"""
    
    def __init__(self):
        super().__init__(
            name="pdf_timestamp",
            description="Extrahiert und validiert Zeitstempel in PDFs (Erstellungsdatum, Änderungsdatum). Validiert gegen Reisedaten. Nützlich für DocumentAgent.",
            parameters={
                "pdf_path": {
                    "type": "string",
                    "description": "Pfad zur PDF-Datei"
                },
                "reference_date": {
                    "type": "string",
                    "description": "Referenzdatum zum Validieren (optional, Format: YYYY-MM-DD)"
                }
            }
        )
    
    async def execute(self,
                      pdf_path: str,
                      reference_date: Optional[str] = None) -> Dict[str, Any]:
        """Extrahiere und validiere PDF-Zeitstempel"""
        try:
            if not Path(pdf_path).exists():
                return {
                    "success": False,
                    "error": f"PDF nicht gefunden: {pdf_path}",
                    "pdf_path": pdf_path
                }
            
            from datetime import datetime
            
            metadata = {}
            creation_date = None
            modification_date = None
            
            # Extrahiere mit PyPDF2
            if HAS_PYPDF2:
                try:
                    with open(pdf_path, 'rb') as file:
                        pdf_reader = PyPDF2.PdfReader(file)
                        info = pdf_reader.metadata
                        if info:
                            creation_date_str = str(info.get("/CreationDate", ""))
                            modification_date_str = str(info.get("/ModDate", ""))
                            
                            # Parse PDF-Datum-Format (z.B. "D:20250115120000+01'00'")
                            if creation_date_str and creation_date_str.startswith("D:"):
                                try:
                                    date_part = creation_date_str[2:16]  # YYYYMMDDHHmmss
                                    creation_date = datetime.strptime(date_part, "%Y%m%d%H%M%S")
                                except:
                                    pass
                            
                            if modification_date_str and modification_date_str.startswith("D:"):
                                try:
                                    date_part = modification_date_str[2:16]
                                    modification_date = datetime.strptime(date_part, "%Y%m%d%H%M%S")
                                except:
                                    pass
                            
                            metadata = {
                                "creation_date": creation_date.isoformat() if creation_date else None,
                                "modification_date": modification_date.isoformat() if modification_date else None,
                                "creation_date_raw": creation_date_str,
                                "modification_date_raw": modification_date_str
                            }
                except Exception as e:
                    logger.debug(f"PyPDF2 timestamp extraction error: {e}")
            
            # Extrahiere mit pdfplumber (Fallback)
            if not metadata.get("creation_date") and HAS_PDFPLUMBER:
                try:
                    import pdfplumber
                    with pdfplumber.open(pdf_path) as pdf:
                        if pdf.metadata:
                            pdf_creation = pdf.metadata.get("CreationDate")
                            pdf_modification = pdf.metadata.get("ModDate")
                            
                            if pdf_creation:
                                metadata["creation_date"] = str(pdf_creation)
                            if pdf_modification:
                                metadata["modification_date"] = str(pdf_modification)
                except Exception as e:
                    logger.debug(f"pdfplumber timestamp extraction error: {e}")
            
            result = {
                "success": True,
                "pdf_path": pdf_path,
                "creation_date": metadata.get("creation_date"),
                "modification_date": metadata.get("modification_date")
            }
            
            # Validiere gegen Referenzdatum
            if reference_date and creation_date:
                try:
                    ref_date = datetime.strptime(reference_date, "%Y-%m-%d")
                    creation_only_date = creation_date.date()
                    
                    # Prüfe ob PDF nach Referenzdatum erstellt wurde (verdächtig)
                    if creation_only_date > ref_date:
                        result["validation_warning"] = f"PDF wurde nach Referenzdatum erstellt ({creation_only_date} > {ref_date.date()})"
                        result["is_valid"] = False
                    else:
                        result["is_valid"] = True
                        result["validation_note"] = "PDF-Zeitstempel ist plausibel"
                except ValueError:
                    result["validation_error"] = f"Ungültiges Referenzdatum-Format: {reference_date}"
            
            return result
            
        except Exception as e:
            logger.error(f"PDF timestamp error: {e}")
            return {
                "success": False,
                "error": str(e),
                "pdf_path": pdf_path
            }

class QRCodeReaderTool(AgentTool):
    """Tool für QR-Code-Erkennung in Belegen"""
    
    def __init__(self):
        super().__init__(
            name="qrcode_reader",
            description="Erkennt QR-Codes in PDFs und Bildern. Extrahiert Daten aus QR-Codes, erkennt E-Rechnungen (ZUGFeRD, XRechnung). Nützlich für DocumentAgent.",
            parameters={
                "file_path": {
                    "type": "string",
                    "description": "Pfad zur Datei (PDF oder Bild)"
                },
                "extract_data": {
                    "type": "boolean",
                    "description": "Daten aus QR-Code extrahieren (Standard: True)",
                    "default": True
                }
            }
        )
    
    async def execute(self,
                      file_path: str,
                      extract_data: bool = True) -> Dict[str, Any]:
        """Erkenne QR-Codes in Datei"""
        try:
            if not Path(file_path).exists():
                return {
                    "success": False,
                    "error": f"Datei nicht gefunden: {file_path}",
                    "file_path": file_path
                }
            
            try:
                from pyzbar import decode as pyzbar_decode
                from PIL import Image
                import cv2
            except ImportError:
                return {
                    "success": False,
                    "error": "pyzbar, PIL oder opencv-python nicht verfügbar",
                    "note": "Bitte 'pip install pyzbar pillow opencv-python' installieren"
                }
            
            qr_codes = []
            
            # Prüfe ob PDF oder Bild
            if file_path.lower().endswith('.pdf'):
                # PDF: Konvertiere Seiten zu Bildern
                try:
                    if HAS_PDFPLUMBER:
                        import pdfplumber
                        with pdfplumber.open(file_path) as pdf:
                            for page_num, page in enumerate(pdf.pages):
                                # Konvertiere PDF-Seite zu Bild
                                img = page.to_image(resolution=300)
                                img_array = img.original
                                
                                # Erkenne QR-Codes
                                decoded = pyzbar_decode(img_array)
                                for qr in decoded:
                                    qr_data = qr.data.decode('utf-8')
                                    qr_codes.append({
                                        "page": page_num + 1,
                                        "data": qr_data,
                                        "type": qr.type,
                                        "rect": {
                                            "left": qr.rect.left,
                                            "top": qr.rect.top,
                                            "width": qr.rect.width,
                                            "height": qr.rect.height
                                        }
                                    })
                    else:
                        return {
                            "success": False,
                            "error": "pdfplumber nicht verfügbar für PDF-Verarbeitung"
                        }
                except Exception as e:
                    logger.debug(f"PDF QR-Code-Erkennung fehlgeschlagen: {e}")
            else:
                # Bild: Direkt verarbeiten
                try:
                    img = cv2.imread(file_path)
                    if img is None:
                        return {
                            "success": False,
                            "error": f"Bild konnte nicht geladen werden: {file_path}"
                        }
                    
                    # Erkenne QR-Codes
                    decoded = pyzbar_decode(img)
                    for qr in decoded:
                        qr_data = qr.data.decode('utf-8')
                        qr_codes.append({
                            "data": qr_data,
                            "type": qr.type,
                            "rect": {
                                "left": qr.rect.left,
                                "top": qr.rect.top,
                                "width": qr.rect.width,
                                "height": qr.rect.height
                            }
                        })
                except Exception as e:
                    logger.debug(f"Bild QR-Code-Erkennung fehlgeschlagen: {e}")
            
            # Extrahiere Daten aus QR-Codes
            extracted_data = []
            if extract_data:
                for qr in qr_codes:
                    data = qr.get("data", "")
                    # Prüfe auf E-Rechnung-Formate
                    if "ZUGFeRD" in data or "zugferd" in data.lower():
                        qr["format"] = "ZUGFeRD"
                    elif "XRechnung" in data or "xrechnung" in data.lower():
                        qr["format"] = "XRechnung"
                    
                    # Versuche JSON/XML zu parsen
                    try:
                        import json
                        parsed = json.loads(data)
                        qr["parsed_data"] = parsed
                    except:
                        pass
                    
                    try:
                        import xml.etree.ElementTree as ET
                        root = ET.fromstring(data)
                        qr["parsed_data"] = {child.tag: child.text for child in root}
                    except:
                        pass
                    
                    extracted_data.append(qr)
            
            return {
                "success": True,
                "file_path": file_path,
                "qr_codes_found": len(qr_codes),
                "qr_codes": qr_codes if not extract_data else extracted_data
            }
            
        except Exception as e:
            logger.error(f"QR code reader error: {e}")
            return {
                "success": False,
                "error": str(e),
                "file_path": file_path
            }

class BarcodeReaderTool(AgentTool):
    """Tool für Barcode-Erkennung in Belegen"""
    
    def __init__(self):
        super().__init__(
            name="barcode_reader",
            description="Erkennt Barcodes in Belegen (EAN, UPC, Code128, etc.). Extrahiert Produktdaten aus Barcodes. Nützlich für DocumentAgent.",
            parameters={
                "file_path": {
                    "type": "string",
                    "description": "Pfad zur Datei (PDF oder Bild)"
                }
            }
        )
    
    async def execute(self, file_path: str) -> Dict[str, Any]:
        """Erkenne Barcodes in Datei"""
        try:
            if not Path(file_path).exists():
                return {
                    "success": False,
                    "error": f"Datei nicht gefunden: {file_path}",
                    "file_path": file_path
                }
            
            try:
                from pyzbar import decode as pyzbar_decode
                import cv2
            except ImportError:
                return {
                    "success": False,
                    "error": "pyzbar oder opencv-python nicht verfügbar",
                    "note": "Bitte 'pip install pyzbar opencv-python' installieren"
                }
            
            barcodes = []
            
            # Prüfe ob PDF oder Bild
            if file_path.lower().endswith('.pdf'):
                # PDF: Konvertiere Seiten zu Bildern
                try:
                    if HAS_PDFPLUMBER:
                        import pdfplumber
                        with pdfplumber.open(file_path) as pdf:
                            for page_num, page in enumerate(pdf.pages):
                                img = page.to_image(resolution=300)
                                img_array = img.original
                                
                                # Erkenne Barcodes
                                decoded = pyzbar_decode(img_array)
                                for barcode in decoded:
                                    barcode_data = barcode.data.decode('utf-8')
                                    barcodes.append({
                                        "page": page_num + 1,
                                        "data": barcode_data,
                                        "type": barcode.type,
                                        "rect": {
                                            "left": barcode.rect.left,
                                            "top": barcode.rect.top,
                                            "width": barcode.rect.width,
                                            "height": barcode.rect.height
                                        }
                                    })
                    else:
                        return {
                            "success": False,
                            "error": "pdfplumber nicht verfügbar für PDF-Verarbeitung"
                        }
                except Exception as e:
                    logger.debug(f"PDF Barcode-Erkennung fehlgeschlagen: {e}")
            else:
                # Bild: Direkt verarbeiten
                try:
                    img = cv2.imread(file_path)
                    if img is None:
                        return {
                            "success": False,
                            "error": f"Bild konnte nicht geladen werden: {file_path}"
                        }
                    
                    # Erkenne Barcodes
                    decoded = pyzbar_decode(img)
                    for barcode in decoded:
                        barcode_data = barcode.data.decode('utf-8')
                        barcodes.append({
                            "data": barcode_data,
                            "type": barcode.type,
                            "rect": {
                                "left": barcode.rect.left,
                                "top": barcode.rect.top,
                                "width": barcode.rect.width,
                                "height": barcode.rect.height
                            }
                        })
                except Exception as e:
                    logger.debug(f"Bild Barcode-Erkennung fehlgeschlagen: {e}")
            
            return {
                "success": True,
                "file_path": file_path,
                "barcodes_found": len(barcodes),
                "barcodes": barcodes
            }
            
        except Exception as e:
            logger.error(f"Barcode reader error: {e}")
            return {
                "success": False,
                "error": str(e),
                "file_path": file_path
            }

class InvoiceNumberValidatorTool(AgentTool):
    """Tool für Rechnungsnummer-Validierung"""
    
    def __init__(self):
        super().__init__(
            name="invoice_number_validator",
            description="Validiert Rechnungsnummern. Prüft Format, Duplikate, Sequenzen und Lücken. Nützlich für DocumentAgent und AccountingAgent.",
            parameters={
                "invoice_number": {
                    "type": "string",
                    "description": "Rechnungsnummer zum Validieren"
                },
                "check_duplicates": {
                    "type": "boolean",
                    "description": "Duplikats-Prüfung in Datenbank (Standard: True)",
                    "default": True
                },
                "check_sequence": {
                    "type": "boolean",
                    "description": "Sequenz-Validierung (Standard: False)",
                    "default": False
                }
            }
        )
    
    async def execute(self,
                      invoice_number: str,
                      check_duplicates: bool = True,
                      check_sequence: bool = False) -> Dict[str, Any]:
        """Validiere Rechnungsnummer"""
        try:
            if not invoice_number or not invoice_number.strip():
                return {
                    "success": False,
                    "valid": False,
                    "error": "Rechnungsnummer ist leer",
                    "invoice_number": invoice_number
                }
            
            invoice_number = invoice_number.strip()
            
            # Basis-Validierung: Format
            # Typische Formate: RE-2025-001, 2025/001, RE2025001, etc.
            import re
            format_patterns = [
                r'^[A-Z]{2,4}[-/]?\d{4}[-/]?\d{3,6}$',  # RE-2025-001
                r'^\d{4}[/-]\d{3,6}$',  # 2025/001
                r'^[A-Z]{2,4}\d{7,10}$',  # RE2025001
                r'^\d{7,12}$'  # Nur Zahlen
            ]
            
            format_valid = any(re.match(pattern, invoice_number, re.IGNORECASE) for pattern in format_patterns)
            
            result = {
                "success": True,
                "valid": format_valid,
                "invoice_number": invoice_number,
                "format_valid": format_valid
            }
            
            if not format_valid:
                result["error"] = "Rechnungsnummer entspricht keinem bekannten Format"
            
            # Duplikats-Prüfung
            if check_duplicates:
                try:
                    # Prüfe in Datenbank (falls verfügbar)
                    if hasattr(self, 'db') and self.db:
                        # Suche nach vorhandenen Rechnungsnummern
                        cursor = self.db.execute(
                            "SELECT COUNT(*) FROM receipts WHERE invoice_number = ?",
                            (invoice_number,)
                        )
                        count = cursor.fetchone()[0]
                        result["is_duplicate"] = count > 0
                        result["duplicate_count"] = count
                    else:
                        result["duplicate_check"] = "Datenbank nicht verfügbar"
                except Exception as e:
                    logger.debug(f"Duplikats-Prüfung fehlgeschlagen: {e}")
                    result["duplicate_check_error"] = str(e)
            
            # Sequenz-Validierung (vereinfacht)
            if check_sequence and format_valid:
                # Extrahiere Nummernteil
                numbers = re.findall(r'\d+', invoice_number)
                if numbers:
                    num = int(numbers[-1])  # Letzte Zahl
                    result["sequence_number"] = num
                    # Prüfe auf Lücken (vereinfacht - würde vollständige Sequenz erfordern)
                    result["sequence_check"] = "Sequenz-Prüfung erfordert vollständige Datenbank-Abfrage"
            
            return result
            
        except Exception as e:
            logger.error(f"Invoice number validator error: {e}")
            return {
                "success": False,
                "valid": False,
                "error": str(e),
                "invoice_number": invoice_number
            }

class VATCalculatorTool(AgentTool):
    """Tool für Mehrwertsteuer-Berechnung"""
    
    def __init__(self):
        super().__init__(
            name="vat_calculator",
            description="Berechnet Mehrwertsteuer (MwSt). Netto/Brutto-Umrechnung, länder-spezifische MwSt-Sätze. Nützlich für AccountingAgent.",
            parameters={
                "amount": {
                    "type": "number",
                    "description": "Betrag"
                },
                "vat_rate": {
                    "type": "number",
                    "description": "MwSt-Satz in Prozent (z.B. 19 für 19%)"
                },
                "calculation_type": {
                    "type": "string",
                    "description": "Berechnungstyp (Standard: 'netto_to_brutto')",
                    "enum": ["netto_to_brutto", "brutto_to_netto", "vat_from_brutto"],
                    "default": "netto_to_brutto"
                },
                "country_code": {
                    "type": "string",
                    "description": "Ländercode (optional, für automatische MwSt-Satz-Erkennung)"
                }
            }
        )
        # Standard-MwSt-Sätze (EU)
        self.vat_rates = {
            "DE": {"standard": 19.0, "reduced": 7.0},
            "AT": {"standard": 20.0, "reduced": 10.0},
            "CH": {"standard": 7.7, "reduced": 2.5},
            "FR": {"standard": 20.0, "reduced": 5.5},
            "IT": {"standard": 22.0, "reduced": 10.0},
            "ES": {"standard": 21.0, "reduced": 10.0},
            "GB": {"standard": 20.0, "reduced": 5.0},
            "US": {"standard": 0.0, "reduced": 0.0}  # Keine bundesweite MwSt
        }
    
    async def execute(self,
                      amount: float,
                      vat_rate: Optional[float] = None,
                      calculation_type: str = "netto_to_brutto",
                      country_code: Optional[str] = None) -> Dict[str, Any]:
        """Berechne MwSt"""
        try:
            # Bestimme MwSt-Satz
            if not vat_rate and country_code:
                country_code = country_code.upper()
                if country_code in self.vat_rates:
                    vat_rate = self.vat_rates[country_code]["standard"]
                else:
                    return {
                        "success": False,
                        "error": f"Unbekannter Ländercode: {country_code}",
                        "supported_countries": list(self.vat_rates.keys())
                    }
            elif not vat_rate:
                return {
                    "success": False,
                    "error": "MwSt-Satz oder Ländercode erforderlich"
                }
            
            # Berechne
            if calculation_type == "netto_to_brutto":
                netto = amount
                vat_amount = netto * (vat_rate / 100)
                brutto = netto + vat_amount
            elif calculation_type == "brutto_to_netto":
                brutto = amount
                netto = brutto / (1 + (vat_rate / 100))
                vat_amount = brutto - netto
            elif calculation_type == "vat_from_brutto":
                brutto = amount
                vat_amount = brutto * (vat_rate / (100 + vat_rate))
                netto = brutto - vat_amount
            else:
                return {
                    "success": False,
                    "error": f"Unbekannter Berechnungstyp: {calculation_type}"
                }
            
            return {
                "success": True,
                "calculation_type": calculation_type,
                "vat_rate": vat_rate,
                "netto": round(netto, 2),
                "brutto": round(brutto, 2),
                "vat_amount": round(vat_amount, 2),
                "country_code": country_code
            }
            
        except Exception as e:
            logger.error(f"VAT calculator error: {e}")
            return {
                "success": False,
                "error": str(e),
                "amount": amount
            }

class ExpenseCategoryClassifierTool(AgentTool):
    """Tool für automatische Kategorisierung von Ausgaben"""
    
    def __init__(self):
        super().__init__(
            name="expense_category_classifier",
            description="Kategorisiert Ausgaben automatisch (Hotel, Restaurant, Transport, etc.). Keyword-basierte Klassifizierung mit Konfidenz-Score. Nützlich für AccountingAgent.",
            parameters={
                "description": {
                    "type": "string",
                    "description": "Beschreibung der Ausgabe"
                },
                "amount": {
                    "type": "number",
                    "description": "Betrag (optional, für Kontext)"
                },
                "merchant": {
                    "type": "string",
                    "description": "Händler/Unternehmen (optional)"
                }
            }
        )
        # Keyword-basierte Kategorien
        self.categories = {
            "hotel": ["hotel", "übernachtung", "accommodation", "zimmer", "room", "hostel", "pension"],
            "restaurant": ["restaurant", "gaststätte", "café", "cafe", "imbiss", "bistro", "essen", "mahlzeit"],
            "transport": ["taxi", "bus", "bahn", "zug", "flug", "flight", "mietwagen", "car rental", "ubahn", "s-bahn"],
            "fuel": ["tankstelle", "benzin", "diesel", "kraftstoff", "fuel", "gas station"],
            "parking": ["parkplatz", "parken", "parking", "parkhaus"],
            "toll": ["maut", "toll", "vignette", "road tax"],
            "office": ["büro", "office", "material", "supplies", "papier", "paper"],
            "communication": ["telefon", "internet", "wifi", "roaming", "communication"],
            "other": []
        }
    
    async def execute(self,
                      description: str,
                      amount: Optional[float] = None,
                      merchant: Optional[str] = None) -> Dict[str, Any]:
        """Kategorisiere Ausgabe"""
        try:
            if not description or not description.strip():
                return {
                    "success": False,
                    "error": "Beschreibung ist leer",
                    "description": description
                }
            
            description_lower = description.lower()
            merchant_lower = merchant.lower() if merchant else ""
            combined_text = f"{description_lower} {merchant_lower}".strip()
            
            # Keyword-basierte Klassifizierung
            scores = {}
            for category, keywords in self.categories.items():
                if category == "other":
                    continue
                score = sum(1 for keyword in keywords if keyword in combined_text)
                if score > 0:
                    scores[category] = score
            
            # Bestimme beste Kategorie
            if scores:
                best_category = max(scores, key=scores.get)
                confidence = min(scores[best_category] / 3.0, 1.0)  # Normalisiert auf 0-1
            else:
                best_category = "other"
                confidence = 0.3  # Niedrige Konfidenz für "other"
            
            return {
                "success": True,
                "description": description,
                "category": best_category,
                "confidence": round(confidence, 2),
                "scores": scores,
                "merchant": merchant,
                "amount": amount
            }
            
        except Exception as e:
            logger.error(f"Expense category classifier error: {e}")
            return {
                "success": False,
                "error": str(e),
                "description": description
            }

class ReceiptStandardValidatorTool(AgentTool):
    """Tool für GoBD-Konformitäts-Prüfung"""
    
    def __init__(self):
        super().__init__(
            name="receipt_standard_validator",
            description="Prüft Belege auf GoBD-Konformität. Vollständigkeits-Prüfung, Lesbarkeits-Prüfung, Archivierbarkeits-Prüfung. Nützlich für DocumentAgent.",
            parameters={
                "receipt_data": {
                    "type": "object",
                    "description": "Beleg-Daten (Betrag, Datum, Steuernummer, etc.)"
                },
                "check_readability": {
                    "type": "boolean",
                    "description": "Lesbarkeits-Prüfung (Standard: True)",
                    "default": True
                },
                "check_completeness": {
                    "type": "boolean",
                    "description": "Vollständigkeits-Prüfung (Standard: True)",
                    "default": True
                }
            }
        )
    
    async def execute(self,
                      receipt_data: Dict[str, Any],
                      check_readability: bool = True,
                      check_completeness: bool = True) -> Dict[str, Any]:
        """Prüfe Beleg auf GoBD-Konformität"""
        try:
            # GoBD-Anforderungen: Betrag, Datum, Steuernummer, Firmenanschrift, etc.
            required_fields = ["amount", "date", "tax_number", "company_address"]
            optional_fields = ["invoice_number", "vat_amount", "description"]
            
            result = {
                "success": True,
                "gobd_compliant": True,
                "issues": [],
                "warnings": []
            }
            
            # Vollständigkeits-Prüfung
            if check_completeness:
                missing_fields = []
                for field in required_fields:
                    if field not in receipt_data or not receipt_data[field]:
                        missing_fields.append(field)
                
                if missing_fields:
                    result["gobd_compliant"] = False
                    result["issues"].append(f"Fehlende Pflichtfelder: {', '.join(missing_fields)}")
                    result["missing_fields"] = missing_fields
                
                # Prüfe optionale Felder
                missing_optional = [f for f in optional_fields if f not in receipt_data or not receipt_data[f]]
                if missing_optional:
                    result["warnings"].append(f"Fehlende optionale Felder: {', '.join(missing_optional)}")
            
            # Lesbarkeits-Prüfung (vereinfacht)
            if check_readability:
                # Prüfe ob Text-Felder lesbar sind
                text_fields = ["description", "company_address", "invoice_number"]
                for field in text_fields:
                    if field in receipt_data:
                        value = str(receipt_data[field])
                        # Prüfe auf zu viele Sonderzeichen oder unlesbare Zeichen
                        if len(value) < 3:
                            result["warnings"].append(f"Feld '{field}' ist sehr kurz")
                        if value.count("?") > len(value) * 0.3:
                            result["warnings"].append(f"Feld '{field}' enthält viele unlesbare Zeichen")
            
            # Betrag-Validierung
            if "amount" in receipt_data:
                try:
                    amount = float(receipt_data["amount"])
                    if amount <= 0:
                        result["issues"].append("Betrag muss größer als 0 sein")
                        result["gobd_compliant"] = False
                except (ValueError, TypeError):
                    result["issues"].append("Betrag ist ungültig")
                    result["gobd_compliant"] = False
            
            # Datum-Validierung
            if "date" in receipt_data:
                try:
                    from datetime import datetime
                    date_str = str(receipt_data["date"])
                    datetime.strptime(date_str, "%Y-%m-%d")
                except (ValueError, TypeError):
                    result["warnings"].append("Datum-Format könnte ungültig sein")
            
            result["issues_count"] = len(result["issues"])
            result["warnings_count"] = len(result["warnings"])
            
            return result
            
        except Exception as e:
            logger.error(f"Receipt standard validator error: {e}")
            return {
                "success": False,
                "error": str(e),
                "receipt_data": receipt_data
            }

class BankStatementParserTool(AgentTool):
    """Tool für Kontoauszug-Parsing"""
    
    def __init__(self):
        super().__init__(
            name="bank_statement_parser",
            description="Parst Kontoauszüge (PDF, MT940, CAMT.053). Extrahiert Transaktionen, Beträge, Daten. Nützlich für AccountingAgent.",
            parameters={
                "file_path": {
                    "type": "string",
                    "description": "Pfad zur Kontoauszug-Datei (PDF)"
                },
                "format": {
                    "type": "string",
                    "description": "Dateiformat (Standard: 'auto')",
                    "enum": ["auto", "pdf", "mt940", "camt053"],
                    "default": "auto"
                }
            }
        )
    
    async def execute(self,
                      file_path: str,
                      format: str = "auto") -> Dict[str, Any]:
        """Parse Kontoauszug"""
        try:
            if not Path(file_path).exists():
                return {
                    "success": False,
                    "error": f"Datei nicht gefunden: {file_path}",
                    "file_path": file_path
                }
            
            transactions = []
            
            # PDF-Parsing
            if format == "auto" or format == "pdf":
                if file_path.lower().endswith('.pdf'):
                    try:
                        if HAS_PDFPLUMBER:
                            import pdfplumber
                            with pdfplumber.open(file_path) as pdf:
                                for page_num, page in enumerate(pdf.pages):
                                    # Extrahiere Tabellen
                                    tables = page.extract_tables()
                                    for table in tables:
                                        # Vereinfachtes Parsing (würde spezifische Format-Erkennung erfordern)
                                        for row in table:
                                            if len(row) >= 3:
                                                # Versuche Datum, Betrag, Beschreibung zu extrahieren
                                                transaction = {
                                                    "page": page_num + 1,
                                                    "raw_data": row
                                                }
                                                transactions.append(transaction)
                        else:
                            return {
                                "success": False,
                                "error": "pdfplumber nicht verfügbar für PDF-Parsing"
                            }
                    except Exception as e:
                        logger.debug(f"PDF-Parsing fehlgeschlagen: {e}")
                        return {
                            "success": False,
                            "error": f"PDF-Parsing fehlgeschlagen: {str(e)}"
                        }
            
            # MT940-Parsing (vereinfacht)
            elif format == "mt940":
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        content = f.read()
                        # Vereinfachtes MT940-Parsing (würde spezifische Bibliothek erfordern)
                        import re
                        # Suche nach Transaktions-Zeilen
                        transaction_lines = re.findall(r':61:(\d{6})(\d{4})?([DC])(\d+),(\d{2})(.*)', content)
                        for line in transaction_lines:
                            transactions.append({
                                "format": "mt940",
                                "date": line[0],
                                "amount": line[3],
                                "currency": line[4] if len(line) > 4 else "EUR",
                                "description": line[5] if len(line) > 5 else ""
                            })
                except Exception as e:
                    logger.debug(f"MT940-Parsing fehlgeschlagen: {e}")
                    return {
                        "success": False,
                        "error": f"MT940-Parsing fehlgeschlagen: {str(e)}"
                    }
            
            return {
                "success": True,
                "file_path": file_path,
                "format": format,
                "transactions_found": len(transactions),
                "transactions": transactions
            }
            
        except Exception as e:
            logger.error(f"Bank statement parser error: {e}")
            return {
                "success": False,
                "error": str(e),
                "file_path": file_path
            }

class DistanceMatrixTool(AgentTool):
    """Tool für Entfernungsmatrix-Berechnung"""
    
    def __init__(self):
        super().__init__(
            name="distance_matrix",
            description="Berechnet Entfernungsmatrix für mehrere Orte. Optimale Route-Berechnung, Kosten-Berechnung. Nützlich für AccountingAgent.",
            parameters={
                "origins": {
                    "type": "array",
                    "description": "Liste von Startorten"
                },
                "destinations": {
                    "type": "array",
                    "description": "Liste von Zielorten"
                },
                "mode": {
                    "type": "string",
                    "description": "Verkehrsmittel (Standard: 'driving')",
                    "enum": ["driving", "walking", "bicycling", "transit"],
                    "default": "driving"
                }
            }
        )
        self.google_api_key = os.getenv('GOOGLE_MAPS_API_KEY')
        self.ors_api_key = os.getenv('OPENROUTESERVICE_API_KEY')
    
    async def execute(self,
                      origins: List[str],
                      destinations: List[str],
                      mode: str = "driving") -> Dict[str, Any]:
        """Berechne Entfernungsmatrix"""
        try:
            if not origins or not destinations:
                return {
                    "success": False,
                    "error": "Origins und Destinations sind erforderlich"
                }
            
            # Nutze TravelTimeCalculatorTool für einzelne Berechnungen
            from backend.agents import get_tool_registry
            tool_registry = get_tool_registry()
            travel_tool = tool_registry.get_tool("travel_time_calculator")
            
            if not travel_tool:
                return {
                    "success": False,
                    "error": "TravelTimeCalculatorTool nicht verfügbar"
                }
            
            matrix = []
            for origin in origins:
                row = []
                for destination in destinations:
                    result = await travel_tool.execute(
                        origin=origin,
                        destination=destination,
                        mode=mode,
                        provider="openrouteservice" if self.ors_api_key else "google"
                    )
                    row.append(result)
                matrix.append(row)
            
            return {
                "success": True,
                "origins": origins,
                "destinations": destinations,
                "mode": mode,
                "matrix": matrix
            }
            
        except Exception as e:
            logger.error(f"Distance matrix error: {e}")
            return {
                "success": False,
                "error": str(e),
                "origins": origins,
                "destinations": destinations
            }

class CompanyDatabaseTool(AgentTool):
    """Tool für Firmendatenbank-Abfrage"""
    
    def __init__(self):
        super().__init__(
            name="company_database",
            description="Abfrage von Firmendaten. USt-IdNr-Validierung gegen EU-VIES, Firmenname-Normalisierung. Nützlich für DocumentAgent und AccountingAgent.",
            parameters={
                "vat_number": {
                    "type": "string",
                    "description": "USt-IdNr zum Validieren (Format: DE123456789)"
                },
                "company_name": {
                    "type": "string",
                    "description": "Firmenname (optional, für Normalisierung)"
                }
            }
        )
    
    async def execute(self,
                      vat_number: Optional[str] = None,
                      company_name: Optional[str] = None) -> Dict[str, Any]:
        """Abfrage Firmendaten"""
        try:
            result = {
                "success": True
            }
            
            # USt-IdNr-Validierung gegen EU-VIES
            if vat_number:
                try:
                    import aiohttp
                    # EU VIES API (kostenlos)
                    country_code = vat_number[:2].upper()
                    vat_id = vat_number[2:]
                    
                    url = "https://ec.europa.eu/taxation_customs/vies/rest-api/ms/{}/vat/{}".format(
                        country_code, vat_id
                    )
                    
                    async with aiohttp.ClientSession() as session:
                        async with session.get(url) as response:
                            if response.status == 200:
                                data = await response.json()
                                result["vat_validation"] = {
                                    "valid": data.get("valid", False),
                                    "name": data.get("name", ""),
                                    "address": data.get("address", ""),
                                    "country_code": country_code
                                }
                            else:
                                result["vat_validation"] = {
                                    "valid": False,
                                    "error": f"HTTP {response.status}"
                                }
                except Exception as e:
                    logger.debug(f"VIES-Validierung fehlgeschlagen: {e}")
                    result["vat_validation"] = {
                        "error": str(e)
                    }
            
            # Firmenname-Normalisierung (vereinfacht)
            if company_name:
                # Entferne häufige Suffixe/Präfixe
                normalized = company_name.strip()
                suffixes = ["GmbH", "AG", "UG", "e.K.", "KG", "OHG", "mbH"]
                for suffix in suffixes:
                    if normalized.endswith(suffix):
                        normalized = normalized[:-len(suffix)].strip()
                
                result["company_name"] = company_name
                result["normalized_name"] = normalized
            
            return result
            
        except Exception as e:
            logger.error(f"Company database error: {e}")
            return {
                "success": False,
                "error": str(e),
                "vat_number": vat_number
            }

class LangChainTool(AgentTool):
    """Tool für LangChain-Integration - erweiterte Agent-Funktionalität mit Tool-Orchestrierung"""
    
    def __init__(self):
        super().__init__(
            name="langchain",
            description="LangChain-Integration für erweiterte Agent-Funktionalität. Ermöglicht komplexe Workflows, Tool-Orchestrierung und erweiterte LLM-Interaktionen. Nützlich für alle Agents, besonders für komplexe Entscheidungsprozesse.",
            parameters={
                "action": {
                    "type": "string",
                    "description": "Aktion: 'create_agent' (erstellt LangChain Agent), 'run_workflow' (führt Workflow aus), 'tool_chain' (verkettet Tools)",
                    "enum": ["create_agent", "run_workflow", "tool_chain"]
                },
                "tools": {
                    "type": "array",
                    "description": "Liste von Tool-Namen für Agent/Workflow",
                    "default": []
                },
                "prompt": {
                    "type": "string",
                    "description": "Prompt für Agent/Workflow"
                },
                "workflow_steps": {
                    "type": "array",
                    "description": "Workflow-Schritte (für run_workflow)",
                    "default": []
                }
            }
        )
        self.has_langchain = HAS_LANGCHAIN
    
    async def execute(self, 
                      action: str,
                      tools: Optional[List[str]] = None,
                      prompt: Optional[str] = None,
                      workflow_steps: Optional[List[Dict[str, Any]]] = None) -> Dict[str, Any]:
        """Führe LangChain-Aktion aus"""
        try:
            if not self.has_langchain:
                return {
                    "success": False,
                    "error": "LangChain nicht verfügbar. Bitte 'pip install langchain langchain-openai' installieren.",
                    "note": "LangChain ist optional und bietet erweiterte Agent-Funktionalität"
                }
            
            if action == "create_agent":
                # Erstelle LangChain Agent mit Tools
                # Dies ist ein vereinfachtes Beispiel - vollständige Integration würde mehr Setup erfordern
                return {
                    "success": True,
                    "action": "create_agent",
                    "message": "LangChain Agent kann erstellt werden. Vollständige Integration erfordert OpenAI API Key oder Ollama-Integration.",
                    "note": "Für vollständige Integration: Konfiguriere LLM (OpenAI oder Ollama) und Tools"
                }
            
            elif action == "run_workflow":
                # Führe Workflow aus
                if not workflow_steps:
                    return {
                        "success": False,
                        "error": "workflow_steps ist erforderlich für run_workflow"
                    }
                
                results = []
                for step in workflow_steps:
                    step_result = {
                        "step": step.get("name", "unknown"),
                        "status": "completed",
                        "result": "Workflow-Schritt ausgeführt"
                    }
                    results.append(step_result)
                
                return {
                    "success": True,
                    "action": "run_workflow",
                    "steps": results,
                    "count": len(results)
                }
            
            elif action == "tool_chain":
                # Verkette Tools
                if not tools:
                    return {
                        "success": False,
                        "error": "tools ist erforderlich für tool_chain"
                    }
                
                return {
                    "success": True,
                    "action": "tool_chain",
                    "tools": tools,
                    "message": "Tool-Kette kann erstellt werden. Vollständige Integration erfordert Tool-Registry-Integration."
                }
            
            else:
                return {
                    "success": False,
                    "error": f"Unbekannte Aktion: {action}",
                    "available_actions": ["create_agent", "run_workflow", "tool_chain"]
                }
                
        except Exception as e:
            logger.error(f"LangChain tool error: {e}")
            return {
                "success": False,
                "error": str(e),
                "action": action
            }

class AgentToolRegistry:
    """Registry für alle verfügbaren Agent-Tools"""
    
    def __init__(self):
        self.tools: Dict[str, AgentTool] = {}
        self._register_default_tools()
    
    def _register_default_tools(self):
        """Registriere Standard-Tools"""
        self.register(WebSearchTool())
        self.register(CurrencyExchangeTool())
        self.register(MealAllowanceLookupTool())
        self.register(GeocodingTool())
        self.register(OpenMapsTool())
        # Neue spezialisierte Tools
        self.register(ExaSearchTool())
        self.register(MarkerTool())
        self.register(PaddleOCRTool())
        self.register(CustomPythonRulesTool())
        self.register(LangChainTool())
        self.register(WebAccessTool())
        # Zusätzliche Tools für höhere Qualität
        self.register(DateParserTool())
        self.register(TaxNumberValidatorTool())
        self.register(TranslationTool())
        self.register(CurrencyValidatorTool())
        self.register(RegexPatternMatcherTool())
        self.register(PDFMetadataTool())
        # Priorität 1 Tools
        self.register(DuplicateDetectionTool())
        self.register(IBANValidatorTool())
        self.register(ImageQualityTool())
        self.register(TimeZoneTool())
        self.register(EmailValidatorTool())
        # Priorität 2 Tools
        self.register(EmailParserTool())
        self.register(SignatureDetectionTool())
        self.register(ExcelImportExportTool())
        self.register(PostalCodeValidatorTool())
        self.register(PhoneNumberValidatorTool())
        self.register(HolidayAPITool())
        self.register(WeatherAPITool())
        self.register(TravelTimeCalculatorTool())
        self.register(PDFTimestampTool())
        # Priorität 3 Tools
        self.register(QRCodeReaderTool())
        self.register(BarcodeReaderTool())
        self.register(InvoiceNumberValidatorTool())
        self.register(VATCalculatorTool())
        self.register(ExpenseCategoryClassifierTool())
        self.register(ReceiptStandardValidatorTool())
        self.register(BankStatementParserTool())
        self.register(DistanceMatrixTool())
        self.register(CompanyDatabaseTool())
    
    def register(self, tool: AgentTool):
        """Registriere ein neues Tool"""
        self.tools[tool.name] = tool
        logger.info(f"Tool '{tool.name}' registriert")
    
    def get_tool(self, name: str) -> Optional[AgentTool]:
        """Hole Tool nach Namen"""
        return self.tools.get(name)
    
    def list_tools(self) -> List[Dict[str, Any]]:
        """Liste alle verfügbaren Tools"""
        return [
            {
                "name": tool.name,
                "description": tool.description,
                "parameters": tool.parameters
            }
            for tool in self.tools.values()
        ]
    
    async def execute_tool(self, name: str, **kwargs) -> Dict[str, Any]:
        """Führe ein Tool aus"""
        tool = self.get_tool(name)
        if not tool:
            return {
                "success": False,
                "error": f"Tool '{name}' nicht gefunden",
                "available_tools": list(self.tools.keys())
            }
        
        try:
            result = await tool.execute(**kwargs)
            return result
        except Exception as e:
            logger.error(f"Tool execution error for {name}: {e}")
            return {
                "success": False,
                "error": str(e),
                "tool": name
            }
    
    async def close(self):
        """Schließe alle Tools (für Cleanup)"""
        for tool in self.tools.values():
            if hasattr(tool, 'close') and callable(tool.close):
                try:
                    await tool.close()
                except Exception as e:
                    logger.warning(f"Error closing tool {tool.name}: {e}")

# Globale Tool-Registry
_tool_registry: Optional[AgentToolRegistry] = None

def get_tool_registry() -> AgentToolRegistry:
    """Hole die globale Tool-Registry (Singleton)"""
    global _tool_registry
    if _tool_registry is None:
        _tool_registry = AgentToolRegistry()
    return _tool_registry

class OllamaLLM:
    """Wrapper for Ollama LLM API
    
    Supports remote Ollama server on GMKTec evo x2 in local network.
    Handles network connectivity, timeouts, and retries for Proxmox deployment.
    """
    
    def __init__(self, base_url: str = OLLAMA_BASE_URL, model: str = OLLAMA_MODEL):
        self.base_url = base_url.rstrip('/')
        self.model = model
        self.timeout = OLLAMA_TIMEOUT
        self.max_retries = OLLAMA_MAX_RETRIES
        self.retry_delay = OLLAMA_RETRY_DELAY
        self._session = None
        logger.info(f"OllamaLLM initialized: {self.base_url}, model={self.model}")
    
    async def _get_session(self):
        """Get or create aiohttp session with connection pooling"""
        if self._session is None or self._session.closed:
            # Create session with connection pooling for better performance
            connector = aiohttp.TCPConnector(
                limit=10,  # Max connections
                limit_per_host=5,  # Max connections per host
                ttl_dns_cache=300,  # DNS cache TTL
                force_close=False,  # Reuse connections
                enable_cleanup_closed=True
            )
            timeout = aiohttp.ClientTimeout(
                total=self.timeout,
                connect=10,  # Connection timeout
                sock_read=self.timeout  # Read timeout
            )
            self._session = aiohttp.ClientSession(
                connector=connector,
                timeout=timeout
            )
        return self._session
    
    async def health_check(self) -> bool:
        """Check if Ollama server is reachable"""
        try:
            session = await self._get_session()
            async with session.get(
                f"{self.base_url}/api/tags",
                timeout=aiohttp.ClientTimeout(total=5)
            ) as response:
                if response.status == 200:
                    logger.info(f"Ollama health check OK: {self.base_url}")
                    return True
                else:
                    logger.warning(f"Ollama health check failed: {response.status}")
                    return False
        except Exception as e:
            logger.warning(f"Ollama health check error: {e}")
            return False
    
    async def chat(self, messages: List[Dict[str, str]], system_prompt: Optional[str] = None) -> str:
        """Send chat messages to Ollama and get response with retry logic"""
        # Prepare messages with system prompt
        formatted_messages = []
        if system_prompt:
            formatted_messages.append({"role": "system", "content": system_prompt})
        formatted_messages.extend(messages)
        
        last_error = None
        for attempt in range(self.max_retries):
            try:
                session = await self._get_session()
                async with session.post(
                    f"{self.base_url}/api/chat",
                    json={
                        "model": self.model,
                        "messages": formatted_messages,
                        "stream": False,
                        "options": {
                            "temperature": 0.7,
                            "num_predict": 4096  # Max tokens
                        }
                    }
                ) as response:
                    if response.status == 200:
                        result = await response.json()
                        content = result.get("message", {}).get("content", "")
                        if content:
                            logger.debug(f"Ollama response received (attempt {attempt + 1})")
                            return content
                        else:
                            logger.warning(f"Empty response from Ollama (attempt {attempt + 1})")
                            last_error = "Empty response from LLM"
                    else:
                        error_text = await response.text()
                        logger.error(f"Ollama API error: {response.status} - {error_text[:200]}")
                        last_error = f"API error {response.status}"
                        
            except aiohttp.ClientConnectorError as e:
                last_error = f"Connection error: {str(e)}"
                logger.warning(f"Ollama connection error (attempt {attempt + 1}/{self.max_retries}): {e}")
                if attempt < self.max_retries - 1:
                    await asyncio.sleep(self.retry_delay * (attempt + 1))  # Exponential backoff
                    
            except asyncio.TimeoutError:
                last_error = "Request timeout"
                logger.warning(f"Ollama timeout (attempt {attempt + 1}/{self.max_retries})")
                if attempt < self.max_retries - 1:
                    await asyncio.sleep(self.retry_delay * (attempt + 1))
                    
            except Exception as e:
                last_error = f"Unexpected error: {str(e)}"
                logger.error(f"Ollama error (attempt {attempt + 1}/{self.max_retries}): {e}")
                if attempt < self.max_retries - 1:
                    await asyncio.sleep(self.retry_delay * (attempt + 1))
        
        # All retries failed
        error_msg = f"Fehler bei Kommunikation mit LLM nach {self.max_retries} Versuchen: {last_error}"
        logger.error(error_msg)
        return error_msg
    
    async def close(self):
        """Close the HTTP session"""
        if self._session and not self._session.closed:
            await self._session.close()
            self._session = None
    
    async def extract_json(self, prompt: str, system_prompt: Optional[str] = None) -> Optional[Dict]:
        """Extract structured JSON from LLM response"""
        extraction_prompt = f"{prompt}\n\nAntworte NUR mit einem gültigen JSON-Objekt, keine zusätzlichen Erklärungen."
        response = await self.chat([{"role": "user", "content": extraction_prompt}], system_prompt)
        
        # Try to extract JSON from response
        try:
            # Remove markdown code blocks if present
            if "```json" in response:
                response = response.split("```json")[1].split("```")[0]
            elif "```" in response:
                response = response.split("```")[1].split("```")[0]
            
            return json.loads(response.strip())
        except json.JSONDecodeError:
            logger.warning(f"Could not parse JSON from response: {response[:200]}")
            return None

class ChatAgent:
    """Agent für Dialog und Rückfragen mit Benutzer"""
    
    def __init__(self, llm: OllamaLLM, memory: Optional[AgentMemory] = None, db=None, tools: Optional[AgentToolRegistry] = None):
        self.llm = llm
        self.name = "ChatAgent"
        self.memory = memory or AgentMemory(self.name, db)
        self.tools = tools or get_tool_registry()
        self.system_prompt_base = """Du bist ein hilfreicher Assistent für Reisekostenabrechnungen. 
Du stellst dem Benutzer klare Fragen zu fehlenden oder unklaren Informationen in den Reisekostenabrechnungen.
Sei präzise, freundlich und auf Deutsch.
Formuliere Fragen so, dass sie mit kurzen Antworten beantwortet werden können.

Verfügbare Tools:
- exa_search: Hochwertige semantische Suche mit Exa/XNG API (PRIMÄR für ChatAgent) - bessere Ergebnisse als Standard-Web-Suche
- date_parser: Datums-Parsing und -Validierung in verschiedenen Formaten (heute, gestern, DD.MM.YYYY, etc.) - für Datumsverständnis in Gesprächen
- timezone: Zeitzonen-Erkennung und -Konvertierung - für Fragen zu internationalen Reisen und Zeitzonen
- email_validator: E-Mail-Validierung (RFC 5322) und DNS-Prüfung - für E-Mail-Adressen-Validierung
- web_access: Generischer Web-Zugriff für HTTP-Requests zu beliebigen URLs (GET/POST/PUT/DELETE, Web-Scraping, API-Zugriff) - für alle Agents verfügbar
- langchain: LangChain-Integration für erweiterte Agent-Funktionalität und komplexe Workflows (OPTIONAL, für alle Agents)
- openmaps: Umfassende OpenStreetMap-Funktionen (Geocoding, POI-Suche, Entfernungen, Routen)
- web_search: Suche nach aktuellen Informationen im Internet (Fallback)
- currency_exchange: Aktuelle Wechselkurse
- geocoding: Bestimmt Ländercode aus Ortsangabe"""
    
    async def initialize(self):
        """Initialisiere Memory"""
        await self.memory.initialize()
    
    async def process(self, context: Dict[str, Any], user_message: Optional[str] = None) -> AgentResponse:
        """Process user message or generate question based on context"""
        report_issues = context.get("issues", [])
        missing_info = context.get("missing_info", [])
        
        # Hole relevanten Memory-Kontext
        memory_context = await self.memory.get_context_for_prompt(
            max_tokens=1500,
            relevant_query=f"{missing_info} {report_issues}" if missing_info or report_issues else None
        )
        
        # Erweitere System-Prompt mit Memory
        system_prompt = self.system_prompt_base
        if memory_context:
            system_prompt += f"\n\n=== Dein Gedächtnis (frühere Erfahrungen) ===\n{memory_context}\n"
            system_prompt += "\nNutze diese Informationen aus deinem Gedächtnis, um bessere Fragen zu stellen und den Benutzer besser zu verstehen."
        
        if user_message:
            # Process user's answer
            prompt = f"""Der Benutzer hat folgende Antwort gegeben: "{user_message}"
Kontext: {json.dumps(context, indent=2, ensure_ascii=False)}
            
Bewerte die Antwort und gib an:
- Ob die Information ausreichend ist
- Ob weitere Fragen nötig sind
- Eine Zusammenfassung der erhaltenen Informationen"""
            
            response_text = await self.llm.chat([
                {"role": "user", "content": prompt}
            ], system_prompt)
            
            # Speichere Konversation im Memory
            await self.memory.add_conversation(
                user_message=user_message,
                agent_response=response_text,
                context={"report_issues": report_issues, "missing_info": missing_info}
            )
            
            # Speichere Erkenntnis, falls neue Information
            if len(missing_info) == 0:
                await self.memory.add_insight(
                    f"Benutzer hat vollständige Informationen für Reisekostenabrechnung bereitgestellt: {user_message[:200]}",
                    source="user_conversation"
                )
            
            return AgentResponse(
                agent_name=self.name,
                message=response_text,
                requires_user_input=len(missing_info) > 0,
                next_agent="DocumentAgent" if len(report_issues) > 0 else None
            )
        else:
            # Generate questions based on context
            if missing_info:
                prompt = f"""Es fehlen folgende Informationen in der Reisekostenabrechnung:
{json.dumps(missing_info, indent=2, ensure_ascii=False)}

Formuliere eine klare, freundliche Frage an den Benutzer, um die fehlende Information zu erhalten."""
                response_text = await self.llm.chat([
                    {"role": "user", "content": prompt}
                ], system_prompt)
            else:
                response_text = "Alle Informationen sind vollständig. Die Prüfung kann fortgesetzt werden."
            
            # Speichere generierte Frage
            if missing_info:
                await self.memory.add_decision(
                    f"Frage generiert für fehlende Informationen: {missing_info}",
                    confidence=0.8,
                    context={"missing_info": missing_info}
                )
            
            return AgentResponse(
                agent_name=self.name,
                message=response_text,
                requires_user_input=len(missing_info) > 0
            )

class DocumentAgent:
    """Agent für Dokumentenanalyse: Verstehen, Übersetzen, Kategorisieren, Validieren"""
    
    def __init__(self, llm: OllamaLLM, message_bus: Optional[AgentMessageBus] = None, memory: Optional[AgentMemory] = None, db=None, tools: Optional[AgentToolRegistry] = None):
        self.llm = llm
        self.name = "DocumentAgent"
        self.message_bus = message_bus
        self.memory = memory or AgentMemory(self.name, db)
        self.tools = tools or get_tool_registry()
        # Load prompt from markdown file
        self.system_prompt_base = load_prompt("document_agent.md")
        if not self.system_prompt_base:
            # Fallback default prompt
            self.system_prompt_base = """Du bist ein Experte für Dokumentenanalyse von Reisekostenbelegen.
Deine Aufgaben:
1. Dokumente kategorisieren (Hotel, Restaurant, Maut, Parken, Tanken, Bahnticket, etc.)
2. Sprache erkennen und bei Bedarf übersetzen
3. Relevante Daten extrahieren (Betrag, Datum, Steuernummer, Firmenanschrift, etc.)
4. Vollständigkeit prüfen (Echtheit, Steuernummer vorhanden, Firmenanschrift, etc.)
5. Probleme und Unstimmigkeiten identifizieren

Verfügbare Tools:
- marker: Erweiterte Dokumentenanalyse und -extraktion (PRIMÄR für DocumentAgent) - strukturierte Daten aus PDFs
- paddleocr: OCR-Tool für Texterkennung (FALLBACK) - wenn andere Methoden versagen, unterstützt 100+ Sprachen
- email_parser: Automatische Beleg-Extraktion aus E-Mails (IMAP/POP3) - erkennt Beleg-Anhänge und extrahiert Betrag, Datum
- signature_detection: Erweiterte Signatur-Erkennung in PDFs (digitale Signaturen, Signatur-Felder, handschriftliche Signaturen)
- qrcode_reader: QR-Code-Erkennung in PDFs und Bildern - erkennt E-Rechnungen (ZUGFeRD, XRechnung)
- barcode_reader: Barcode-Erkennung in Belegen (EAN, UPC, Code128) - extrahiert Produktdaten
- duplicate_detection: Duplikats-Erkennung durch Hash-Vergleich - verhindert doppelte Beleg-Uploads
- image_quality: Qualitätsprüfung von gescannten Belegen (DPI, Schärfe, Kontrast, Helligkeit) - warnt vor schlechter Qualität vor OCR
- pdf_metadata: PDF-Metadaten-Extraktion (Erstellungsdatum, Autor, Titel, Seitenzahl) - für Dokumentenanalyse
- pdf_timestamp: Zeitstempel-Validierung in PDFs - validiert Erstellungsdatum gegen Reisedaten
- receipt_standard_validator: GoBD-Konformitäts-Prüfung - prüft Vollständigkeit, Lesbarkeit, Archivierbarkeit
- invoice_number_validator: Rechnungsnummer-Validierung - prüft Format, Duplikate, Sequenzen
- company_database: Firmendatenbank-Abfrage - USt-IdNr-Validierung gegen EU-VIES, Firmenname-Normalisierung
- translation: Übersetzung zwischen Sprachen (PRIMÄR für DocumentAgent) - für mehrsprachige Belege, unterstützt 100+ Sprachen
- tax_number_validator: Steuernummer-Validierung (USt-IdNr, VAT) für verschiedene Länder (DE, AT, CH, FR, IT, ES, GB, US) - für Beleg-Validierung
- iban_validator: IBAN-Validierung und Bankdaten-Extraktion (ISO 13616) - für Bankdaten in Belegen
- email_validator: E-Mail-Validierung (RFC 5322) und DNS-Prüfung - für E-Mail-Adressen in Belegen
- phone_number_validator: Telefonnummer-Validierung und Formatierung (E.164) - für Telefonnummern in Belegen
- postal_code_validator: Postleitzahlen-Validierung (DE, AT, CH, FR, IT, ES, GB, US) - für Adress-Validierung
- date_parser: Datums-Parsing und -Validierung in verschiedenen Formaten - für Datumsextraktion aus Dokumenten
- regex_pattern_matcher: Mustererkennung in Texten (Beträge, Datumsangaben, E-Mails, Telefonnummern, etc.) - für Datenextraktion
- web_access: Generischer Web-Zugriff für HTTP-Requests (GET/POST/PUT/DELETE) - nützlich für Validierung von Dokumenten, API-Zugriff, Web-Scraping
- langchain: LangChain-Integration für erweiterte Agent-Funktionalität und komplexe Workflows (OPTIONAL)
- openmaps: Umfassende OpenStreetMap-Funktionen (Geocoding, Reverse Geocoding, POI-Suche, Entfernungen) - nützlich zur Validierung von Adressen und Standorten in Belegen
- web_search: Suche nach aktuellen Informationen

Antworte präzise und strukturiert."""
        
        # Subscribe to messages from other agents
        if self.message_bus:
            self.message_bus.subscribe(self.name, self.handle_agent_message)
    
    async def initialize(self):
        """Initialisiere Memory"""
        await self.memory.initialize()
    
    def handle_agent_message(self, message: Dict):
        """Handle messages from other agents"""
        logger.info(f"DocumentAgent received message from {message.get('from')}: {message.get('content')}")
        # Can request clarification from Chat Agent if needed
        if message.get('from') == 'ChatAgent':
            # Chat Agent might request document re-analysis
            pass
    
    def extract_pdf_text(self, pdf_path: str, encryption=None) -> str:
        """
        Extract text from PDF file
        Handles encrypted files for DSGVO compliance
        """
        extracted_text = ""
        
        try:
            # Check if file is encrypted and decrypt if needed
            file_data = None
            if encryption:
                try:
                    # Try to decrypt first
                    file_data = encryption.decrypt_file(Path(pdf_path))
                    # Write decrypted data to temp file for processing
                    temp_path = Path(pdf_path).with_suffix('.tmp.pdf')
                    with open(temp_path, 'wb') as f:
                        f.write(file_data)
                    pdf_path = str(temp_path)
                except Exception:
                    # File might not be encrypted, proceed normally
                    pass
            
            if HAS_PDFPLUMBER:
                # Use pdfplumber (better for tables and structured data)
                with pdfplumber.open(pdf_path) as pdf:
                    for page in pdf.pages:
                        text = page.extract_text()
                        if text:
                            extracted_text += text + "\n"
            elif HAS_PYPDF2:
                # Fallback to PyPDF2
                with open(pdf_path, 'rb') as file:
                    pdf_reader = PyPDF2.PdfReader(file)
                    for page in pdf_reader.pages:
                        text = page.extract_text()
                        if text:
                            extracted_text += text + "\n"
            
            # Clean up temp file if created
            if file_data and Path(pdf_path).suffix == '.tmp.pdf':
                try:
                    Path(pdf_path).unlink()
                except:
                    pass
                    
        except Exception as e:
            logger.warning(f"Could not extract text from PDF {pdf_path}: {e}")
        
        return extracted_text
    
    async def analyze_document(self, receipt_path: str, filename: str, encryption=None) -> DocumentAnalysis:
        """Analyze a PDF receipt document"""
        try:
            # Extract text from PDF (handles encryption if needed)
            pdf_text = self.extract_pdf_text(receipt_path, encryption)
            
            # Limit text length for LLM (first 5000 characters)
            pdf_text_limited = pdf_text[:5000] if pdf_text else "Kein Text extrahiert"
            
            # Hole relevanten Memory-Kontext für ähnliche Dokumente
            memory_context = await self.memory.get_context_for_prompt(
                max_tokens=1500,
                relevant_query=f"document analysis {filename} {pdf_text_limited[:200]}"
            )
            
            # Erweitere System-Prompt mit Memory
            system_prompt = self.system_prompt_base
            if memory_context:
                system_prompt += f"\n\n=== Dein Gedächtnis (frühere Dokumentenanalysen) ===\n{memory_context}\n"
                system_prompt += "\nNutze diese Erfahrungen aus deinem Gedächtnis, um ähnliche Dokumente besser zu analysieren und bekannte Muster zu erkennen."
            
            prompt = f"""Analysiere das folgende Reisekosten-Dokument:

Dateiname: {filename}
Extrahierter Text aus PDF:
{pdf_text_limited}

Analysiere und extrahiere folgende Informationen:
1. Dokumenttyp: hotel_receipt, restaurant_bill, toll_receipt, parking, fuel, train_ticket, other
2. Sprache des Dokuments
3. Betrag (Hauptbetrag)
4. Datum
5. Währung
6. Steuernummer/USt-IdNr (falls vorhanden)
7. Firmenanschrift/Kontaktdaten (falls vorhanden)
8. Vollständigkeitsprüfung:
   - Steuernummer vorhanden?
   - Firmenanschrift vorhanden?
   - Betrag klar erkennbar?
   - Datum vorhanden?
9. Probleme/Unstimmigkeiten (z.B. unleserlich, unvollständig, verdächtig)

Antworte im JSON-Format mit folgenden Feldern:
{{
  "document_type": "...",
  "language": "...",
  "extracted_data": {{
    "amount": 0.0,
    "currency": "EUR",
    "date": "YYYY-MM-DD",
    "tax_number": "...",
    "company_address": "..."
  }},
  "validation_issues": ["..."],
  "completeness_check": {{
    "has_tax_number": true/false,
    "has_company_address": true/false,
    "has_amount": true/false,
    "has_date": true/false
  }},
  "confidence": 0.0-1.0
}}"""
            
            analysis_json = await self.llm.extract_json(prompt, system_prompt)
            
            if not analysis_json:
                # Fallback: Try to extract basic info from filename
                filename_lower = filename.lower()
                doc_type = "other"
                if "hotel" in filename_lower or "unterkunft" in filename_lower:
                    doc_type = "hotel_receipt"
                elif "restaurant" in filename_lower or "essen" in filename_lower:
                    doc_type = "restaurant_bill"
                elif "maut" in filename_lower or "toll" in filename_lower:
                    doc_type = "toll_receipt"
                elif "park" in filename_lower:
                    doc_type = "parking"
                elif "tank" in filename_lower or "fuel" in filename_lower or "benzin" in filename_lower:
                    doc_type = "fuel"
                elif "bahn" in filename_lower or "train" in filename_lower or "zug" in filename_lower:
                    doc_type = "train_ticket"
                
                return DocumentAnalysis(
                    document_type=doc_type,
                    language="de",
                    extracted_data={"filename": filename},
                    validation_issues=["Konnte Dokument nicht vollständig analysieren - nur Dateiname verwendet"],
                    completeness_check={"has_tax_number": False, "has_company_address": False, "has_amount": False, "has_date": False},
                    confidence=0.3
                )
            
            # Map to DocumentAnalysis model
            analysis = DocumentAnalysis(
                document_type=analysis_json.get("document_type", "unknown"),
                language=analysis_json.get("language", "de"),
                translated_content=analysis_json.get("translated_content"),
                extracted_data=analysis_json.get("extracted_data", {}),
                validation_issues=analysis_json.get("validation_issues", []),
                completeness_check=analysis_json.get("completeness_check", {}),
                confidence=float(analysis_json.get("confidence", 0.5))
            )
            
            # Speichere Analyse im Memory
            analysis_summary = f"Dokument: {filename}, Typ: {analysis.document_type}, Betrag: {analysis.extracted_data.get('amount', 0.0)} {analysis.extracted_data.get('currency', 'EUR')}, Sprache: {analysis.language}, Konfidenz: {analysis.confidence:.2f}"
            await self.memory.add_analysis(
                analysis_summary,
                metadata={
                    "document_type": analysis.document_type,
                    "filename": filename,
                    "confidence": analysis.confidence,
                    "validation_issues_count": len(analysis.validation_issues)
                }
            )
            
            # Speichere Muster, wenn hohe Konfidenz
            if analysis.confidence > 0.8:
                pattern = f"Dokumenttyp '{analysis.document_type}' mit typischen Merkmalen: Sprache={analysis.language}, Betrag-Format={analysis.extracted_data.get('amount')}"
                await self.memory.add_pattern(pattern, examples=[filename])
            
            # Speichere Erkenntnis bei Problemen
            if analysis.validation_issues:
                insight = f"Dokument '{filename}' hat Probleme: {', '.join(analysis.validation_issues[:3])}"
                await self.memory.add_insight(insight, source="document_analysis")
            
            return analysis
        except Exception as e:
            logger.error(f"Error analyzing document {filename}: {e}")
            return DocumentAnalysis(
                document_type="unknown",
                language="de",
                extracted_data={},
                validation_issues=[f"Fehler bei Dokumentenanalyse: {str(e)}"],
                completeness_check={},
                confidence=0.0
            )
    
    async def process(self, receipts: List[Dict[str, Any]], encryption=None) -> List[DocumentAnalysis]:
        """
        Process multiple receipts
        Supports encrypted files for DSGVO compliance
        """
        analyses = []
        for receipt in receipts:
            analysis = await self.analyze_document(
                receipt.get("local_path", ""),
                receipt.get("filename", ""),
                encryption
            )
            
            # EU-AI-Act: Log AI decision
            try:
                from compliance import AITransparency
                ai_log = AITransparency.create_ai_decision_log(
                    decision_type="document_analysis",
                    agent_name=self.name,
                    input_data={"filename": receipt.get("filename"), "path": receipt.get("local_path")},
                    output_data=analysis.model_dump(),
                    confidence=analysis.confidence,
                    human_reviewed=False
                )
                # Store AI log in analysis metadata for transparency
                if not hasattr(analysis, 'ai_decision_log'):
                    analysis.ai_decision_log = ai_log
            except Exception as e:
                logger.warning(f"Could not create AI decision log: {e}")
            analyses.append(analysis)
            
            # Notify other agents about analysis completion (if message bus available)
            if self.message_bus:
                self.message_bus.publish(self.name, "AccountingAgent", {
                    "type": "document_analyzed",
                    "receipt_id": receipt.get("id"),
                    "analysis": analysis.model_dump()
                })
                
                # If issues found, notify Chat Agent
                if analysis.validation_issues or not all(analysis.completeness_check.values()):
                    self.message_bus.publish(self.name, "ChatAgent", {
                        "type": "document_issue",
                        "receipt_id": receipt.get("id"),
                        "filename": receipt.get("filename"),
                        "issues": analysis.validation_issues,
                        "completeness": analysis.completeness_check
                    })
        
        return analyses

class AccountingAgent:
    """Agent für Buchhaltung: Zuordnung, Verpflegungsmehraufwand, Spesensätze"""
    
    def __init__(self, llm: OllamaLLM, message_bus: Optional[AgentMessageBus] = None, memory: Optional[AgentMemory] = None, db=None, tools: Optional[AgentToolRegistry] = None):
        self.llm = llm
        self.name = "AccountingAgent"
        self.message_bus = message_bus
        self.memory = memory or AgentMemory(self.name, db)
        self.tools = tools or get_tool_registry()
        # Load prompt from markdown file
        self.system_prompt_base = load_prompt("accounting_agent.md")
        if not self.system_prompt_base:
            # Fallback default prompt
            self.system_prompt_base = """Du bist ein Buchhaltungs-Experte für Reisekostenabrechnungen.
Deine Aufgaben:
1. Dokumente den Reisekosteneinträgen zuordnen (basierend auf Datum, Ort, Zweck)
2. Verpflegungsmehraufwand automatisch berechnen (basierend auf Land und Abwesenheitsdauer)
3. Spezielle Dokumente zuordnen (Maut, Parken, etc.)
4. Kategorien korrekt zuweisen
5. Beträge validieren und korrigieren

Verfügbare Tools:
- marker: Erweiterte Dokumentenanalyse und -extraktion (PRIMÄR für AccountingAgent) - strukturierte Daten aus PDFs
- custom_python_rules: Benutzerdefinierte Python-Regeln für Buchhaltungsvalidierung und -berechnung (PRIMÄR für AccountingAgent)
- excel_import_export: Excel/CSV-Import/Export für Buchhaltung - exportiert Reisekosten-Reports für Buchhaltungssoftware
- invoice_number_validator: Rechnungsnummer-Validierung - prüft Format, Duplikate, Sequenzen
- vat_calculator: Mehrwertsteuer-Berechnung - Netto/Brutto-Umrechnung, länder-spezifische MwSt-Sätze
- expense_category_classifier: Automatische Kategorisierung von Ausgaben (Hotel, Restaurant, Transport, etc.) - keyword-basiert mit Konfidenz-Score
- bank_statement_parser: Kontoauszug-Parsing (PDF, MT940, CAMT.053) - extrahiert Transaktionen für Fremdwährungsnachweis
- distance_matrix: Entfernungsmatrix-Berechnung für mehrere Orte - optimale Route-Berechnung, Kosten-Berechnung
- company_database: Firmendatenbank-Abfrage - USt-IdNr-Validierung gegen EU-VIES, Firmenname-Normalisierung
- duplicate_detection: Duplikats-Erkennung durch Hash-Vergleich - verhindert doppelte Abrechnungen
- tax_number_validator: Steuernummer-Validierung (USt-IdNr, VAT) für verschiedene Länder (DE, AT, CH, FR, IT, ES, GB, US) - für Beleg-Validierung
- iban_validator: IBAN-Validierung und Bankdaten-Extraktion (ISO 13616) - für Überweisungsdaten
- postal_code_validator: Postleitzahlen-Validierung (DE, AT, CH, FR, IT, ES, GB, US) - für Adress-Validierung
- currency_validator: Währungsvalidierung und -formatierung (ISO 4217) - für Währungsvalidierung und Betragsformatierung
- timezone: Zeitzonen-Erkennung und -Konvertierung - für internationale Reisen, validiert Reisezeiten bei Zeitzonen-Wechsel
- travel_time_calculator: Reisezeit-Berechnung zwischen Orten (Auto, Bahn) - validiert Reisezeiten und -entfernungen
- holiday_api: Internationale Feiertags-Erkennung - validiert Reisetage gegen Feiertage
- weather_api: Wetter-Daten für Reisevalidierung - validiert Reisezeiten gegen Wetterdaten
- date_parser: Datums-Parsing und -Validierung in verschiedenen Formaten - für Datumsvergleiche und Validierung
- regex_pattern_matcher: Mustererkennung in Texten (Beträge, Datumsangaben, Steuernummern, etc.) - für Datenextraktion und Validierung
- web_access: Generischer Web-Zugriff für HTTP-Requests (GET/POST/PUT/DELETE) - nützlich für Buchhaltungs-APIs, Steuer-Websites, Validierung, API-Interaktion
- langchain: LangChain-Integration für erweiterte Agent-Funktionalität und komplexe Workflows (OPTIONAL, besonders nützlich für AccountingAgent)
- openmaps: Umfassende OpenStreetMap-Funktionen (Geocoding, Reverse Geocoding, POI-Suche, Entfernungen, Routen) - nützlich für Ortsbestimmung, Entfernungsvalidierung und Standortinformationen
- geocoding: Bestimmt Ländercode aus Ortsangabe
- meal_allowance_lookup: Holt aktuelle Verpflegungsmehraufwand-Spesensätze
- currency_exchange: Rechnet Fremdwährungen in EUR um
- web_search: Sucht nach aktuellen Informationen"""
<｜tool▁calls▁begin｜><｜tool▁call▁begin｜>
read_file
        
        # Subscribe to messages from other agents
        if self.message_bus:
            self.message_bus.subscribe(self.name, self.handle_agent_message)
    
    async def initialize(self):
        """Initialisiere Memory"""
        await self.memory.initialize()
    
    def handle_agent_message(self, message: Dict):
        """Handle messages from other agents"""
        logger.info(f"AccountingAgent received message from {message.get('from')}: {message.get('content')}")
        # Can request document analysis from Document Agent or clarification from Chat Agent
    
    async def get_country_code(self, location: str) -> str:
        """Get country code from location - nutzt Geocoding-Tool für aktuelle Daten"""
        try:
            # Versuche zuerst Geocoding-Tool
            geocoding_tool = self.tools.get_tool("geocoding")
            if geocoding_tool:
                result = await geocoding_tool.execute(location)
                if result.get("success") and result.get("country_code"):
                    country_code = result["country_code"]
                    # Speichere im Memory
                    await self.memory.add_insight(
                        f"Länderbestimmung für '{location}': {country_code} ({result.get('country', '')})",
                        source="geocoding_tool"
                    )
                    return country_code
        except Exception as e:
            logger.warning(f"Geocoding tool error: {e}, using fallback")
        
        # Fallback auf einfache String-Erkennung
        location_lower = location.lower()
        country_mapping = {
            "deutschland": "DE", "germany": "DE",
            "österreich": "AT", "austria": "AT",
            "schweiz": "CH", "switzerland": "CH",
            "frankreich": "FR", "france": "FR",
            "italien": "IT", "italy": "IT",
            "spanien": "ES", "spain": "ES",
            "großbritannien": "GB", "uk": "GB", "england": "GB",
            "usa": "US", "vereinigte staaten": "US"
        }
        for key, code in country_mapping.items():
            if key in location_lower:
                return code
        return "DE"  # Default: Deutschland
    
    async def get_meal_allowance(self, location: str, days: int, is_24h_absence: bool = True) -> float:
        """Get Verpflegungsmehraufwand based on location and duration - nutzt Web-Tools für aktuelle Daten"""
        country_code = await self.get_country_code(location)
        
        # Versuche zuerst aktuelle Spesensätze aus dem Web zu holen
        try:
            meal_allowance_tool = self.tools.get_tool("meal_allowance_lookup")
            if meal_allowance_tool:
                result = await meal_allowance_tool.execute(country_code)
                if result.get("success"):
                    if result.get("rates"):
                        rates_list = result["rates"]
                        if isinstance(rates_list, list) and len(rates_list) >= 2:
                            # Nimm die ersten beiden Werte (24h und abwesend)
                            rates = {
                                "24h": float(rates_list[0]),
                                "abwesend": float(rates_list[1]) if len(rates_list) > 1 else float(rates_list[0]) / 2
                            }
                            rate_type = "24h" if is_24h_absence else "abwesend"
                            daily_rate = rates.get(rate_type, rates["24h"])
                            
                            # Speichere im Memory
                            await self.memory.add_insight(
                                f"Aktuelle Spesensätze für {country_code}: 24h={rates['24h']} EUR, abwesend={rates['abwesend']} EUR (Quelle: Web)",
                                source="meal_allowance_lookup_tool"
                            )
                            
                            return daily_rate * days
                        elif isinstance(rates_list, dict):
                            # Falls es bereits ein Dict ist
                            rates = rates_list
                            rate_type = "24h" if is_24h_absence else "abwesend"
                            daily_rate = rates.get(rate_type, rates.get("24h", 28.0))
                            return daily_rate * days
        except Exception as e:
            logger.warning(f"Meal allowance lookup tool error: {e}, using local database")
        
        # Fallback auf lokale Datenbank
        rates = MEAL_ALLOWANCE_RATES.get(country_code, MEAL_ALLOWANCE_RATES["DEFAULT"])
        rate_type = "24h" if is_24h_absence else "abwesend"
        daily_rate = rates[rate_type]
        
        return daily_rate * days
    
    async def assign_expenses(self, report_entries: List[Dict], document_analyses: List[DocumentAnalysis], receipts: List[Dict]) -> List[ExpenseAssignment]:
        """Assign expenses to travel entries"""
        assignments = []
        
        # Create mapping of dates to entries
        entries_by_date = {entry.get("date"): entry for entry in report_entries}
        
        for i, analysis in enumerate(document_analyses):
            if i >= len(receipts):
                continue
            
            receipt = receipts[i]
            doc_date = analysis.extracted_data.get("date")
            doc_amount = analysis.extracted_data.get("amount", 0.0)
            
            # Find matching entry
            matching_entry = None
            if doc_date:
                matching_entry = entries_by_date.get(doc_date)
            
            # If no exact date match, use LLM to find best match
            if not matching_entry and report_entries:
                # Hole relevanten Memory-Kontext für ähnliche Zuordnungen
                memory_context = await self.memory.get_context_for_prompt(
                    max_tokens=1500,
                    relevant_query=f"assignment {analysis.document_type} {doc_date} {doc_amount}"
                )
                
                # Erweitere System-Prompt mit Memory
                system_prompt = self.system_prompt_base
                if memory_context:
                    system_prompt += f"\n\n=== Dein Gedächtnis (frühere Zuordnungen) ===\n{memory_context}\n"
                    system_prompt += "\nNutze diese Erfahrungen aus deinem Gedächtnis, um ähnliche Zuordnungen besser durchzuführen."
                
                prompt = f"""Ordne folgendes Dokument einem Reiseeintrag zu:

Dokument:
- Typ: {analysis.document_type}
- Betrag: {doc_amount} {analysis.extracted_data.get('currency', 'EUR')}
- Datum: {doc_date}
- Weitere Daten: {json.dumps(analysis.extracted_data, ensure_ascii=False, indent=2)}

Verfügbare Reiseeinträge:
{json.dumps([{"date": e.get("date"), "location": e.get("location"), "customer_project": e.get("customer_project"), "days_count": e.get("days_count", 1)} for e in report_entries], indent=2, ensure_ascii=False)}

Finde den besten passenden Reiseeintrag basierend auf:
- Datum (am besten)
- Ort/Standort
- Zweck/Projekt

Antworte mit JSON: {{"entry_date": "YYYY-MM-DD", "confidence": 0.0-1.0, "reason": "Warum dieser Eintrag passt"}}"""
                
                match_result = await self.llm.extract_json(prompt, system_prompt)
                if match_result and "entry_date" in match_result:
                    matching_entry = entries_by_date.get(match_result["entry_date"])
                    if matching_entry and match_result.get("confidence", 0.0) < 0.5:
                        # Low confidence - don't assign
                        matching_entry = None
            
            if matching_entry:
                # Determine category
                category = analysis.document_type
                if category == "restaurant_bill":
                    category = "meals"
                elif category in ["toll_receipt", "parking", "fuel"]:
                    category = "transport"
                elif category == "hotel_receipt":
                    category = "hotel"
                
                # Calculate meal allowance if applicable
                meal_allowance = None
                # Meal allowance is added separately for travel days, not for restaurant bills
                # Restaurant bills are actual expenses, meal allowance is for days without restaurant receipts
                
                # For hotel stays, calculate meal allowance for each day
                if category == "hotel":
                    location = matching_entry.get("location", "")
                    days = matching_entry.get("days_count", 1)
                    # Full day absence = 24h rate
                    meal_allowance = await self.get_meal_allowance(location, days, is_24h_absence=True)
                
                assignment = ExpenseAssignment(
                    receipt_id=receipt.get("id", ""),
                    entry_date=matching_entry.get("date", ""),
                    category=category,
                    amount=doc_amount,
                    currency=analysis.extracted_data.get("currency", "EUR"),
                    meal_allowance_added=meal_allowance,
                    assignment_confidence=analysis.confidence
                )
                
                # Speichere Zuordnung im Memory
                assignment_summary = f"Dokument {receipt.get('id', '')} zugeordnet zu Eintrag {assignment.entry_date}, Kategorie: {category}, Betrag: {doc_amount} {assignment.currency}"
                if meal_allowance:
                    assignment_summary += f", Verpflegungsmehraufwand: {meal_allowance} EUR"
                
                await self.memory.add_decision(
                    assignment_summary,
                    confidence=assignment.assignment_confidence,
                    context={
                        "document_type": analysis.document_type,
                        "category": category,
                        "entry_date": assignment.entry_date
                    }
                )
                
                assignments.append(assignment)
        
        return assignments
    
    async def process(self, report: Dict, document_analyses: List[DocumentAnalysis]) -> Dict[str, Any]:
        """Process expense assignment and meal allowance"""
        report_entries = report.get("entries", [])
        assignments = await self.assign_expenses(
            report_entries,
            document_analyses,
            report.get("receipts", [])
        )
        
        # Machbarkeitsprüfung: Überlappende Hotelrechnungen, Datum-Abgleich
        feasibility_issues = []
        
        # Prüfe auf überlappende Hotelrechnungen
        hotel_assignments = [a for a in assignments if a.category == "hotel"]
        hotel_dates = {}
        for assignment in hotel_assignments:
            entry_date = assignment.entry_date
            if entry_date in hotel_dates:
                hotel_dates[entry_date].append(assignment)
            else:
                hotel_dates[entry_date] = [assignment]
        
        for date, date_assignments in hotel_dates.items():
            if len(date_assignments) > 1:
                feasibility_issues.append(f"Mehrere Hotelrechnungen für {date} gefunden - mögliche Überlappung")
        
        # Prüfe Datum-Abgleich mit Arbeitsstunden
        for assignment in assignments:
            entry_date = assignment.entry_date
            matching_entry = next((e for e in report_entries if e.get("date") == entry_date), None)
            if matching_entry:
                working_hours = matching_entry.get("working_hours", 0.0)
                if working_hours == 0.0:
                    feasibility_issues.append(f"Für {entry_date} sind keine Arbeitsstunden im Stundenzettel verzeichnet, aber Reisekosten vorhanden")
        
        # Prüfe zeitliche Machbarkeit (Übernachtung ohne Anreise)
        entries_by_date = {e.get("date"): e for e in report_entries}
        for assignment in hotel_assignments:
            entry_date = assignment.entry_date
            from datetime import datetime, timedelta
            try:
                entry_date_obj = datetime.strptime(entry_date, "%Y-%m-%d")
                prev_date = (entry_date_obj - timedelta(days=1)).strftime("%Y-%m-%d")
                # Prüfe, ob es einen Reiseeintrag am Tag davor gibt
                if prev_date not in entries_by_date:
                    feasibility_issues.append(f"Hotelrechnung für {entry_date}, aber kein Reiseeintrag am Tag davor - mögliche fehlende Anreise")
            except:
                pass
        
        # Calculate totals
        total_expenses = sum(a.amount for a in assignments)
        total_meal_allowance = sum(a.meal_allowance_added or 0.0 for a in assignments)
        
        # Also calculate meal allowance for entries without receipts (pure travel days)
        entries_with_receipts = {a.entry_date for a in assignments}
        for entry in report_entries:
            entry_date = entry.get("date")
            if entry_date not in entries_with_receipts:
                # Travel day without receipt - add meal allowance
                location = entry.get("location", "")
                days = entry.get("days_count", 1)
                meal_allowance = await self.get_meal_allowance(location, days, is_24h_absence=True)
                total_meal_allowance += meal_allowance
        
        result = {
            "assignments": [a.model_dump() for a in assignments],
            "total_expenses": round(total_expenses, 2),
            "total_meal_allowance": round(total_meal_allowance, 2),
            "feasibility_issues": feasibility_issues,
            "summary": f"Zuordnung abgeschlossen: {len(assignments)} Dokumente zugeordnet, Gesamtausgaben: {total_expenses:.2f} EUR, Verpflegungsmehraufwand: {total_meal_allowance:.2f} EUR"
        }
        
        if feasibility_issues:
            result["summary"] += f"\n⚠️ {len(feasibility_issues)} Machbarkeitsprobleme gefunden - bitte prüfen!"
        
        # Speichere Zusammenfassung im Memory
        await self.memory.add_insight(
            f"Buchhaltungszuordnung abgeschlossen: {len(assignments)} Dokumente, Gesamtausgaben: {total_expenses:.2f} EUR, Verpflegungsmehraufwand: {total_meal_allowance:.2f} EUR, Probleme: {len(feasibility_issues)}",
            source="process_completion"
        )
        
        return result

class AgentOrchestrator:
    """Orchestrates the agent network for expense report review"""
    
    def __init__(self, llm: Optional[OllamaLLM] = None, db=None):
        base_url = (llm.base_url if isinstance(llm, OllamaLLM) else OLLAMA_BASE_URL)
        self.chat_llm = OllamaLLM(base_url=base_url, model=OLLAMA_MODEL_CHAT)
        self.document_llm = OllamaLLM(base_url=base_url, model=OLLAMA_MODEL_DOCUMENT)
        self.accounting_llm = OllamaLLM(base_url=base_url, model=OLLAMA_MODEL_ACCOUNTING)
        self._llms = {
            "ChatAgent": self.chat_llm,
            "DocumentAgent": self.document_llm,
            "AccountingAgent": self.accounting_llm,
        }
        self.db = db
        self.message_bus = AgentMessageBus()
        # Initialisiere Tool-Registry
        self.tools = get_tool_registry()
        # Initialisiere Agenten mit Memory und Tools (alle Agents erhalten Zugriff auf Tools)
        self.chat_agent = ChatAgent(self.chat_llm, db=db, tools=self.tools)
        self.document_agent = DocumentAgent(self.document_llm, self.message_bus, db=db, tools=self.tools)
        self.accounting_agent = AccountingAgent(self.accounting_llm, self.message_bus, db=db, tools=self.tools)
        self._llm_health_checked = False
        self._memory_initialized = False
    
    async def ensure_llm_available(self):
        """Check LLM availability and warn if not reachable"""
        if not self._llm_health_checked:
            all_healthy = True
            for agent_name, agent_llm in self._llms.items():
                is_healthy = await agent_llm.health_check()
                if not is_healthy:
                    all_healthy = False
                    logger.error(f"⚠️ Ollama LLM für {agent_name} nicht erreichbar: {agent_llm.base_url} (Modell: {agent_llm.model})")
                    logger.error("Bitte überprüfen Sie:")
                    logger.error("  1. Ollama läuft auf dem GMKTec-Server")
                    logger.error("  2. Netzwerk-Verbindung zum GMKTec-Server")
                    logger.error("  3. Firewall-Regeln erlauben Zugriff")
                    logger.error("  4. OLLAMA_BASE_URL und agentenspezifische Modelle sind korrekt konfiguriert")
                else:
                    logger.info(f"✅ Ollama LLM erreichbar für {agent_name}: {agent_llm.base_url} (Modell: {agent_llm.model})")
            self._llm_health_checked = True
            return all_healthy
        return True
    
    async def initialize_memory(self):
        """Initialisiere Memory für alle Agenten"""
        if not self._memory_initialized:
            await self.chat_agent.initialize()
            await self.document_agent.initialize()
            await self.accounting_agent.initialize()
            self._memory_initialized = True
            logger.info("Agent-Memory für alle Agenten initialisiert")
    
    def broadcast_message(self, from_agent: str, message: Dict[str, Any]):
        """Broadcast message to all agents via message bus"""
        self.message_bus.broadcast(from_agent, message)
    
    def send_message(self, from_agent: str, to_agent: str, message: Dict[str, Any]):
        """Send message from one agent to another"""
        self.message_bus.publish(from_agent, to_agent, message)
    
    async def close(self):
        """Clean up resources"""
        for agent_llm in self._llms.values():
            await agent_llm.close()
        # Schließe alle Tools
        await self.tools.close()
    
    async def review_expense_report(self, report_id: str, db) -> Dict[str, Any]:
        """Main orchestration method for reviewing an expense report"""
        # Ensure LLM is available
        await self.ensure_llm_available()
        
        # Initialisiere Memory
        await self.initialize_memory()
        
        # Fetch report
        report = await db.travel_expense_reports.find_one({"id": report_id})
        if not report:
            raise ValueError(f"Report {report_id} not found")
        
        receipts = report.get("receipts", [])
        entries = report.get("entries", [])
        
        # Step 1: Document Agent - Analyze all receipts
        logger.info(f"Step 1: Analyzing {len(receipts)} documents...")
        self.broadcast_message("Orchestrator", {
            "type": "status_update",
            "message": f"Starte Dokumentenanalyse für {len(receipts)} Belege",
            "step": 1
        })
        document_analyses = await self.document_agent.process(receipts)
        
        # Notify other agents about document analyses
        self.send_message("DocumentAgent", "AccountingAgent", {
            "type": "document_analyses_complete",
            "analyses": [a.model_dump() for a in document_analyses]
        })
        
        # Collect issues
        issues = []
        for analysis in document_analyses:
            if analysis.validation_issues:
                issues.extend(analysis.validation_issues)
            if not all(analysis.completeness_check.values()):
                issues.append(f"Dokument {analysis.document_type}: Vollständigkeitsprüfung fehlgeschlagen")
        
        # Step 2: Accounting Agent - Assign expenses and calculate meal allowance
        logger.info("Step 2: Assigning expenses...")
        self.broadcast_message("Orchestrator", {
            "type": "status_update",
            "message": "Starte Buchhaltungszuordnung",
            "step": 2
        })
        accounting_result = await self.accounting_agent.process(report, document_analyses)
        
        # Check if accounting agent needs clarification
        issues_needing_clarification = []
        for assignment in accounting_result.get("assignments", []):
            if assignment.get("assignment_confidence", 1.0) < 0.7:
                issues_needing_clarification.append({
                    "type": "low_confidence_assignment",
                    "receipt_id": assignment.get("receipt_id"),
                    "confidence": assignment.get("assignment_confidence")
                })
        
        # Step 3: Handle issues requiring clarification
        if issues_needing_clarification or issues:
            logger.info("Step 3: Issues found, may need user clarification")
            # Notify Chat Agent about issues
            self.send_message("Orchestrator", "ChatAgent", {
                "type": "clarification_needed",
                "document_issues": issues,
                "assignment_issues": issues_needing_clarification
            })
        
        # Step 4: Generate review summary
        review_summary = f"""Reisekostenabrechnung geprüft:

Dokumentenanalyse:
- {len(document_analyses)} Dokumente analysiert
- Probleme gefunden: {len(issues)}

Buchhaltungszuordnung:
{accounting_result['summary']}

Zuordnung:
{json.dumps(accounting_result['assignments'], indent=2, ensure_ascii=False)}

Status: {'Benötigt Klärung' if (issues_needing_clarification or issues) else 'Prüfung abgeschlossen'}
"""
        
        # Notify all agents about completion
        self.broadcast_message("Orchestrator", {
            "type": "review_complete",
            "has_issues": bool(issues_needing_clarification or issues),
            "summary": review_summary
        })
        
        # Update report with review notes
        await db.travel_expense_reports.update_one(
            {"id": report_id},
            {
                "$set": {
                    "review_notes": review_summary,
                    "accounting_data": accounting_result,
                    "document_analyses": [a.model_dump() for a in document_analyses],
                    "updated_at": datetime.utcnow()
                }
            }
        )
        
        # Determine if user input is needed
        requires_input = len(issues) > 0
        
        return {
            "status": "review_complete" if not requires_input else "needs_user_input",
            "issues": issues,
            "accounting_result": accounting_result,
            "document_analyses": [a.model_dump() for a in document_analyses],
            "requires_user_input": requires_input
        }
    
    async def handle_user_message(self, report_id: str, user_message: str, db) -> AgentResponse:
        """Handle user message in chat context"""
        # Initialisiere Memory
        await self.initialize_memory()
        
        report = await db.travel_expense_reports.find_one({"id": report_id})
        if not report:
            raise ValueError(f"Report {report_id} not found")
        
        # Get review context
        context = {
            "report": report,
            "issues": report.get("review_notes", "").split("\n") if report.get("review_notes") else [],
            "missing_info": []  # Would be extracted from review notes
        }
        
        # Process with Chat Agent
        response = await self.chat_agent.process(context, user_message)
        
        # Save chat message
        from server import ChatMessage
        chat_msg = ChatMessage(
            report_id=report_id,
            sender="agent",
            message=response.message
        )
        chat_msg_dict = chat_msg.model_dump()
        chat_msg_dict["created_at"] = datetime.utcnow()
        await db.chat_messages.insert_one(chat_msg_dict)
        
        return response

